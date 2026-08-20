"""
Módulo Compras (Estoque) — blueprint do JOGA Analytics.

Montado sob /estoque pelo server.py. Autenticação, sessão e RBAC vêm do app principal:
este módulo NÃO tem login próprio — a senha única (ESTOQUE_SENHA) foi removida na fusão,
junto do /login, /logout e do before_request que a guardava.

Consome os datasets Power BI "Estoque" (+ RCA p/ comprador/venda).
"""

import io
import os
import re
import csv
import json
from datetime import date, timedelta

from flask import Blueprint, jsonify, request, send_from_directory, Response, session

from . import pbi
from . import queries as Q
from . import core
from . import store
from . import historico
from . import provider_sql as PS   # modo DATA_SOURCE=postgres (lê do joga_demo)

bp = Blueprint("estoque", __name__, url_prefix="/estoque")


def _pg():
    """True quando o módulo lê do banco analítico (joga_demo) em vez do Power BI."""
    return pbi.CONFIG["data_source"] == "postgres"

# Diretório do próprio pacote — o index.html do módulo mora aqui, não na raiz do app.
_PKG_DIR = os.path.dirname(os.path.abspath(__file__))


def _mes_atual():
    return request.args.get("mes") or date.today().strftime("%Y-%m")


# Logo do CLIENTE impresso no PDF do pedido de compra (o produto é JOGA; o logo aqui é de
# quem emite o pedido ao fornecedor). Configurável por instância via CLIENTE_LOGO, com o
# arquivo da Multpel como padrão.
# ⚠️ O caminho mudou na fusão: os assets saíram de estoque/static/ para static/estoque/ na
# raiz do app. Como o Image() abaixo está sob try/except, o caminho errado não dava erro —
# o pedido simplesmente saía sem logo.
def compradores_reais(prod_map=None, forn_map=None):
    """[{codcomprador, comprador}] — só quem de fato COMPRA PARA REVENDA.

    ⚠️ Não confundir com `_compradores_map()`: aquele é a PCEMPR crua, ou seja, TODO
    funcionário (vendedores, financeiro, etc.). O comprador de verdade se deriva da base:
    fornecedor que tem produto de revenda → PCFORNEC.CODCOMPRADOR. Sem esse cruzamento a
    lista vem cheia de gente que não compra nada.

    Usada pela tela do módulo e pelo Admin (vínculo usuário↔comprador), para as duas
    mostrarem exatamente o mesmo conjunto.
    """
    if prod_map is None:
        prod_map = _cadastro_produtos()
    if forn_map is None:
        forn_map = _cadastro_fornecedores()
    comp_map = _compradores_map()
    forns_revenda = {int(core._n(p.get("CODFORNEC"))) for p in prod_map.values()
                     if p.get("CODFORNEC") not in (None, "")}
    cods = {int(core._n(forn_map[cf].get("CODCOMPRADOR"))) for cf in forns_revenda
            if cf in forn_map and forn_map[cf].get("CODCOMPRADOR") not in (None, "")}
    return sorted(
        [{"codcomprador": c, "comprador": comp_map.get(c, f"COMPRADOR {c}")} for c in cods],
        key=lambda x: x["comprador"] or "")


def _logo_cliente():
    """Logo impresso no PDF do pedido de compra. `None` = cabeçalho sem logo.

    ⚠️ O fallback é o logo do cliente ATUAL — o que, numa instância de demonstração, imprimiria a
    marca dele num pedido fictício. Por isso o fallback só vale quando a instância NÃO declarou um
    emitente próprio (`EMPRESA_RAZAO`): quem troca a empresa do pedido não quer o logo da outra."""
    env = os.getenv("CLIENTE_LOGO", "").strip()
    if env and os.path.exists(env):
        return env
    if os.getenv("EMPRESA_RAZAO", "").strip():
        return None
    raiz = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(raiz, "static", "estoque", "logo-multpel-trofeu.png")


# ───────────────────────── cadastro (cache 24h) ─────────────────────────
@pbi.cached(ttl=86400, key_fn=lambda: "cad:prod")
def _cadastro_produtos():
    if pbi.CONFIG["data_source"] == "postgres":
        return PS.cadastro_produto()
    rows = pbi.run_dax(Q.q_cadastro_produto())
    return {int(core._n(r["CODPROD"])): r for r in rows}


@pbi.cached(ttl=86400, key_fn=lambda: "cad:forn")
def _cadastro_fornecedores():
    if pbi.CONFIG["data_source"] == "postgres":
        return PS.cadastro_fornecedor()
    rows = pbi.run_dax(Q.q_cadastro_fornecedor())
    return {int(core._n(r["CODFORNEC"])): r for r in rows}


@pbi.cached(ttl=86400, key_fn=lambda: "filiais")
def _filiais_disponiveis():
    rows = PS.filiais() if pbi.CONFIG["data_source"] == "postgres" else pbi.run_dax(Q.q_filiais())
    fs = sorted({str(r["CODFILIAL"]).strip() for r in rows if r.get("CODFILIAL") not in (None, "")},
                key=lambda x: (len(x), x))
    return fs


@pbi.cached(ttl=86400, key_fn=lambda: "compradores")
def _compradores_map():
    """{matricula: nome} — PCEMPR no dataset Estoque (fallback RCA).

    ⚠️ Aqui existiu um mascarador de nomes (COMPRADOR_DEMO), de quando a demo rodava sobre a base
    REAL da Multpel. Removido em 08/2026: hoje a demo tem base própria (`joga_demo`, sintética),
    então não há nome real para esconder — mascarar de novo só reintroduziria um flag que alguém
    esqueceria ligado em produção."""
    if pbi.CONFIG["data_source"] == "postgres":
        return {int(core._n(r["MATRICULA"])): r["NOME"]
                for r in PS.compradores() if r.get("MATRICULA") not in (None, "") and r.get("NOME")}
    for runner, q in ((pbi.run_dax, Q.q_compradores_estoque()),
                      (pbi.run_dax_rca, Q.q_compradores_rca())):
        try:
            rows = runner(q)
            m = {int(core._n(r["MATRICULA"])): r["NOME"]
                 for r in rows if r.get("MATRICULA") not in (None, "") and r.get("NOME")}
            if m:
                return m
        except Exception:
            continue
    return {}


# ───────────────────────── snapshot (cache 30min por filial-set) ─────────────────────────
def _filiais_param():
    raw = request.args.get("filiais", "").strip()
    if not raw:
        return list(Q.FILIAIS_PADRAO)  # default: CDs 3 e 5 (estoque endereçado)
    return [f.strip() for f in raw.split(",") if f.strip()]


def _filiais_key(filiais):
    return ",".join(sorted(filiais)) if filiais else "ALL"


# ───────────────────────── unidades de negócio (estrutura da Multpel) ─────────────────────────
# Estoque físico (PCEST) e Venda (RCA faturamento) vivem em filiais DIFERENTES por unidade.
# Atacado: estoque nos CDs 3+5, mas fatura em 3+7+8 (5=depósito, 7/8=venda sem estoque).
# Lojas (A&M/AC) e JID são autossuficientes. Filiais 1,2,6,10-13,15 excluídas; 8 sem estoque.
# ⚠️ Rótulo sem marca de propósito ("Matriz", não "Multpel Matriz"): esta linha aparece no topo de
# TODA tela, inclusive na instância de demonstração, que roda a mesma imagem. Dentro do app do
# próprio cliente "Matriz" já é inequívoco — a marca ali só servia para vazar numa apresentação.
#
# ⚠️ São dado de INSTÂNCIA, não código — mesma régua já aplicada ao `EMPRESA_*` logo abaixo.
# A estrutura de filiais é da Multpel, e demo e produção rodam a MESMA imagem: sem sobrescrita,
# a instância de demonstração exibia "A&M", "JID", "AC" e "Telemarketing" para qualquer prospect,
# e ainda oferecia três unidades que abrem VAZIAS (a base sintética só tem as filiais 3 e 5) —
# tela em branco numa apresentação custa mais caro que nome estranho.
# Sem as env vars o comportamento é idêntico ao de sempre: os defaults são os dicionários da
# Multpel, então produção não precisa de configuração nenhuma.
_NOMES_FILIAL_PADRAO = {"3": "Matriz", "4": "A&M", "5": "Deposito",
                        "7": "Telemarketing", "8": "Atacado", "9": "JID", "14": "AC"}
_UNIDADES_PADRAO = {
    "atacado": {"nome": "Atacado", "estoque": ["3", "5"],            "venda": ["3", "7", "8"]},
    "am":      {"nome": "A&M",     "estoque": ["4"],                 "venda": ["4"]},
    "ac":      {"nome": "AC",      "estoque": ["14"],                "venda": ["14"]},
    "jid":     {"nome": "JID",     "estoque": ["9"],                 "venda": ["9"]},
    "todas":   {"nome": "Todas",   "estoque": ["3", "5", "4", "14", "9"], "venda": ["3", "7", "8", "4", "14", "9"]},
}


def _env_json(nome, padrao, validar=None):
    """Lê env var com JSON e cai no padrão se estiver ausente ou torta.

    Nunca levanta: estas variáveis são editadas à mão no Portainer, e uma vírgula sobrando não
    pode impedir o app de subir. Degrada para o default e AVISA no log — silêncio aqui seria
    pior, porque a instância rodaria com a nomenclatura errada sem ninguém perceber."""
    bruto = os.getenv(nome)
    if not bruto:
        return padrao
    try:
        val = json.loads(bruto)
        if validar and not validar(val):
            raise ValueError("estrutura inválida")
        return val
    except Exception as e:
        print(f"[config] {nome} ignorada ({e}); usando o padrão.")
        return padrao


def _valida_unidades(u):
    return (isinstance(u, dict) and u and all(
        isinstance(v, dict) and v.get("nome")
        and isinstance(v.get("estoque"), list) and isinstance(v.get("venda"), list)
        for v in u.values()))


NOMES_FILIAL = _env_json("NOMES_FILIAL_JSON", _NOMES_FILIAL_PADRAO,
                         lambda d: isinstance(d, dict))
UNIDADES = _env_json("UNIDADES_JSON", _UNIDADES_PADRAO, _valida_unidades)
# padrão tem de EXISTIR no dicionário em uso: apontar para uma unidade inexistente faria toda
# tela cair no KeyError de UNIDADES[_unidade()]
UNIDADE_PADRAO = os.getenv("UNIDADE_PADRAO", "atacado")
if UNIDADE_PADRAO not in UNIDADES:
    UNIDADE_PADRAO = next(iter(UNIDADES))

# Emitente do pedido de compra (cabeçalho do PDF, estilo relatório 211) + CNPJ que identifica
# TRANSFERÊNCIA entre filiais no orçamento (mesma raiz de CNPJ ⇒ não é compra).
#
# ⚠️ É dado de INSTÂNCIA, não código: cada `EMPRESA_*` sobrescreve por env var. O default segue a
# Multpel (cliente atual) para não exigir env em produção; a demo define os seus no compose e
# assim o PDF de pedido dela sai sem nenhum dado real — sem precisar de flag de "modo demo".
MULTPEL_EMPRESA = {
    "razao":    os.getenv("EMPRESA_RAZAO", "MULTPEL COM. DE PAPEIS E EMBALAGENS LTDA"),
    "cnpj":     os.getenv("EMPRESA_CNPJ", "02.262.785/0001-04"),
    "ie":       os.getenv("EMPRESA_IE", "081924950"),
    "endereco": os.getenv("EMPRESA_ENDERECO", "Rua Antonio Pedro Carleto, 56"),
    "bairro":   os.getenv("EMPRESA_BAIRRO", "Vila Rica"),
    "cep":      os.getenv("EMPRESA_CEP", "29301-200"),
    "cidade":   os.getenv("EMPRESA_CIDADE", "Cachoeiro de Itapemirim"),
    "uf":       os.getenv("EMPRESA_UF", "ES"),
    "tel":      os.getenv("EMPRESA_TEL", "(28) 3526-1450"),
    "email":    os.getenv("EMPRESA_EMAIL", "fiscal@mutpelatacado.com.br"),
}


def _unidade():
    u = (request.args.get("unidade") or UNIDADE_PADRAO).lower()
    return u if u in UNIDADES else UNIDADE_PADRAO


def _filiais_estoque():
    """Filiais de ESTOQUE físico (PCEST) da unidade atual."""
    return list(UNIDADES[_unidade()]["estoque"])


def _filiais_venda():
    """Filiais de VENDA/faturamento (RCA) da unidade atual."""
    return list(UNIDADES[_unidade()]["venda"])


def _snapshot_rows(filiais):
    key = f"snap:{_filiais_key(filiais)}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    rows = (PS.snapshot_estoque(filiais) if pbi.CONFIG["data_source"] == "postgres"
            else pbi.run_dax(Q.q_snapshot_estoque(filiais)))
    pbi._CACHE.set(key, rows, 1800)
    return rows


def _bloqueio_map(filiais):
    """{cod: [{qtbloq, dtultent, qtultent}, ...]} — bloqueio POR FILIAL (ver core.qt_em_transicao).

    ⚠️ Degrada para `{}` se a query falhar: sem o mapa, a pré-entrada volta ao cálculo agregado
    (que erra a favor de comprar de novo). Uma tabela nova não pode derrubar o módulo inteiro."""
    key = f"bloq:{_filiais_key(filiais)}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    m = {}
    try:
        rows = (PS.bloqueio_filial(filiais) if pbi.CONFIG["data_source"] == "postgres"
                else pbi.run_dax(Q.q_bloqueio_filial(filiais)))
        for r in rows:
            m.setdefault(int(core._n(r["CODPROD"])), []).append(
                {"qtbloq": r.get("qtbloq"), "qtultent": r.get("qtultent"),
                 "dtultent": r.get("dtultent")})
    except Exception as e:
        print(f"[bloqueio] posição por filial indisponível ({e}); pré-entrada no modo agregado.")
    pbi._CACHE.set(key, m, 1800)
    return m


def _endereco_map(filiais):
    """{cod: qt_end} — estoque endereçado (RUA<>99) nas filiais."""
    key = f"end:{_filiais_key(filiais)}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    rows = (PS.estoque_endereco(filiais) if pbi.CONFIG["data_source"] == "postgres"
            else pbi.run_dax(Q.q_estoque_endereco(filiais)))
    m = {int(core._n(r["CODPROD"])): core._n(r.get("qt_end")) for r in rows}
    pbi._CACHE.set(key, m, 1800)
    return m


def _posicoes_map(filiais):
    """{cod: nº de posições WMS ocupadas} — DISTINCTCOUNT(CODENDERECO), QT>0, RUA<>99."""
    key = f"pos:{_filiais_key(filiais)}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    try:
        rows = (PS.posicoes_por_produto(filiais) if pbi.CONFIG["data_source"] == "postgres"
                else pbi.run_dax(Q.q_posicoes_por_produto(filiais)))
        m = {int(core._n(r["CODPROD"])): int(core._n(r.get("pos"))) for r in rows}
    except Exception as e:
        print(f"[posicoes] WMS indisponível ({e}).")
        m = {}
    pbi._CACHE.set(key, m, 1800)
    return m


# ───────────────────────── embalagem / pedido real (cache) ─────────────────────────
@pbi.cached(ttl=86400, key_fn=lambda: "embalagem")
def _embalagem_map():
    """{cod: {qtunit, volume, ...}} — caixa/cubagem do PCEMBALAGEM (Estoque)."""
    try:
        rows = (PS.embalagem() if pbi.CONFIG["data_source"] == "postgres"
                else pbi.run_dax(Q.q_embalagem()))
        return {int(core._n(r["CODPROD"])): r for r in rows if r.get("CODPROD") not in (None, "")}
    except Exception as e:
        print(f"[embalagem] indisponível ({e}).")
        return {}


def _pedidos_data(filiais, hoje):
    """{'cab': [...PCPEDIDO], 'ja_pedida': {cod: qt_aberta}} — pedido de compra REAL (Winthor).
    Reutilizado pelo abastecimento (já-pedido) e pelo orçamento/acompanhamento (cabeçalho).
    Degrada p/ vazio se indisponível. Cache 30min por filial-set + data."""
    key = f"peddata:{_filiais_key(filiais)}:{hoje.isoformat()}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    data = {"cab": [], "itens": [], "ja_pedida": {}}
    _pg = pbi.CONFIG["data_source"] == "postgres"
    try:
        cab = (PS.pedido_cab(hoje - timedelta(days=180), filiais) if _pg
               else pbi.run_dax(Q.q_pedido_cab(hoje - timedelta(days=180), filiais)))
        if cab:
            numped_min = min(int(core._n(r["NUMPED"])) for r in cab)
            itens = (PS.pedido_itens(numped_min) if _pg
                     else pbi.run_dax(Q.q_pedido_itens(numped_min)))
            data = {"cab": cab, "itens": itens,
                    "ja_pedida": core.montar_ja_pedida(cab, itens, hoje=hoje, dias=180),
                    # IPI/ST efetivos praticados por (fornecedor, produto) — mesma matéria-prima
                    # do já-pedido, NENHUMA query a mais. Alimenta o valor sugerido na régua da NF.
                    "tributacao": core.montar_tributacao(cab, itens, hoje=hoje, dias=180)}
    except Exception as e:
        print(f"[pedidos] Winthor indisponível ({e}). Pedido real desabilitado.")
    pbi._CACHE.set(key, data, 1800)
    return data


def _trib_entrada_map():
    """{(codprod, filial, uf, tipofornec): {ipi, st}} — tributação de entrada do ERP (rotina 212).
    Cache 6h: é cadastro fiscal, muda pouco (mas muda ANTES do histórico — é por isso que ela é
    a fonte primária). Ausente (instância sem a tabela publicada) → {} e a cascata degrada."""
    key = "tribentrada:v1"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    out = {}
    try:
        rows = PS.trib_entrada() if pbi.CONFIG["data_source"] == "postgres" else pbi.run_dax(Q.q_trib_entrada())
        out = core.montar_trib_entrada(rows)
    except Exception as e:
        print(f"[tributacao] TRIB_ENTRADA indisponível ({e}). Cascata cai no cadastro/histórico.")
    pbi._CACHE.set(key, out, 21600)
    return out


def _hoje():
    """O "hoje" do módulo. Em modo BD ele é ANCORADO NO DADO, não no calendário.

    ⚠️ O Comercial já fazia isso (`provider_sql.hoje_analitico`, ~25 chamadas entre
    `provider_sql.py` e `server.py`); o Compras ficou de fora e usava `date.today()` puro. O
    resultado, medido na demo em 18/08/2026: o fato sintético terminava em 24/07 e o relógio dizia
    18/08, então TODA janela de data do módulo caía 25 dias à frente do dado —
    venda "mês atual" = R$ 0, meta do orçamento = 0, `dias_sem_venda` inflado em 25 dias (mais
    itens virando "parado") e, o mais visível, `core._aplicar_curva` carimbando o catálogo INTEIRO
    como curva C (3.781 produtos, nenhum A ou B) porque o total de venda da janela era zero.

    Só o modo `postgres` muda — o caminho Power BI segue com `date.today()`, que é o correto no
    cliente real. É a regra do README: caminho novo atrás de `data_source == 'postgres'`, o default
    sai idêntico.

    (`hoje_analitico` cacheia por processo, igual no Comercial: base que recebe dado novo só vê a
    âncora nova depois de reiniciar. Aceito — é o comportamento que o Comercial já tem.)"""
    h = request.args.get("hoje")
    if h:
        try:
            return date.fromisoformat(h)
        except ValueError:
            pass
    if pbi.CONFIG["data_source"] == "postgres":
        try:
            return PS.hoje_analitico()
        except Exception as e:      # banco analítico fora não pode derrubar toda tela
            print(f"[hoje] âncora analítica indisponível ({e}); caindo no calendário.")
    return date.today()


# ───────────────────────── lead time (aba Análise → Lead time) ─────────────────────────
def _leadtime_raw(hoje):
    """Linhas cruas do lead time: cabeçalho de pedidos 12m (PCPEDIDO) + ponte de entrada
    (PEDIDO_ENTRADA). Cache próprio p/ o drill por fornecedor reusar sem novo fetch.
    SEM recorte de filial de propósito: lead time é característica do FORNECEDOR, não da
    filial — e assim o cache é um só. Cache 30min."""
    key = f"leadtime:raw:{hoje.isoformat()}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    if pbi.CONFIG["data_source"] == "postgres":
        raw = {"cab": PS.pedido_cab(hoje - timedelta(days=365), None), "entradas": PS.pedido_entrada()}
    else:
        raw = {"cab": pbi.run_dax(Q.q_pedido_cab(hoje - timedelta(days=365), filiais=None)),
               "entradas": pbi.run_dax(Q.q_pedido_entrada())}
    pbi._CACHE.set(key, raw, 1800)
    return raw


# ───────────────────────── verbas (aba Análise → Verbas) ─────────────────────────
def _verbas_raw(hoje):
    """Linhas cruas de PCVERBA + PCAPLICVERBA (2024+ na origem). Cache 30min."""
    key = f"verbas:raw:{hoje.isoformat()}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    if pbi.CONFIG["data_source"] == "postgres":
        raw = {"verbas": PS.verbas(), "aplics": PS.verba_aplic()}
    else:
        raw = {"verbas": pbi.run_dax(Q.q_verbas()),
               "aplics": pbi.run_dax(Q.q_verba_aplic())}
    pbi._CACHE.set(key, raw, 1800)
    return raw


def _compras_12m_map(hoje):
    """{codfornec: Σ VLTOTAL dos pedidos 12m} — REUSA o cab do lead time (nenhuma query
    nova); transferência entre filiais (raiz de CNPJ) fica fora, igual ao orçamento."""
    forn_map = _cadastro_fornecedores()
    raiz = core._cnpj_raiz(MULTPEL_EMPRESA["cnpj"])
    out = {}
    for r in _leadtime_raw(hoje)["cab"]:
        cod = int(core._n(r.get("CODFORNEC")))
        if raiz and core._cnpj_raiz((forn_map.get(cod) or {}).get("CGC")) == raiz:
            continue
        out[cod] = out.get(cod, 0.0) + core._n(r.get("VLTOTAL"))
    return out


def _forn_extra_map(hoje, periodo, filiais):
    """{codfornec: {n_pedidos, ciclo_dias, ultima_compra, verba, verba_campanha}} — as colunas
    novas da aba Fornecedores (ciclo de compras + verba), pedido do diretor 07/2026.

    NENHUMA query nova: o cabeçalho de pedidos vem do cache do Lead time (`_leadtime_raw`, 12m)
    e a verba do cache da aba Verbas (`_verbas_raw`). Degrada p/ {} se qualquer um cair — a aba
    continua funcionando sem as colunas novas em vez de quebrar inteira.

    Janelas (ver os docstrings em core): pedidos e verba na janela do seletor "Venda" (a mesma do
    lucro, senão a soma lucro+verba não faria sentido); o CICLO em 12m fixos.
    """
    key = f"fornextra:{_filiais_key(filiais)}:{periodo}:{hoje.isoformat()}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    ini, fim = _venda_datas(periodo, hoje)
    forn_map = _cadastro_fornecedores()
    cnpj = MULTPEL_EMPRESA["cnpj"]
    out = {}
    try:
        ciclo = core.ciclo_compras(_leadtime_raw(hoje)["cab"], ini, fim, filiais=filiais,
                                   forn_map=forn_map, cnpj_empresa=cnpj,
                                   ciclo_desde=hoje - timedelta(days=365))
        for cod, v in ciclo.items():
            out.setdefault(cod, {}).update(v)
    except Exception as e:
        print(f"[fornecedores] ciclo de compras indisponível ({e}).")
    try:
        raw = _verbas_raw(hoje)
        vb = core.verba_por_fornecedor(raw["verbas"], raw["aplics"], ini, fim,
                                       forn_map=forn_map, cnpj_empresa=cnpj)
        for cod, v in vb.items():
            out.setdefault(cod, {}).update(v)
    except Exception as e:
        print(f"[fornecedores] verba indisponível ({e}).")
    # crescimento YoY na régua COMPLETA (todos os produtos vendidos, não só os que ainda estão em
    # estoque) — ver core.yoy_fornecedor. Reusa os mesmos mapas de venda que o snapshot já busca,
    # então não custa query nova; degrada p/ o cálculo antigo se o RCA cair.
    try:
        fil_v = _filiais_venda()
        fornec_de_prod = {c: int(core._n(r.get("CODFORNEC")))
                          for c, r in _cadastro_produtos().items()
                          if core._n(r.get("CODFORNEC"))}
        yoy = core.yoy_fornecedor(_vendas_map(periodo, hoje, fil_v),
                                  _vendas_ano_ant_map(periodo, hoje, fil_v), fornec_de_prod)
        for cod, v in yoy.items():
            out.setdefault(cod, {}).update(v)
    except Exception as e:
        print(f"[fornecedores] YoY completo indisponível ({e}).")
    pbi._CACHE.set(key, out, 1800)
    return out


def _verbas_res(hoje, comprador=None, fornec=None):
    """Agregado da aba Verbas (core.verbas_fornecedores) — fecha o TRIPÉ cruzando com a
    compra 12m e o lead real já calculados. Degrada p/ vazio se o BI cair. Cache 30min.

    `comprador` e `fornec` recortam no CORE (não no cliente) para o resumo dos cards, o gráfico
    mensal e o "por conta" falarem do mesmo universo da tabela. O raw (a query cara) é
    compartilhado entre todos, então o recorte custa só a reagregação em memória.

    ⚠️ A chave de cache carrega os DOIS. Recorte novo que não entre aqui faz o primeiro
    fornecedor consultado ser servido a todos os outros por 30 min — e números plausíveis do
    fornecedor errado não denunciam nada na tela."""
    key = f"verbas:{hoje.isoformat()}:{comprador or 'todos'}:{fornec or 'todos'}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    res = {"resumo": {}, "contas": [], "meses": [], "fornecedores": [], "grandes_sem_verba": []}
    try:
        raw = _verbas_raw(hoje)
        lead_map = {f["codfornec"]: f.get("lead_real")
                    for f in _leadtime_res(hoje).get("fornecedores", [])}
        res = core.verbas_fornecedores(raw["verbas"], raw["aplics"], _cadastro_fornecedores(),
                                       _compradores_map(), compras_map=_compras_12m_map(hoje),
                                       lead_map=lead_map, hoje=hoje,
                                       cnpj_empresa=MULTPEL_EMPRESA["cnpj"],
                                       comprador=comprador, fornec=fornec)
    except Exception as e:
        print(f"[verbas] indisponível ({e}).")
    pbi._CACHE.set(key, res, 1800)
    return res


def _leadtime_res(hoje, comprador=None):
    """Agregado da aba Lead time (core.leadtime_fornecedores sobre o raw acima).
    Degrada p/ vazio se o BI estiver fora. Cache 30min (chave por comprador — ver _verbas_res).

    ⚠️ Quem consome isto como MAPA por fornecedor (verbas, /api/fornecedores_extra) chama SEM
    comprador de propósito: lá o recorte é do consumidor, e um mapa parcial viraria coluna vazia."""
    key = f"leadtime:{hoje.isoformat()}:{comprador or 'todos'}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    res = {"resumo": {}, "faixas": [], "fornecedores": []}
    try:
        raw = _leadtime_raw(hoje)
        res = core.leadtime_fornecedores(raw["cab"], raw["entradas"], _cadastro_fornecedores(),
                                         _compradores_map(), hoje=hoje,
                                         cnpj_empresa=MULTPEL_EMPRESA["cnpj"],
                                         comprador=comprador)
    except Exception as e:
        print(f"[leadtime] indisponível ({e}).")
    pbi._CACHE.set(key, res, 1800)
    return res


# ───────────────────────── venda real (dataset RCA, cache 30min) ─────────────────────────
def _venda_datas(periodo, hoje):
    if periodo == "30d":
        return hoje - timedelta(days=30), hoje
    if periodo == "90d":
        return hoje - timedelta(days=90), hoje
    if periodo == "6m":
        return hoje - timedelta(days=180), hoje
    if periodo == "12m":
        return hoje - timedelta(days=365), hoje
    return hoje.replace(day=1), hoje  # mês atual (default)


def _vendas_map(periodo, hoje, filiais=None):
    """{cod: {venda, custo, qtd}} líquido (venda − devoluções) do RCA p/ o período selecionado."""
    ini, fim = _venda_datas(periodo, hoje)
    return _vendas_liquidas(ini, fim, filiais)


def _ano_antes(d):
    """Mesma data no ano anterior (29/02 → 28/02)."""
    try:
        return d.replace(year=d.year - 1)
    except ValueError:
        return d.replace(year=d.year - 1, day=28)


def _vendas_ano_ant_map(periodo, hoje, filiais=None):
    """{cod: {venda, custo, qtd}} líquido do MESMO período no ANO ANTERIOR — base do
    crescimento (YoY) por item e, agregado, por fornecedor. Mesma fórmula da venda atual
    (líquida), senão o crescimento sairia inflado (bruta × líquida)."""
    ini, fim = _venda_datas(periodo, hoje)
    return _vendas_liquidas(_ano_antes(ini), _ano_antes(fim), filiais)


def _vendas_liquidas(ini, fim, filiais=None):
    """{cod: {venda, custo, qtd}} líquido (venda − devoluções − devol. avaria) do RCA no
    intervalo, escopado por filiais de VENDA. Degrada p/ {} se o RCA estiver indisponível."""
    if pbi.CONFIG["data_source"] == "postgres":   # Inc.2: venda por produto do joga_demo
        return PS.vendas_por_produto(ini, fim, filiais)
    key = f"vendaliq:{_filiais_key(filiais)}:{ini}:{fim}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    m = {}
    try:
        for r in pbi.run_dax_rca(Q.q_vendas_rca(ini, fim, filiais)):
            c = int(core._n(r["CODPROD"]))
            m[c] = {"venda": core._n(r.get("venda")), "custo": core._n(r.get("custo")), "qtd": core._n(r.get("qtd"))}
        # ⚠️ A QUANTIDADE também é abatida (08/2026). Antes só valor e custo eram: a "Qtd vendida"
        # saía bruta e qualquer preço médio (valor líquido ÷ qtd bruta) misturava duas réguas.
        # Medido em julho no BI real: 140.608 un devolvidas contra 1.206.349 vendidas = 11,7%.
        # `max(0, ...)` porque devolução por DTENT e venda por DTSAIDA são janelas diferentes:
        # um item devolvido em julho pode ter saído em junho, e a subtração ficaria negativa.
        for r in pbi.run_dax_rca(Q.q_devol_rca(ini, fim, filiais)):
            c = int(core._n(r["CODPROD"]))
            if c in m:
                m[c]["venda"] -= core._n(r.get("dev")); m[c]["custo"] -= core._n(r.get("cdev"))
                m[c]["qtd"] = max(0.0, m[c]["qtd"] - core._n(r.get("qtdev")))
        for r in pbi.run_dax_rca(Q.q_devol_av_rca(ini, fim, filiais)):
            c = int(core._n(r["CODPROD"]))
            if c in m:
                m[c]["venda"] -= core._n(r.get("devav")); m[c]["custo"] -= core._n(r.get("cdevav"))
                m[c]["qtd"] = max(0.0, m[c]["qtd"] - core._n(r.get("qtdevav")))
    except Exception as e:
        print(f"[venda] RCA indisponível ({e}). Camada de vendas desabilitada.")
        m = {}
    pbi._CACHE.set(key, m, 1800)
    return m


def _vendas_forn_mensal_map(hoje, filiais=None, meses=24):
    """{codfornec: {AnoMes(int): venda_LÍQUIDA R$}} — série do drawer 360° do FORNECEDOR.

    **24 meses** (não 12) para o gráfico sobrepor o MESMO mês do ano anterior: sem isso o
    diretor vê "−20,4%" na coluna e não enxerga *quando* caiu. Cabe folgado: agregada por
    fornecedor são ~6k linhas (a série por produto na mesma janela daria ~54k, e o
    `executeQueries` corta em 100.000).

    ⚠️ Régua COMPLETA — agrega no fato por `CODFORNEC`, não somando os produtos da tela.
    Ver `q_venda_fornecedor_mensal_rca` e `core.yoy_fornecedor` para o porquê (item que saiu de
    linha some do histórico e chega a inverter o sinal do crescimento).

    Sob demanda (só quando abre um fornecedor) e cacheado 12h, igual ao mapa do produto."""
    ini = hoje.replace(day=1)
    for _ in range(int(meses)):
        ini = (ini - timedelta(days=1)).replace(day=1)
    key = f"vfornmes:{_filiais_key(filiais)}:{meses}:{hoje.strftime('%Y-%m')}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    _pgm = pbi.CONFIG["data_source"] == "postgres"
    m = {}
    try:
        for r in (PS.venda_fornecedor_mensal(ini, filiais) if _pgm
                  else pbi.run_dax_rca(Q.q_venda_fornecedor_mensal_rca(ini, filiais))):
            cf = int(core._n(r.get("CODFORNEC")))
            am = int(core._n(r.get("AnoMes")))
            if cf and am:
                m.setdefault(cf, {})[am] = core._n(r.get("venda"))
        for r in (PS.devol_fornecedor_mensal(ini, filiais) if _pgm
                  else pbi.run_dax_rca(Q.q_devol_fornecedor_mensal_rca(ini, filiais))):
            cf = int(core._n(r.get("CODFORNEC")))
            am = int(core._n(r.get("AnoMes")))
            if cf in m and am in m[cf]:
                m[cf][am] -= core._n(r.get("dev"))
    except Exception as e:
        print(f"[venda forn 24m] RCA indisponível ({e}). Drawer do fornecedor sai sem o gráfico.")
        m = {}
    pbi._CACHE.set(key, m, 43200)   # 12h
    return m


def _vendas_mensal_rs_map(hoje, filiais=None):
    """({cod: {AnoMes: venda_LÍQUIDA R$}}, {cod: {AnoMes: clientes}}) dos últimos 12 meses —
    alimenta SÓ o gráfico "venda 12m" do drawer 360°. Mapa SEPARADO do de quantidade (que
    alimenta giro/forecast) de propósito: mexer naquele quebraria a metodologia de giro já
    calibrada. Líquida (bruta − devolução) p/ bater com o "Venda no período" do próprio drawer.
    Buscado sob demanda (só quando abre um produto) e cacheado 12h p/ servir os demais.

    O 2º mapa é a POSITIVAÇÃO do item (clientes distintos no mês), pedido do diretor 07/2026
    para separar queda por perda de BASE de queda por VOLUME. Vem da mesma query, sem custo de
    linha. ⚠️ NÃO leva desconto de devolução: positivação é visita ("o cliente comprou este item
    no mês"), e devolver não desfaz a visita — diferente do valor, que é líquido."""
    ini = (hoje.replace(day=1) - timedelta(days=1)).replace(day=1)
    for _ in range(12):
        ini = (ini - timedelta(days=1)).replace(day=1)
    key = f"vmesrs:{_filiais_key(filiais)}:{hoje.strftime('%Y-%m')}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        # tolera entrada gravada pelo formato antigo (só o mapa de venda): o cache é em memória e
        # morre no deploy, mas um processo de longa vida não pode quebrar por causa disso
        return hit if isinstance(hit, tuple) else (hit, {})
    _pg = pbi.CONFIG["data_source"] == "postgres"
    m, cli = {}, {}
    try:
        venda_rows = PS.venda_produto_mensal(ini, filiais) if _pg else pbi.run_dax_rca(Q.q_venda_produto_mensal_rca(ini, filiais))
        for r in venda_rows:
            c = int(core._n(r["CODPROD"])); am = int(core._n(r.get("AnoMes")))
            if am:
                m.setdefault(c, {})[am] = core._n(r.get("venda"))
                if r.get("clientes") is not None:
                    cli.setdefault(c, {})[am] = int(core._n(r.get("clientes")))
        devol_rows = PS.devol_produto_mensal(ini, filiais) if _pg else pbi.run_dax_rca(Q.q_devol_produto_mensal_rca(ini, filiais))
        for r in devol_rows:
            c = int(core._n(r["CODPROD"])); am = int(core._n(r.get("AnoMes")))
            if am and c in m and am in m[c]:
                m[c][am] -= core._n(r.get("dev"))
    except Exception as e:
        print(f"[venda 12m] RCA indisponível ({e}). Gráfico de venda do 360° cai p/ unidades.")
        m, cli = {}, {}
    pbi._CACHE.set(key, (m, cli), 43200)  # 12h
    return m, cli


def _vendas_mensal_map(meses, hoje, profundo=False, filiais=None):
    """{cod: {AnoMes: qtd}} — venda mensal (QT) do RCA p/ o forecast, escopada por filiais de
    VENDA da unidade. Cache 12h. Degrada p/ {} se RCA indisponível. Só quando forecast ligado.
    profundo=True (sazonalidade) força ≥25 meses de histórico p/ o fator ano-a-ano."""
    fetch = max(25, int(meses)) if profundo else max(1, int(meses))
    ini = (hoje.replace(day=1) - timedelta(days=1)).replace(day=1)  # 1º dia do mês anterior
    for _ in range(fetch):
        ini = (ini - timedelta(days=1)).replace(day=1)
    key = f"vmes:{fetch}:{_filiais_key(filiais)}:{hoje.strftime('%Y-%m')}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    m = {}
    try:
        rows = (PS.vendas_mensal_qt(ini, filiais) if pbi.CONFIG["data_source"] == "postgres"
                else pbi.run_dax_rca(Q.q_vendas_mensal_rca(ini, filiais)))
        for r in rows:
            c = int(core._n(r["CODPROD"]))
            am = int(core._n(r.get("AM")))
            if am:
                m.setdefault(c, {})[am] = core._n(r.get("qtd"))
    except Exception as e:
        print(f"[forecast] RCA mensal indisponível ({e}). Forecast desabilitado.")
        m = {}
    pbi._CACHE.set(key, m, 43200)  # 12h
    return m


def _desempenho_data(periodo, hoje, filiais=None):
    """{resumo, compradores} — desempenho comercial por comprador (RCA), escopado por filiais de
    VENDA da unidade. Espelha RECEITA COMPRADOR + comparativo ano×ano. Cache 30min."""
    ini, fim = _venda_datas(periodo, hoje)
    _pg = pbi.CONFIG["data_source"] == "postgres"
    key = f"desemp:{periodo}:{_filiais_key(filiais)}:{ini}:{fim}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    res = {"resumo": {}, "compradores": []}
    try:
        receita = (PS.receita_comprador(ini, fim, filiais) if _pg
                   else pbi.run_dax_rca(Q.q_receita_comprador_rca(ini, fim, filiais)))
        devol, custo_dev = {}, {}
        _devol_rows = (PS.devol_comprador(ini, fim, filiais) if _pg
                       else pbi.run_dax_rca(Q.q_devol_comprador_rca(ini, fim, filiais)))
        for r in _devol_rows:
            cc = r.get("CODCOMPRADOR")
            if cc not in (None, ""):
                k = int(core._n(cc))
                devol[k] = core._n(r.get("dev"))
                custo_dev[k] = core._n(r.get("cdev"))
        # mesmo período no ano anterior (comparativo YoY de venda E lucro por comprador)
        ini_ant = ini.replace(year=ini.year - 1)
        fim_ant = fim.replace(year=fim.year - 1)
        venda_ant, custo_ant = {}, {}
        _va_rows = (PS.venda_comprador_periodo(ini_ant, fim_ant, filiais) if _pg
                    else pbi.run_dax_rca(Q.q_venda_comprador_periodo_rca(ini_ant, fim_ant, filiais)))
        for r in _va_rows:
            cc = r.get("CODCOMPRADOR")
            if cc not in (None, ""):
                venda_ant[int(core._n(cc))] = core._n(r.get("venda"))
                custo_ant[int(core._n(cc))] = core._n(r.get("custo"))
        res = core.desempenho_comprador(receita, devol, _compradores_map(), venda_ant, custo_ant,
                                        custo_dev)
    except Exception as e:
        print(f"[desempenho] RCA indisponível ({e}). Aba de desempenho desabilitada.")
        res = {"resumo": {}, "compradores": []}
    pbi._CACHE.set(key, res, 1800)
    return res


def _venda_comprador_30d(fil_estoque, fil_venda, hoje):
    """{nome_comprador: venda_liquida_30d} — base da meta de orçamento (65%). Estoque e venda
    escopados por filiais diferentes da unidade (ex.: Atacado = estoque 3+5, venda 3+7+8)."""
    key = f"vcomp30:{_filiais_key(fil_estoque)}:{_filiais_key(fil_venda)}:{hoje.isoformat()}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    produtos = core.construir_produtos(
        _snapshot_rows(fil_estoque), _endereco_map(fil_estoque), _cadastro_produtos(),
        _cadastro_fornecedores(), _compradores_map(), _vendas_map("30d", hoje, fil_venda),
        dict(core.DEFAULTS), hoje=hoje)
    m = {g["comprador"]: g["venda"] for g in core.por_comprador(produtos) if g.get("comprador")}
    pbi._CACHE.set(key, m, 1800)
    return m


def _build_produtos():
    """Constrói a lista enriquecida de produtos para a unidade/params atuais.
    Estoque (snapshot/endereço/pedido) usa as filiais de ESTOQUE; venda/forecast usam as de VENDA."""
    filiais_e = _filiais_estoque()
    filiais_v = _filiais_venda()
    params = core.merge_params(request.args.to_dict())
    snap = _snapshot_rows(filiais_e)
    end_map = _endereco_map(filiais_e)
    prod_map = _cadastro_produtos()
    forn_map = _cadastro_fornecedores()
    comp_map = _compradores_map()
    venda_map = _vendas_map(request.args.get("venda_periodo", "mes"), _hoje(), filiais_v)
    # série mensal do RCA (QT): sempre buscada (cache 12h). Com forecast ligado alimenta o
    # forecast; desligado, serve ao fallback de giro dos ITENS NOVOS e à série de 12 meses do
    # 360°. Janela de 12m p/ o gráfico de venda do drawer ficar completo — o giro/forecast leem
    # só os meses que precisam, então a janela maior não muda nenhum cálculo.
    if params.get("forecast"):
        venda_mensal = _vendas_mensal_map(max(12, int(params["forecast_meses"])), _hoje(),
                                          profundo=bool(params.get("forecast_sazonal")), filiais=filiais_v)
    else:
        venda_mensal = _vendas_mensal_map(12, _hoje(), filiais=filiais_v)
    _peddata = _pedidos_data(filiais_e, _hoje())
    ja_pedida = _peddata["ja_pedida"]
    embalagem = _embalagem_map()
    preco_venda = _preco_venda_map(filiais_v)
    # crescimento (YoY): venda líquida do MESMO período no ano anterior, por produto.
    # Serve o item (aba Produtos) e, agregado por fornecedor, a aba Fornecedores.
    venda_ant = _vendas_ano_ant_map(request.args.get("venda_periodo", "mes"), _hoje(), filiais_v)
    produtos = core.construir_produtos(snap, end_map, prod_map, forn_map, comp_map, venda_map, params,
                                       hoje=_hoje(), venda_mensal_map=venda_mensal,
                                       ja_pedida_map=ja_pedida, embalagem_map=embalagem,
                                       preco_venda_map=preco_venda, venda_ant_map=venda_ant,
                                       tributacao_map=_peddata.get("tributacao"),
                                       trib_entrada_map=_trib_entrada_map(),
                                       bloqueio_map=_bloqueio_map(filiais_e))
    # ocupação WMS: nº de posições por item + volume endereçado (m³) + flag "espaço morto".
    pos_map = _posicoes_map(filiais_e)
    for p in produtos:
        pe = pos_map.get(p["codprod"], 0)
        p["pos_end"] = pe
        cx = p.get("caixa") or 1
        vol_un = (p.get("cubagem_caixa_m3") or 0) / cx if cx else 0
        p["m3_end"] = round((p.get("qt_end") or 0) * vol_un, 3)
        # espaço morto: ocupa >=3 posições e praticamente não gira
        dsv = p.get("dias_sem_venda")
        p["espaco_morto"] = pe >= 3 and ((p.get("giro_mes") or 0) <= 0 or (dsv is not None and dsv >= 90))
    return produtos, params, filiais_e


def _preco_venda_map(filiais):
    """{cod: preço de venda unitário} — realizado médio dos ÚLTIMOS 3 MESES (RCA), janela FIXA
    (independe do filtro de período), p/ a 'venda perdida' não variar com o seletor de venda e
    alinhar com a janela do giro (também 3m). O preço de tabela do BI (PCPRODUT[PVENDA]) está
    vazio; usar o realizado 3m como referência. Cache mensal (6h)."""
    hoje = _hoje()
    if pbi.CONFIG["data_source"] == "postgres":   # Inc.2: preço realizado 3m do joga_demo
        vp = PS.vendas_por_produto(hoje - timedelta(days=90), hoje, filiais)
        return {c: d["venda"] / d["qtd"] for c, d in vp.items()
                if (d.get("qtd") or 0) > 0 and (d.get("venda") or 0) > 0}
    key = f"precov:{_filiais_key(filiais)}:{hoje.isoformat()[:7]}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    m = {}
    try:
        for r in pbi.run_dax_rca(Q.q_vendas_rca(hoje - timedelta(days=90), hoje, filiais)):
            c = int(core._n(r["CODPROD"])); v = core._n(r.get("venda")); q = core._n(r.get("qtd"))
            if q > 0 and v > 0:
                m[c] = v / q
    except Exception as e:
        print(f"[preco_venda] RCA indisponível ({e}). Venda perdida cai no custo.")
    pbi._CACHE.set(key, m, 6 * 3600)
    return m


# ───────────────────────── páginas ─────────────────────────
@bp.route("/")
def index():
    return send_from_directory(_PKG_DIR, "index.html")


# ───────────────────────── API ─────────────────────────
@bp.route("/api/filtros")
def api_filtros():
    prod_map = _cadastro_produtos()
    forn_map = _cadastro_fornecedores()
    deptos = sorted({str(p.get("CODEPTO")) for p in prod_map.values()
                     if p.get("CODEPTO") not in (None, "")})
    fornecedores = sorted(
        [{"codfornec": cf, "fornecedor": f.get("FORNECEDOR") or f"FORN {cf}"}
         for cf, f in forn_map.items()],
        key=lambda x: x["fornecedor"] or "")
    compradores = compradores_reais(prod_map, forn_map)
    return jsonify({
        "ok": True,
        "filiais": _filiais_disponiveis(),
        "filiais_padrao": list(Q.FILIAIS_PADRAO),
        "unidades": [{"id": uid, "nome": u["nome"],
                      "cod": "" if uid == "todas" else ",".join(sorted(set(u["estoque"] + u["venda"]), key=int))}
                     for uid, u in UNIDADES.items()],
        "unidade_padrao": UNIDADE_PADRAO,
        "nomes_filial": NOMES_FILIAL,
        "deptos": deptos,
        "fornecedores": fornecedores,
        "compradores": compradores,
        # Comprador vinculado ao usuário no Admin: é o filtro INICIAL da tela, não uma trava.
        # Ele pode trocar para qualquer outro (e a escolha passa a valer via prefs locais).
        "comprador_padrao": session.get("codcomprador"),
        "defaults": core.DEFAULTS,
    })


@bp.route("/api/snapshot")
def api_snapshot():
    produtos, params, filiais = _build_produtos()
    # ⚠️ ciclo de compras / verba NÃO entram aqui de propósito: exigiriam o cab de pedidos 12m e a
    # PCVERBA, que hoje só são pagos por quem abre Lead time/Verbas. Pendurá-los no snapshot (o
    # endpoint mais quente do módulo, que TODA tela carrega) faria a tela inicial pagar por duas
    # abas que a maioria não abre. Vão no /api/fornecedores_extra, buscado só pela aba.
    return jsonify({
        "ok": True,
        "gerado_em": date.today().isoformat(),
        "bi_refresh": pbi.get_dataset_refresh(),
        "filiais": filiais or "ALL",
        "unidade": _unidade(),
        "unidade_nome": UNIDADES[_unidade()]["nome"],
        "params": params,
        "n": len(produtos),
        "produtos": produtos,
        # ⚠️ `params` viaja porque a watchlist "Em desaceleração" é parametrizável (janela,
        # cobertura e piso de valor). Sem ele o card sairia sempre nos defaults e os campos de
        # ⚙ Parâmetros não moveriam nada — a mesma falha silenciosa do `novo_dias`.
        "cockpit": core.cockpit(produtos, params),
        "fornecedores": core.fornecedores(produtos, params),
        "compradores": core.por_comprador(produtos),
    })


@bp.route("/api/fornecedores_extra")
def api_fornecedores_extra():
    """Colunas da aba Fornecedores que não saem da posição de estoque: ciclo de compras
    (PCPEDIDO) + verba negociada (PCVERBA), por fornecedor.

    Endpoint SEPARADO do snapshot de propósito — ver o comentário em `api_snapshot`. A aba busca
    isto ao abrir; enquanto não chega, ela renderiza sem as colunas novas em vez de travar.

    Desde 07/2026 leva também o **lead real** por fornecedor, porque a aba Abastecimento passou a
    exibi-lo na linha (pedido do diretor: "não preciso entrar na aba lead time para ver o prazo
    do fornecedor, só olhar ali e ajustar no parâmetro"). Sai por aqui e NÃO pelo `/api/snapshot`
    pela mesma razão do ciclo/verba: o snapshot é carregado por TODAS as telas, e a maioria não
    usa lead. `_leadtime_res` já é cache de 30min — custo zero de query.
    """
    hoje = _hoje()
    _, _, filiais = _build_produtos()
    extra = dict(_forn_extra_map(hoje, request.args.get("venda_periodo", "mes"), filiais) or {})
    try:
        for f in _leadtime_res(hoje).get("fornecedores", []):
            cf = f.get("codfornec")
            if cf is None:
                continue
            # `confiavel` viaja junto de propósito: o comprador vai MEXER no parâmetro com base
            # neste número, e 14d medidos em 2 entradas não sustentam a mesma decisão que 14d
            # medidos em 40. Sem a marca, a tela apresentaria amostra fraca como fato.
            extra[cf] = {**(extra.get(cf) or {}), "lead_real": f.get("lead_real"),
                         "lead_n": f.get("n"), "lead_confiavel": f.get("confiavel")}
    except Exception as e:      # lead fora não pode derrubar ciclo e verba junto
        print(f"[fornecedores_extra] lead time indisponível ({e}).")
    return jsonify({"ok": True, "extra": extra})


@bp.route("/api/desempenho")
def api_desempenho():
    """Desempenho comercial por comprador: venda líquida, lucro, margem ponderada, positivação
    (clientes distintos), devolução e comparativo ano×ano. Período via ?venda_periodo=."""
    periodo = request.args.get("venda_periodo", "mes")
    d = _desempenho_data(periodo, _hoje(), _filiais_venda())
    return jsonify({"ok": True, "periodo": periodo,
                    "resumo": d["resumo"], "compradores": d["compradores"]})


@bp.route("/api/leadtime")
def api_leadtime():
    """Lead time por fornecedor (12m): 1º recebimento (PEDIDO_ENTRADA) − emissão (PCPEDIDO).
    Devolve os DOIS leads (com/sem 'digitado na hora') — ver core.leadtime_fornecedores.
    `?comprador_cod=` recorta a base (inclusive o resumo dos cards)."""
    res = _leadtime_res(_hoje(), comprador=request.args.get("comprador_cod"))
    return jsonify({"ok": True, "gerado_em": date.today().isoformat(),
                    "bi_refresh": pbi.get_dataset_refresh(), **res})


def _qualidade_cadastro_res(comprador=None):
    """Bloco de cadastro logístico da aba Qualidade. Sem query nova — o cadastro de produtos
    e o mapa de embalagem já estão em cache. Cache 30min, chave por comprador."""
    key = f"qualcad:{comprador or 'todos'}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    res = {"resumo": {"total": 0, "base": 0, "contagem": {}, "rotulos": {}}, "produtos": []}
    try:
        res = core.qualidade_cadastro(_cadastro_produtos(), _embalagem_map(),
                                      _cadastro_fornecedores(), _compradores_map(),
                                      comprador=comprador)
    except Exception as e:
        print(f"[qualidade] cadastro indisponível ({e}).")
    pbi._CACHE.set(key, res, 1800)
    return res


@bp.route("/api/qualidade-cadastro")
def api_qualidade_cadastro():
    """Problemas de CADASTRO logístico sobre a BASE INTEIRA — ver core.qualidade_cadastro.

    ⚠️ Universo diferente do resto da aba Qualidade (que roda sobre o snapshot, recortado por
    filial). É proposital e a tela declara: 72 produtos na base contra 21 no snapshot do
    Atacado. `?comprador_cod=` recorta por comprador do fornecedor."""
    res = _qualidade_cadastro_res(request.args.get("comprador_cod"))
    return jsonify({"ok": True, "gerado_em": date.today().isoformat(), **res})


@bp.route("/api/verbas")
def api_verbas():
    """Verbas de fornecedor (rotina 1801): negociado × aplicado × saldo + tripé com
    compra 12m e lead time — ver core.verbas_fornecedores.
    `?comprador_cod=` e `?fornec=` recortam a base (inclusive resumo, gráfico e "por conta").
    O nome `fornec` acompanha o resto do módulo (export e /api/verbas/detalhe), não `fornec_cod`."""
    res = _verbas_res(_hoje(), comprador=request.args.get("comprador_cod"),
                      fornec=request.args.get("fornec"))
    return jsonify({"ok": True, "gerado_em": date.today().isoformat(),
                    "bi_refresh": pbi.get_dataset_refresh(), **res})


@bp.route("/api/verbas/detalhe")
def api_verbas_detalhe():
    """Drill da aba Verbas: as verbas de UM fornecedor (?fornec=cod), uma a uma."""
    try:
        cod = int(request.args.get("fornec") or 0)
    except ValueError:
        cod = 0
    if not cod:
        return jsonify({"ok": False, "error": "fornec obrigatório"}), 400
    try:
        raw = _verbas_raw(_hoje())
    except Exception as e:
        print(f"[verbas] drill indisponível ({e}).")
        return jsonify({"ok": False, "error": "BI indisponível"}), 503
    det = core.verbas_detalhe(raw["verbas"], raw["aplics"], cod, hoje=_hoje())
    return jsonify({"ok": True, **det})


@bp.route("/api/leadtime/pedidos")
def api_leadtime_pedidos():
    """Drill da aba Lead time: tudo que compõe o número de UM fornecedor (?fornec=cod) —
    pedidos um a um (incl. abertos/negativos), evolução trimestral, faixas e promessa
    (DTPREVENT real × entrada). Reusa o cache raw do /api/leadtime — nenhuma query nova."""
    try:
        cod = int(request.args.get("fornec") or 0)
    except ValueError:
        cod = 0
    if not cod:
        return jsonify({"ok": False, "error": "fornec obrigatório"}), 400
    try:
        raw = _leadtime_raw(_hoje())
    except Exception as e:
        print(f"[leadtime] drill indisponível ({e}).")
        return jsonify({"ok": False, "error": "BI indisponível"}), 503
    det = core.leadtime_detalhe(raw["cab"], raw["entradas"], cod,
                                _cadastro_fornecedores(), _compradores_map(), hoje=_hoje())
    return jsonify({"ok": True, **det})


# ───────────────────────── pesquisa de preço (captura em campo) ─────────────────────────
@bp.route("/pesquisa")
def pagina_pesquisa():
    """Tela de campo da pesquisa de preço — página PRÓPRIA, fora do SPA.

    ⚠️ Não é aba e não entra no menu (decisão 08/2026): o módulo já tem 21 abas, e este é outro
    CONTEXTO — pessoa em pé, uma mão, sinal ruim, um item por vez. As telas do painel pressupõem
    o oposto (sentado, tela larga, tabela de 20 colunas sobre o snapshot inteiro).

    A guarda de acesso vem do `before_request` do blueprint (login + área `compras`), então esta
    função não leva decorador nenhum."""
    return send_from_directory(_PKG_DIR, "pesquisa.html")


@bp.route("/api/busca")
def api_busca():
    """Busca LEVE de produto por código ou descrição — para a tela de campo.

    ⚠️ Existe porque não havia nada assim: a única rota por produto era
    `/api/produto/<cod>`, que exige código exato e ainda roda `_build_produtos()` (snapshot
    inteiro). No SPA a busca é client-side sobre `S.produtosAll`, que o celular não pode carregar.

    Usa só o cadastro (cache 24h) — nenhuma consulta ao BI."""
    q = (request.args.get("q") or "").strip()
    fornec = request.args.get("fornec")

    # `?tipo=fornec` — busca o FORNECEDOR, para o comprador montar o roteiro da visita
    # (pedido do diretor 08/2026: "filtrar por fornecedor, aí traz os itens daquele fornecedor").
    if request.args.get("tipo") == "fornec":
        if len(q) < 2:
            return jsonify({"ok": True, "fornecedores": []})
        ql = q.lower()
        fs = [{"codfornec": cf, "fornecedor": (f.get("FORNECEDOR") or "").strip()}
              for cf, f in _cadastro_fornecedores().items()
              if ql in str(cf) or ql in (f.get("FORNECEDOR") or "").lower()]
        fs.sort(key=lambda x: x["fornecedor"])
        return jsonify({"ok": True, "fornecedores": fs[:20]})

    # sem fornecedor escolhido, exige termo: varrer o cadastro inteiro por nada é caro e inútil
    if not fornec and len(q) < 2:
        return jsonify({"ok": True, "produtos": []})
    prod, emb = _cadastro_produtos(), _embalagem_map()
    fcod = int(core._n(fornec)) if fornec else None
    out = []
    for cod, cad in prod.items():
        if fcod and int(core._n(cad.get("CODFORNEC"))) != fcod:
            continue
        if q and not core.casa_busca({**cad, "CODPROD": cod}, q):
            continue
        qe = core._n((emb.get(cod) or {}).get("qtunit"))
        out.append({"codprod": cod,
                    "descricao": (cad.get("DESCRICAO") or "").strip(),
                    "embalagem": cad.get("EMBALAGEM"),
                    "qtunitcx": qe if qe > 1 else core._n(cad.get("QTUNITCX")) or None})
        # busca livre corta em 30 (tela de polegar); ROTEIRO de fornecedor vai até 200 — a lista
        # dele É o trabalho da visita, e cortar em 30 deixaria o comprador sem saber o que falta
        if len(out) >= (200 if fcod else 30):
            break
    out.sort(key=lambda x: ((x["descricao"] or "").upper(), x["codprod"]))
    lim = 200 if fcod else 30
    return jsonify({"ok": True, "produtos": out, "truncado": len(out) >= lim})


@bp.route("/api/pesquisa-preco", methods=["GET", "POST"])
def api_pesquisa_preco():
    """GET: histórico de um produto. POST: grava uma medição.

    A validação recusa preço <= 0 e produto fora do cadastro — dado de campo entra torto com
    facilidade, e uma linha inválida contamina a comparação sem avisar."""
    if not store.ensure():
        return jsonify({"ok": False, "error": "Postgres indisponível"}), 503
    if request.method == "GET":
        # `?ultimas=1` devolve a última medição de CADA produto — é o que o modal de pedido
        # precisa (ele já tem a lista de itens em mãos e recorta no cliente). A tabela cresce
        # por visita, não por catálogo, então não vale paginar ainda.
        # `?lista=1` — a consulta da tela de campo: o que já foi pesquisado, com autor.
        # Recorte de fornecedor aqui (e não no store): fornecedor vem do CADASTRO do produto.
        if request.args.get("lista"):
            dias = int(core._n(request.args.get("dias")) or 90)
            linhas = _pesquisa_enriquecida(store.pesquisa_lista(dias=dias),
                                           request.args.get("fornec"))
            return jsonify({"ok": True, "medicoes": linhas})
        if request.args.get("ultimas"):
            return jsonify({"ok": True, "ultimas": store.pesquisa_ultima()})
        cod = request.args.get("codprod")
        if not cod:
            return jsonify({"ok": False, "error": "codprod obrigatório"}), 400
        return jsonify({"ok": True, "medicoes": store.pesquisa_do_produto(int(core._n(cod)))})

    d = request.get_json() or {}
    cod = int(core._n(d.get("codprod")))
    preco = core._n(d.get("preco"))
    if not cod:
        return jsonify({"ok": False, "error": "codprod obrigatório"}), 400
    if preco <= 0:
        return jsonify({"ok": False, "error": "preço tem de ser maior que zero"}), 400
    cad = _cadastro_produtos().get(cod)
    if cad is None:
        return jsonify({"ok": False, "error": f"produto {cod} não está no cadastro de revenda"}), 400
    if (d.get("unidade") or "un") not in ("un", "cx"):
        return jsonify({"ok": False, "error": "unidade tem de ser 'un' ou 'cx'"}), 400
    # o fator viaja GRAVADO: o cadastro muda e o passado não pode se reinterpretar
    qe = core._n((_embalagem_map().get(cod) or {}).get("qtunit"))
    d["qtunitcx"] = d.get("qtunitcx") or (qe if qe > 1 else core._n(cad.get("QTUNITCX")) or None)
    d["codprod"] = cod
    d["usuario_id"] = session.get("user_id") or session.get("uid")
    return jsonify({"ok": True, "id": store.pesquisa_add(d)})


def _pesquisa_enriquecida(medicoes, fornec=None):
    """Medições + produto, fornecedor e a comparação com o NOSSO PREÇO DE VENDA.

    Fonte ÚNICA da tela de campo, do drawer 360° e dos exports — três montagens do mesmo número
    é como as telas divergem (o app já tem cicatriz dessa família na aba Fornecedores).

    ⚠️ **A referência é o preço de VENDA, não o custo.** O pedido foi lido errado até 08/2026:
    o comprador vai a atacados CONCORRENTES ver por quanto ELES vendem ("pesquisar o preço que
    eles estão vendendo o mesmo produto que nós vendemos", diretor), para saber se o nosso preço
    está dentro da praça. Comparando com o `CUSTOFIN` a tela errava duas vezes: como o custo é
    sempre menor que o preço de venda, o gap saía enviesado para o VERDE — dizia "estamos bem"
    mesmo quando vendíamos acima do mercado — e o documento que vai ao fornecedor levava o nosso
    custo de aquisição, a única coisa que não se manda a fornecedor.

    ⚠️ **NÃO use o `preco_venda` do produto.** Em `core.construir_produtos` esse campo cai em
    `custofin` quando o item não vendeu em 3 meses (a venda perdida precisa do fallback). Aqui
    ele devolveria o CUSTO rotulado como "nosso preço de venda" — o vazamento de volta, agora
    com etiqueta errada. Sem preço realizado, a coluna sai VAZIA.
    """
    prod, forn = _cadastro_produtos(), _cadastro_fornecedores()
    # Preço de venda = realizado médio dos ÚLTIMOS 3 MESES: o `PCPRODUT[PVENDA]` está vazio nesta
    # base, e o diretor aprovou a régua ("pode pegar a média de preço dos últimos 3 meses").
    # Cache de 6h do `_preco_venda_map` — esta tela não custa consulta nova ao BI.
    try:
        pv_map = _preco_venda_map(_filiais_venda())
    except Exception as e:                                    # noqa: BLE001
        print(f"[pesquisa] preço de venda indisponível ({e}).")
        pv_map = {}
    fcod = int(core._n(fornec)) if fornec else None
    out = []
    for m in medicoes:
        cad = prod.get(m["codprod"]) or {}
        cf = int(core._n(cad.get("CODFORNEC"))) or None
        if fcod and cf != fcod:
            continue
        n = core.normaliza_pesquisa(m["preco"], m.get("unidade"), m.get("qtunitcx"))
        nosso = core._n(pv_map.get(m["codprod"])) or None
        g = core.gap_pesquisa(n["preco_un"], nosso) if n["comparavel"] else {"delta": None, "delta_pct": None}
        out.append({**m,
                    "descricao": (cad.get("DESCRICAO") or f"PRODUTO {m['codprod']}").strip(),
                    "codfornec": cf, "fornecedor": (forn.get(cf) or {}).get("FORNECEDOR"),
                    "preco_venda_unit": core._round(nosso, 4) if nosso else None,
                    "preco_un": n["preco_un"], "comparavel": n["comparavel"],
                    "delta": g["delta"], "delta_pct": g["delta_pct"]})
    return out


@bp.route("/api/evolucao")
def api_evolucao():
    """Serie historica do estoque — a aba Evolucao (ver `estoque/historico.py`).

    ⚠️ **Restrita ao ADM por enquanto** (decisao 08/2026): a aba nasce sem historico nenhum e so
    fica util depois de ~4 semanas de foto. Deixa-la visivel para o comprador antes disso seria
    entregar tela vazia como se fosse produto. O gate e no SERVIDOR, nao so no menu — esconder
    aba no JS nao e controle de acesso.

    Tudo e DERIVADO da foto crua a cada chamada: os ⚙ Parametros viajam na querystring e
    reescrevem a serie inteira (e o que a decisao de guardar ingrediente comprou). O resumo traz
    a VARIACAO na janela, nao so o nivel — "caiu R$ 380 mil desde o inicio" e a pergunta que a
    aba responde; "ha R$ 4,3 mi em estoque" o Cockpit ja respondia."""
    if (session.get("role") or "") != "admin":
        return jsonify({"ok": False, "error": "Aba restrita ao administrador"}), 403
    if not store.ensure():
        return jsonify({"ok": True, "dias": [], "log": [], "resumo": {},
                        "indisponivel": "Postgres indisponivel"})
    unidade = _unidade()
    ini = core._parse_dt(request.args.get("ini"))
    fim = core._parse_dt(request.args.get("fim"))
    comprador = request.args.get("comprador_cod") or None
    fornec = request.args.get("fornec") or None
    # curva/XYZ vêm da FOTO daquele dia, não do cadastro de hoje — é o que permite ver a ruptura
    # da curva A ao longo do tempo sem reclassificar o passado com a régua de agora.
    # Multi-seleção não entra: a barra manda "A,B" e aqui só uma vale; o front recorta 1 por vez.
    curva = (request.args.get("curva") or "").strip().upper()[:1] or None
    xyz = (request.args.get("xyz") or "").strip().upper()[:1] or None
    dias = historico.serie(unidade, ini, fim, comprador, fornec, request.args.to_dict(),
                           curva=curva, xyz=xyz)
    log = historico.dias_com_foto(unidade, ini, fim)
    return jsonify({"ok": True, "unidade": unidade, "dias": dias, "log": log,
                    "filtros": {"comprador": comprador, "fornec": fornec,
                                "curva": curva, "xyz": xyz},
                    "resumo": _resumo_evolucao(dias, log)})


# Metricas com direcao INEQUIVOCA — so elas ganham cor na tela. O valor de estoque fica de fora
# de proposito: estoque caindo pode ser boa gestao OU desabastecimento, e pintar de verde faria
# a aba um dia comemorar uma ruptura. E a ruptura estavel ao lado da queda que prova gestao.
_EVO_DIRECAO = {"valor_parado": "menor_melhor", "n_ruptura": "menor_melhor",
                "pct_ruptura": "menor_melhor",
                # A watchlist cai quando o comprador age (vendeu, devolveu, parou de repor) — e,
                # ao contrario do valor de estoque, ela nao tem leitura ruim quando cai: nenhum
                # item sai daqui por desabastecimento, porque entrar exige cobertura ALTA.
                # ⚠️ E aqui que o piso de VALOR (em vez de um "top 50") se paga: lista de tamanho
                # fixo daria uma serie constante, que nao pode melhorar nem piorar.
                "valor_desacel": "menor_melhor",
                # ⚠️ Ocupação SEM cor, mesma regra do valor de estoque: subir pode ser depósito
                # enchendo (ruim, acaba o espaço) ou giro entrando (bom), e cair pode ser
                # organização (bom) ou desabastecimento (ruim). Pintar de verde ou vermelho seria
                # afirmar uma direção que o número sozinho não tem — quem dá o sentido é ela ao
                # lado do valor de estoque e da ruptura, que é o motivo de estarem na mesma aba.
                "ocupacao_pct": None,
                "pct_ideal": "maior_melhor", "valor_estoque": None}


def _resumo_evolucao(dias, log):
    """Primeiro x ultimo ponto da janela + o estado de maturidade da serie.

    `fotos`/`maturidade` alimentam o aviso honesto da tela enquanto o historico nao enche: a aba
    mostra a foto de hoje e diz quantas faltam, em vez de um "sem dados" que faria o diretor
    achar que a ferramenta nao esta trabalhando."""
    if not dias:
        return {"fotos": len(log), "maturidade": "vazia", "direcao": _EVO_DIRECAO}
    a, b = dias[0], dias[-1]
    # ⚠️ Com UMA foto, `a` e `b` sao o MESMO ponto e todo delta sai zero — a tela dizia
    # "R$ 0,00 (0%) na janela" nos quatro KPIs no dia da 1a foto em producao. Isso se le como
    # "nao mudou nada", quando o certo e "ainda nao da para comparar": afirma uma medicao que
    # nao foi feita. `None` cai no `—` que o front ja sabe renderizar (`evoDelta`), na mesma
    # politica do resto da aba (o aviso de maturidade ja dizia que faltavam 27 dias).
    um_ponto = len(dias) == 1
    var = {}
    for k in _EVO_DIRECAO:
        ini_v, fim_v = a.get(k), b.get(k)
        if um_ponto or ini_v is None or fim_v is None:
            var[k] = None
            continue
        var[k] = {"ini": ini_v, "fim": fim_v, "delta": core._round(fim_v - ini_v, 4),
                  "delta_pct": core._round((fim_v - ini_v) / ini_v * 100, 1) if ini_v else None}
    n = len(dias)
    return {
        "de": a["data"], "ate": b["data"], "fotos": n,
        # 4 semanas e o minimo para a linha dizer algo; 90 dias e onde ela vira tendencia
        "maturidade": "tendencia" if n >= 90 else ("util" if n >= 28 else "enchendo"),
        "faltam_para_util": max(0, 28 - n),
        "variacao": var, "direcao": _EVO_DIRECAO,
    }


@bp.route("/api/validade")
def api_validade():
    produtos, params, filiais = _build_produtos()
    idx = {p["codprod"]: p for p in produtos}
    hoje = _hoje()
    dias = int(params["horizonte_val"])
    _janela = (hoje, hoje + timedelta(days=dias))
    lotes = PS.validade(*_janela, filiais) if _pg() else pbi.run_dax(Q.q_validade(*_janela, filiais))
    fefo = core.validade_fefo(lotes, idx, params, hoje=hoje)
    resumo = {
        "n": len(fefo),
        "valor_risco": core._round(sum(l["valor_risco"] or 0 for l in fefo)),
        "valor_risco_critico": core._round(sum(l["valor_risco"] or 0 for l in fefo if l["classificacao"] == "critico")),
        "critico": sum(1 for l in fefo if l["classificacao"] == "critico"),
        "atencao": sum(1 for l in fefo if l["classificacao"] == "atencao"),
        "planejar": sum(1 for l in fefo if l["classificacao"] == "planejar"),
        "giro_zero": sum(1 for l in fefo if l["risco"] == "giro_zero"),
    }
    return jsonify({"ok": True, "horizonte": dias, "resumo": resumo, "lotes": fefo})


def _venda_liq_mensal(filiais_venda):
    """Venda LÍQUIDA (bruta − devolução) por comprador × mês, p/ o % da aba Vencidos.
    Devolve (por_mes {AnoMes 'YYYY-MM': liq}, por_comprador {cc: liq},
    por_comp_mes {'cc|YYYY-MM': liq}). Mesma fórmula da aba Desempenho
    (líq = [VENDA BRUTA] − [TOTAL DEVOLUCAO]). Venda começa em 2024 no RCA;
    meses anteriores ficam sem % (numerador sem denominador).

    ⚠️ O terceiro mapa (CRUZADO) existe porque a query já traz o grão comprador × mês
    (`q_venda_comprador_mensal_rca` agrupa por CODCOMPRADOR **e** AnoMes) e as duas primeiras
    agregações o descartavam — cada uma colapsando uma dimensão. Sem o cruzado, a aba Vencidos
    com comprador filtrado não tinha denominador: a linha de % sumia do gráfico e o card caía no
    % all-time daquele comprador, **ignorando o seletor de período** sem avisar (07/2026).
    A devolução (`dev_idx`) já mantinha o grão — era só a venda que perdia."""
    ini = date(2024, 1, 1)

    def _am(v):
        s = str(int(v))
        return f"{s[:4]}-{s[4:]}"

    vendas = PS.venda_comprador_mensal(ini, filiais_venda) if _pg() else pbi.run_dax_rca(Q.q_venda_comprador_mensal_rca(ini, filiais_venda))
    devol = PS.devol_comprador_mensal(ini, filiais_venda) if _pg() else pbi.run_dax_rca(Q.q_devol_comprador_mensal_rca(ini, filiais_venda))
    dev_idx = {(int(core._n(r["CODCOMPRADOR"])), _am(r["AnoMes"])): core._n(r.get("dev")) for r in devol}
    por_mes, por_comp, por_comp_mes = {}, {}, {}
    for r in vendas:
        cc = int(core._n(r["CODCOMPRADOR"]))
        mes = _am(r["AnoMes"])
        liq = core._n(r.get("venda")) - dev_idx.get((cc, mes), 0.0)
        por_mes[mes] = por_mes.get(mes, 0.0) + liq
        por_comp[cc] = por_comp.get(cc, 0.0) + liq
        # chave em string: este mapa viaja no JSON, e tupla não sobrevive à serialização
        k = f"{cc}|{mes}"
        por_comp_mes[k] = por_comp_mes.get(k, 0.0) + liq
    return por_mes, por_comp, por_comp_mes


@bp.route("/api/vencidos")
def api_vencidos():
    """Perda por validade (conta 200042) por mês — replica a planilha VENCIDOS do diretor,
    cruza com o estoque atual (item que já venceu e ainda está na casa) e traz o % da perda
    sobre a venda líquida (por mês e por comprador)."""
    produtos, params, filiais = _build_produtos()
    idx = {p["codprod"]: p for p in produtos}
    rows = PS.vencidos(filiais) if _pg() else pbi.run_dax(Q.q_vencidos(filiais))
    venda_mes, venda_comp, venda_comp_mes = _venda_liq_mensal(_filiais_venda())
    res = core.vencidos_por_mes(rows, idx, venda_mes_map=venda_mes, venda_comp_map=venda_comp,
                                venda_comp_mes_map=venda_comp_mes)
    # próximo vencimento do estoque ATUAL dos itens que já venceram e continuam na casa —
    # "já perdi isso; e o que tenho vence quando?". Só p/ os em_estoque (lista curta).
    em = res.get("em_estoque") or []
    cods = [p["codprod"] for p in em if p.get("codprod") is not None]
    if cods:
        pv = PS.prox_venc(cods, _hoje(), filiais) if _pg() else pbi.run_dax(Q.q_prox_venc(cods, _hoje(), filiais))
        pvmap = {int(core._n(r["CODPROD"])): r.get("prox_venc") for r in pv}
        for p in em:  # mesmas refs de res['produtos'] → atualiza os dois
            dtv = pvmap.get(p["codprod"])
            p["prox_venc"] = (str(dtv)[:10] if dtv else None)
    return jsonify({"ok": True, **res})


@bp.route("/api/resumos")
def api_resumos():
    """Painel gerencial do diretor: 2 blocos-resumo (itens a vencer por faixa de validade +
    cobertura de estoque por faixa de dias). Validade busca a janela inteira (não só 30d).

    Filtros (correção 07/2026): antes só o comprador era respeitado — curva/XYZ/fornecedor/depto
    saíam do front e morriam aqui, então o painel não reagia ao filtro de curva. Agora passa pelo
    `_aplicar_filtros_cliente` (o mesmo dos exports), o que também alinha tela e CSV/Excel/PDF.
    EXCEÇÃO: o **orçamento** continua só por comprador — meta (65% da venda líq. 30d) e comprado
    (pedido real do Winthor) não têm quebra por curva ABC; recortar um lado e não o outro daria um
    "% da meta" falso. O front avisa na tela quando há filtro que o orçamento não consegue honrar."""
    produtos, params, filiais = _build_produtos()
    hoje = _hoje()
    comprador = request.args.get("comprador") or "TODOS"
    todos = comprador in ("", "TODOS")
    # respeita o filtro de comprador do topo (painéis recalculam por comprador)
    prod_f = produtos if todos else [p for p in produtos if (p.get("comprador") or "") == comprador]
    # + demais filtros da UI (curva, XYZ, fornecedor, depto, busca…)
    prod_f = _aplicar_filtros_cliente(prod_f)
    idx = {p["codprod"]: p for p in prod_f}
    _jan = (hoje, hoje + timedelta(days=3650))
    lotes = PS.validade(*_jan, filiais) if _pg() else pbi.run_dax(Q.q_validade(*_jan, filiais))
    # recorta os lotes sempre que QUALQUER filtro estiver ativo (antes só quando havia comprador,
    # o que faria a validade divergir dos outros blocos assim que a curva passasse a filtrar)
    if len(prod_f) != len(produtos):
        cods = set(idx)
        lotes = [l for l in lotes if int(core._n(l.get("CODPROD"))) in cods]
    cab = _pedidos_data(filiais, hoje)["cab"]
    venda_comp = _venda_comprador_30d(filiais, _filiais_venda(), hoje)
    orc = core.orcamento_winthor(cab, venda_comp, _compradores_map(), _cadastro_fornecedores(),
                                 _mes_atual(), comprador, pct=0.65, hoje=hoje, meta_override=None,
                                 cnpj_empresa=MULTPEL_EMPRESA["cnpj"])
    # filtros ativos que o ORÇAMENTO não consegue honrar (não existe quebra por curva/XYZ/… na
    # meta nem no pedido do Winthor) → o front mostra o aviso no card em vez de exibir % falso
    _rot = {"curva": "curva", "xyz": "XYZ", "fornec": "fornecedor", "depto": "depto", "busca": "busca"}
    orc_ignora = [lbl for arg, lbl in _rot.items() if request.args.get(arg)]
    return jsonify({
        "ok": True,
        "gerado_em": hoje.isoformat(),
        "n_produtos": len(prod_f),
        "validade": core.resumo_validade(lotes, idx, hoje=hoje),
        "cobertura": core.resumo_cobertura(prod_f),
        # régua do Estoque ideal vem do ⚙ Parâmetros (limiar em dias + meta em %); o clamp da
        # querystring mora no core.regua_estoque_ideal — ver o porquê lá.
        "estoque_ideal": core.resumo_estoque_ideal(prod_f, *core.regua_estoque_ideal(params)),
        "orcamento": orc["resumo"],
        "orcamento_ignora": orc_ignora,
        "ruptura": core.resumo_ruptura(prod_f),
    })


def _resumo_fornecedor(codfornec, itens, forn, extra, lead):
    """Os ESCALARES do 360° do fornecedor, num dict achatado.

    Extraído do `api_fornecedor` quando a ficha exportável nasceu (07/2026): drawer e export
    passaram a ler do MESMO lugar. Duplicar esta conta era garantir que um dia a ficha e a tela
    diriam números diferentes para o mesmo fornecedor — o app já tem cicatriz dessa família."""
    _sug = sum(core._valor_sugerido_compra(p, "valor_sugerido_nf") for p in itens)
    resumo = {
        "codfornec": codfornec,
        "fornecedor": forn.get("FORNECEDOR") or f"FORN {codfornec}",
        "fantasia": forn.get("FANTASIA"), "estado": forn.get("ESTADO"),
        "comprador": (_compradores_map() or {}).get(int(core._n(forn.get("CODCOMPRADOR")))) or None,
        "prazo_entrega": core._n(forn.get("PRAZOENTREGA")) or None,
        "n_produtos": len(itens),
        "estoque": core._round(sum(p.get("valor") or 0 for p in itens)),
        "venda": core._round(sum(p.get("venda") or 0 for p in itens)),
        "lucro": core._round(sum(p.get("lucro") or 0 for p in itens)),
        "n_ruptura": sum(1 for p in itens if (p.get("qtdisp") or 0) <= 0 and (p.get("giro_dia") or 0) > 0),
        "sugestao_nf": core._round(_sug),          # a comprar, régua da NF (c/ impostos)
        # `core.eh_parado`, não a verdade do campo: `novo` é mercadoria recém-chegada, não capital
        # parado. Sem isto o drawer divergia do export desta MESMA aba (que já usava a régua certa,
        # ver `_export_data` view="fornecedores") — dois números para o mesmo fornecedor.
        "valor_parado": core._round(sum(p.get("valor") or 0 for p in itens if core.eh_parado(p))),
        # lead REAL (mediana das entradas >=2d) é o que a aba Lead time mostra; `confiavel`
        # diz se há amostra suficiente — sem isso o drawer exibiria um número frágil como fato
        "lead_real": (lead or {}).get("lead_real"),
        "lead_todos": (lead or {}).get("lead_todos"),
        "lead_n": (lead or {}).get("n"),
        "lead_confiavel": (lead or {}).get("confiavel"),
        "ultima_compra": extra.get("ultima_compra"),
        **{k: extra.get(k) for k in ("n_pedidos", "ciclo_dias", "verba", "verba_campanha",
                                     "venda_yoy", "venda_ant_yoy")},
    }
    resumo["margem"] = core._round(resumo["lucro"] / resumo["venda"] * 100, 1) if resumo["venda"] else None
    return resumo


@bp.route("/api/fornecedor/<int:codfornec>")
def api_fornecedor(codfornec):
    """Drawer 360° do FORNECEDOR — o "igual tem a do produto" pedido pelo diretor (07/2026).

    Reúne o que já existe sobre o fornecedor, tudo de caches vivos: **nenhuma query nova além
    da série mensal** (`_vendas_forn_mensal_map`, 12h de cache, ~6k linhas).

    ⚠️ A série é a régua COMPLETA (agregada no fato por CODFORNEC). Os KPIs seguem o seletor
    "Venda" do topo, como a tabela; a série é sempre 24m — é histórico, não recorte de tela
    (mesma política do Ciclo 12m, ver README)."""
    produtos, params, filiais = _build_produtos()
    hoje = _hoje()
    periodo = request.args.get("venda_periodo", "mes")
    forn = _cadastro_fornecedores().get(codfornec) or {}
    itens = [p for p in produtos if p.get("codfornec") == codfornec]

    meses = core._meses_ate(hoje, 12)
    serie_map = _vendas_forn_mensal_map(hoje, _filiais_venda()).get(codfornec) or {}
    serie = [core._round(core._n(serie_map.get(am)), 2) for am in meses]
    # mesmo mês do ano anterior, para o gráfico responder "quando caiu" e não só "caiu"
    serie_ant = [core._round(core._n(serie_map.get(am - 100)), 2) for am in meses]

    extra = (_forn_extra_map(hoje, periodo, filiais) or {}).get(codfornec) or {}
    lead = next((f for f in _leadtime_res(hoje)["fornecedores"]
                 if f.get("codfornec") == codfornec), None)
    # top produtos do fornecedor por venda no período (o drill natural do drawer)
    top = sorted(itens, key=lambda p: -(p.get("venda") or 0))[:10]
    # pedidos de compra REAIS em aberto (o que já está a caminho deste fornecedor)
    try:
        cab = _pedidos_data(filiais, hoje)["cab"]
        abertos = sorted((c for c in cab
                          if int(core._n(c.get("CODFORNEC"))) == codfornec
                          and core._n(c.get("VLENTREGUE")) <= 0),
                         key=lambda c: str(c.get("DTEMISSAO") or ""), reverse=True)[:8]
        pedidos = [{"numped": int(core._n(c.get("NUMPED"))),
                    "dtemissao": core._parse_dt(c.get("DTEMISSAO")).isoformat()
                    if core._parse_dt(c.get("DTEMISSAO")) else None,
                    "valor": core._round(core._n(c.get("VLTOTAL"))),
                    "dtprevent": core._parse_dt(c.get("DTPREVENT")).isoformat()
                    if core._parse_dt(c.get("DTPREVENT")) else None} for c in abertos]
    except Exception as e:
        print(f"[fornecedor {codfornec}] pedidos indisponíveis ({e}).")
        pedidos = []

    resumo = _resumo_fornecedor(codfornec, itens, forn, extra, lead)
    # Top 3 vendedores do fornecedor no período — segue o seletor "Venda" do topo, como os KPIs
    # (a série é que é sempre 24m). Degrada para [] se o RCA cair: o drawer não some por isso.
    top_vend = _top_vendedores_fornecedor(codfornec, periodo, _filiais_venda(), hoje)
    return jsonify({"ok": True, "fornecedor": resumo, "meses": meses,
                    "top_vendedores": top_vend,
                    "serie": serie, "serie_ant": serie_ant,
                    "top_produtos": [{k: p.get(k) for k in
                                      ("codprod", "descricao", "venda", "lucro", "margem",
                                       "valor", "qtdisp", "cobertura", "curva_abc", "status_abast")}
                                     for p in top],
                    "pedidos_abertos": pedidos})


def _vendedores_nomes():
    """{codusur_str: nome} do mapa de vendedores do Comercial (PCUSUARI, cache 24h).

    ⚠️ NÃO usa `from server import ...`. Em produção o container roda `python server.py`
    (Dockerfile), então aquele arquivo é o módulo **`__main__`** — e o import criaria uma SEGUNDA
    cópia de um módulo de ~8,7 mil linhas, reexecutando-o inteiro (outro Flask app, outra conexão
    Redis) a cada abertura de drawer. Foi o que aconteceu: em produção o mapa vinha vazio e TODO
    vendedor aparecia como "RCA 950".

    O teste nunca pegaria: na suíte o `import server` registra o módulo com esse nome e o import
    funciona. É por isso que a busca é pelo módulo JÁ CARREGADO, nos dois nomes possíveis, em vez
    de importar de novo."""
    import sys
    for nome_mod in ("server", "__main__"):
        mod = sys.modules.get(nome_mod)
        fn = getattr(mod, "_carregar_vendedores_map", None)
        if callable(fn):
            try:
                return {k: (v.get("nome") if isinstance(v, dict) else v)
                        for k, v in (fn() or {}).items()}
            except Exception as e:
                print(f"[vendedores] mapa indisponível ({e}).")
                return {}
    print("[vendedores] mapa de nomes não encontrado em nenhum módulo carregado.")
    return {}


def _vendedores_tecnicos():
    """Códigos que NÃO são vendedores de verdade (bonificação/brinde/transferência).

    Lê do Comercial pelo módulo já carregado — mesma razão do `_vendedores_nomes`: em produção o
    server.py é `__main__`, e um `from server import ...` traria uma segunda cópia do módulo.
    Sem a lista, degrada para vazio: perder o filtro é feio, derrubar o drawer é pior."""
    import sys
    for nome_mod in ("server", "__main__"):
        v = getattr(sys.modules.get(nome_mod), "VENDEDORES_TECNICOS", None)
        if v:
            return set(v)
    return set()


def _top_vendedores_produto(codprod, periodo, filiais_venda, hoje, n=3):
    """Top N vendedores DESTE produto no período: [{codusur, nome, qtd, valor}].

    Pedido do diretor 07/2026 ("trazer o melhor vendedor do item nessa janela"). O uso que
    justifica: num item parado ou com excesso, saber quem já escoou aquilo transforma o drawer de
    diagnóstico em ação — hoje o comprador vê o capital imobilizado e não tem o que fazer com a
    informação.

    Consulta por PRODUTO, sob demanda: o corte transversal (todos os produtos × todos os
    vendedores) daria ~145 mil linhas e estouraria o limite de 100.000 do executeQueries em
    silêncio. Cache de 30min por produto+período.

    Nome vem do mapa de vendedores do COMERCIAL (`_carregar_vendedores_map`, cache 24h, já
    resolvido nos dois modos de fonte) — import tardio porque o server importa este blueprint, e
    no topo isso seria circular. Sem o mapa, cai no código: ranking com número vale mais que
    ranking nenhum."""
    ini, fim = _venda_datas(periodo, hoje)
    key = f"topvend:{codprod}:{periodo}:{_filiais_key(filiais_venda)}:{hoje.isoformat()}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    try:
        rows = (PS.vendedores_do_produto(codprod, ini, fim, filiais_venda) if _pg()
                else pbi.run_dax_rca(Q.q_vendedores_do_produto_rca(codprod, ini, fim, filiais_venda)))
    except Exception as e:
        print(f"[top vendedores {codprod}] indisponível ({e}).")
        return []
    nomes = _vendedores_nomes()
    # Contas TÉCNICAS fora do ranking (08/2026). O diretor viu "2. RCA 999 — 37.440 un · R$ 0,00"
    # no drawer: 999 é conta técnica (bonificação/brinde/transferência), não vendedor. Aparecia
    # como "RCA 999" justamente porque o mapa de nomes do Comercial já a exclui, e o fallback a
    # mostrava mesmo assim. A pergunta da seção é "para quem eu ligo para escoar este item" — e
    # não existe ninguém para ligar. Mesma lista que o Comercial usa em todas as suas telas.
    tecnicos = _vendedores_tecnicos()
    saida = []
    for r in rows:
        cu = core._n(r.get("CODUSUR"))
        if not cu:
            continue
        cu = int(cu)
        if cu in tecnicos:
            continue
        saida.append({"codusur": cu, "nome": nomes.get(str(cu)) or f"RCA {cu}",
                      "qtd": core._round(core._n(r.get("qtd"))),
                      "valor": core._round(core._n(r.get("valor")))})
    # ordena por QUANTIDADE: a pergunta é quem gira volume do item (é o que escoa estoque).
    # O faturamento viaja junto na tela para o comprador ver quando os dois discordam — vender
    # muita unidade barata e vender pouca cara são conversas diferentes.
    saida.sort(key=lambda x: (-(x["qtd"] or 0), -(x["valor"] or 0)))
    saida = saida[:n]
    pbi._CACHE.set(key, saida, 1800)
    return saida


def _top_vendedores_fornecedor(codfornec, periodo, filiais_venda, hoje, n=3):
    """Top N vendedores DESTE fornecedor no período: [{codusur, nome, qtd, valor}].

    Pedido do diretor 08/2026: "trazer aqui no fornecedor o Top 3 vendedores em venda, igual
    traz nos produtos". Mesma mecânica de `_top_vendedores_produto`, com UMA diferença que não
    é detalhe: filtra por `CODFORNEC` no FATO, não pela lista de produtos do cadastro. A lista
    só tem o que é revenda HOJE e perderia item fora de linha — medido no MAGNATECH, 90d:
    R$ 16.017,39 pela lista contra R$ 16.055,89 no fato, sempre subestimando. Ver a docstring
    de `q_vendedores_do_fornecedor_rca`.

    Nome e lista de técnicos vêm de `_vendedores_nomes`/`_vendedores_tecnicos` — reusados, NÃO
    reimplementados. Aquele caminho já carrega a cicatriz do "RCA 950": em produção o server.py
    é `__main__`, e um `from server import ...` traz uma 2ª cópia do módulo e devolve mapa
    vazio. Repetir o import aqui reintroduziria o bug num lugar que os gates de lá não olham."""
    ini, fim = _venda_datas(periodo, hoje)
    key = f"topvendforn:{codfornec}:{periodo}:{_filiais_key(filiais_venda)}:{hoje.isoformat()}"
    hit = pbi._CACHE.get(key)
    if hit is not None:
        return hit
    try:
        rows = (PS.vendedores_do_fornecedor(codfornec, ini, fim, filiais_venda) if _pg()
                else pbi.run_dax_rca(Q.q_vendedores_do_fornecedor_rca(codfornec, ini, fim,
                                                                      filiais_venda)))
    except Exception as e:
        print(f"[top vendedores forn {codfornec}] indisponível ({e}).")
        return []
    nomes = _vendedores_nomes()
    tecnicos = _vendedores_tecnicos()
    saida = []
    for r in rows:
        cu = core._n(r.get("CODUSUR"))
        if not cu:
            continue
        cu = int(cu)
        if cu in tecnicos:
            continue
        saida.append({"codusur": cu, "nome": nomes.get(str(cu)) or f"RCA {cu}",
                      "qtd": core._round(core._n(r.get("qtd"))),
                      "valor": core._round(core._n(r.get("valor")))})
    # ⚠️ Ordena por VALOR aqui, não por quantidade como no produto. O pedido foi "Top 3
    # vendedores em VENDA", e no fornecedor a quantidade soma unidades de produtos diferentes
    # (uma caixa de bobina com um pote de 60ml) — número sem significado físico. No produto a
    # unidade é a mesma para todos, e por isso lá a quantidade manda.
    saida.sort(key=lambda x: (-(x["valor"] or 0), -(x["qtd"] or 0)))
    saida = saida[:n]
    pbi._CACHE.set(key, saida, 1800)
    return saida


@bp.route("/api/produto/<int:codprod>")
def api_produto(codprod):
    produtos, params, filiais = _build_produtos()
    idx = {p["codprod"]: p for p in produtos}
    p = idx.get(codprod)
    lotes_raw = PS.lotes_produto(codprod, filiais) if _pg() else pbi.run_dax(Q.q_lotes_produto(codprod, filiais))
    lotes = core.validade_fefo(lotes_raw, idx, params, hoje=_hoje()) if p else []
    enderecos = []
    if p:
        try:
            enderecos = PS.produto_enderecos(codprod, filiais) if _pg() else pbi.run_dax(Q.q_produto_enderecos(codprod, filiais))
        except Exception as e:
            print(f"[enderecos] WMS indisponível p/ {codprod} ({e}).")
        p = {**p, "plano": core.plano_reposicao(p, params, hoje=_hoje())}
        # série de VENDA dos últimos 12 meses (R$ líquido + unidades) p/ o gráfico do 360°.
        # Buscada aqui (sob demanda) e não no snapshot inteiro — cacheada 12h p/ os próximos.
        meses = core._meses_ate(_hoje(), 12)   # inclui o mês corrente (venda em andamento)
        _mrs, _mcli = _vendas_mensal_rs_map(_hoje(), _filiais_venda())
        rs = (_mrs or {}).get(codprod) or {}
        cli = (_mcli or {}).get(codprod) or {}
        p["serie_mensal_meses"] = meses
        p["serie_mensal_rs"] = [core._round(core._n(rs.get(am)), 2) for am in meses] if rs else None
        # positivação do item por mês. Mês sem venda = 0 clientes (não é dado faltando: ninguém
        # comprou), então preenche com 0 em vez de null — o gráfico precisa da queda desenhada.
        p["serie_mensal_clientes"] = [int(core._n(cli.get(am))) for am in meses] if cli else None
    top_vend = _top_vendedores_produto(codprod, request.args.get("venda_periodo", "mes"),
                                       _filiais_venda(), _hoje()) if p else []
    # histórico de pesquisa de preço do item (drawer). Degrada p/ [] — Postgres fora não pode
    # derrubar o 360° inteiro por causa de um bloco acessório.
    # ⚠️ Enriquecido AQUI, pela mesma função da tela de campo e dos exports. Antes vinha cru e o
    # JS refazia o gap sozinho — duas implementações do mesmo número, que já tinham divergido
    # (o servidor comparava preço com imposto contra mercadoria e o drawer não comparava).
    try:
        _pesq = _pesquisa_enriquecida(store.pesquisa_do_produto(codprod)) if store.ensure() else []
    except Exception as e:                                    # noqa: BLE001
        print(f"[pesquisa] histórico indisponível ({e}).")
        _pesq = []
    return jsonify({"ok": bool(p), "produto": p, "pesquisas": _pesq, "lotes": lotes, "enderecos": enderecos,
                    "top_vendedores": top_vend})


def _rua_conferencia(rua):
    """Lista de conferência de uma rua (com estoque + reservadas vazias) já com descrição."""
    filiais = _filiais_estoque()
    key = f"ruaconf:{_filiais_key(filiais)}:{rua}"
    hit = pbi._CACHE.get(key)
    if hit is None:
        _itens = PS.rua_itens(rua, filiais) if _pg() else pbi.run_dax(Q.q_rua_itens(rua, filiais))
        _vaz = PS.ocupacao_vazias(filiais, rua=rua) if _pg() else pbi.run_dax(Q.q_ocupacao_vazias(filiais, rua=rua))
        hit = core.rua_conferencia(rua, _itens, _vaz)
        cods = sorted({x["codprod"] for x in hit if x.get("codprod") is not None})
        if cods:
            try:
                _dd = PS.desc_de(cods) if _pg() else pbi.run_dax(Q.q_desc_de(cods))
                dm = {int(core._n(r["CODPROD"])): r.get("DESCRICAO") for r in _dd}
            except Exception as e:
                print(f"[rua] descrições indisponíveis ({e}).")
                dm = {}
            for x in hit:
                cp = x.get("codprod")
                x["descricao"] = dm.get(cp) or (f"Produto {cp}" if cp else "— sem produto")
        pbi._CACHE.set(key, hit, 900)
    return hit


@bp.route("/api/rua/<int:rua>")
def api_rua(rua):
    itens = _rua_conferencia(rua)
    return jsonify({"ok": True, "rua": rua, "itens": itens, "n_pos": len(itens),
                    "n_itens": len({x["codprod"] for x in itens if x.get("codprod")}),
                    "n_vazias": sum(1 for x in itens if (x.get("situacao") or "").startswith("VAZIA"))})


@bp.route("/api/ocupacao")
def api_ocupacao():
    filiais = _filiais_estoque()
    key = f"ocup:{_filiais_key(filiais)}"
    hit = pbi._CACHE.get(key)
    if hit is None:
        try:
            if _pg():
                hit = core.ocupacao_resumo(PS.ocupacao_kpis(filiais), PS.ocupacao_por_rua(filiais),
                                           PS.ocupacao_por_tipo(filiais), PS.ocupacao_vazias(filiais))
            else:
                hit = core.ocupacao_resumo(pbi.run_dax(Q.q_ocupacao_kpis(filiais)),
                                           pbi.run_dax(Q.q_ocupacao_por_rua(filiais)),
                                           pbi.run_dax(Q.q_ocupacao_por_tipo(filiais)),
                                           pbi.run_dax(Q.q_ocupacao_vazias(filiais)))
            # descrição das vagas reservadas vem do PCPRODUT completo (itens zerados/FL
            # não estão no snapshot nem no cadastro de revenda)
            cods = sorted({v["codprod"] for v in hit.get("vazias", []) if v.get("codprod") is not None})
            if cods:
                try:
                    _dd = PS.desc_de(cods) if _pg() else pbi.run_dax(Q.q_desc_de(cods))
                    dm = {int(core._n(r["CODPROD"])): r.get("DESCRICAO") for r in _dd}
                    for v in hit["vazias"]:
                        if v.get("codprod") is not None and dm.get(v["codprod"]):
                            v["descricao"] = dm[v["codprod"]]
                except Exception as e:
                    print(f"[ocupacao] descrições das vagas indisponíveis ({e}).")
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 502
        pbi._CACHE.set(key, hit, 1800)
    return jsonify({"ok": True, **hit})


@bp.route("/api/plano_reposicao")
def api_plano_reposicao():
    """Plano DRP de todos os produtos com giro>0 e sugestão>0 — alimenta a aba 'Plano reposição'."""
    produtos, params, _ = _build_produtos()
    hoje = _hoje()
    itens = []
    for p in produtos:
        if (p.get("giro_dia") or 0) <= 0 or (p.get("sugestao_compra") or 0) <= 0:
            continue
        plano = core.plano_reposicao(p, params, hoje=hoje)
        if not plano["liberacoes"]:
            continue
        itens.append({
            "codprod": p["codprod"], "descricao": p["descricao"],
            "codfornec": p["codfornec"], "fornecedor": p["fornecedor"],
            "comprador": p.get("comprador"), "qtdisp": p["qtdisp"],
            "cobertura": p["cobertura"], "giro_mes": p["giro_mes"],
            "custo_unit": p["custo_unit"], "lead_efetivo": p["lead_efetivo"],
            "qtunitcx": p.get("qtunitcx"), "giro_fonte": p.get("giro_fonte"),
            "liberacoes": plano["liberacoes"],
        })
    return jsonify({"ok": True, "gerado_em": hoje.isoformat(), "n": len(itens), "itens": itens})


# ───────────────────────── ficha 360° (escalares) ─────────────────────────
# Pedido do diretor 07/2026: "exportar os dois [drawers] e colocar em horizontal no Excel e PDF".
# Só os campos ESCALARES, por decisão dele — as listas do drawer (lotes, endereços, pedidos em
# aberto, top produtos, plano semanal) têm tamanho variável e não cabem numa linha.
#
# Esta spec é FONTE ÚNICA e serve três saídas: a ficha de um item (Excel + PDF), e as colunas
# extras do export da lista. Campo novo entra aqui e aparece nos três — era o jeito de não
# repetir a dívida da aba Fornecedores, que é calculada em dois lugares e exige lembrar dos dois.
# Formato: (seção, chave, rótulo, tipo). A seção só é usada pela grade do PDF.
_FICHA_COLS = {
    "produto": [
        ("Identificação", "codprod", "Cód", "text"),
        ("Identificação", "descricao", "Produto", "text"),
        ("Identificação", "fornecedor", "Fornecedor", "text"),
        ("Identificação", "comprador", "Comprador", "text"),
        ("Identificação", "curva_abc", "ABC", "text"),
        ("Identificação", "xyz", "XYZ", "text"),
        ("Posição de estoque", "qtdisp", "Disponível", "int"),
        ("Posição de estoque", "valor", "Valor em estoque", "money"),
        ("Posição de estoque", "giro_mes", "Giro/mês", "int"),
        ("Posição de estoque", "cobertura", "Cobertura (d)", "int"),
        ("Venda no período", "venda", "Venda", "money"),
        ("Venda no período", "lucro", "Lucro", "money"),
        ("Venda no período", "margem", "Margem", "pct"),
        ("Venda no período", "qtd_vendida", "Qtd vendida", "int"),
        ("Situação", "status_abast", "Abastecimento", "text"),
        ("Situação", "status_ruptura", "Ruptura", "text"),
        ("Situação", "status_parado", "Parado", "text"),
        ("Situação", "dtultent", "Última entrada", "date"),
        ("Situação", "dias_sem_entrada", "Dias s/ entrada", "int"),
        ("Situação", "dtultsaida", "Última saída", "date"),
        ("Situação", "dias_sem_venda", "Dias s/ venda", "int"),
        ("Abastecimento", "lead_efetivo", "Lead (d)", "int"),
        ("Abastecimento", "embalagem_caixa", "Embalagem", "text"),
        ("Abastecimento", "caixa", "Un/caixa", "int"),
        ("Abastecimento", "qtd_ja_pedida", "Já pedido (aberto)", "int"),
        ("Abastecimento", "qt_transicao", "Recebido (pré-entrada)", "int"),
        ("Abastecimento", "estoque_projetado", "Estoque projetado", "int"),
        ("Abastecimento", "cobertura_proj", "Cob. projetada (d)", "int"),
        ("Abastecimento", "est_alvo", "Estoque alvo", "int"),
        ("Abastecimento", "sugestao_compra", "Sugestão (un)", "int"),
        ("Abastecimento", "sugestao_cx", "Sugestão (cx)", "int"),
        # as duas réguas, como em toda tela que diz "quanto vou gastar"
        ("Abastecimento", "valor_sugerido_nf", "A comprar (c/ impostos)", "money"),
        ("Abastecimento", "valor_sugerido_liq", "A comprar (mercadoria)", "money"),
        ("Abastecimento", "status_exec", "Status", "text"),
    ],
    "fornecedor": [
        ("Identificação", "codfornec", "Cód", "text"),
        ("Identificação", "fornecedor", "Fornecedor", "text"),
        ("Identificação", "fantasia", "Fantasia", "text"),
        ("Identificação", "estado", "UF", "text"),
        ("Identificação", "comprador", "Comprador", "text"),
        ("Posição de estoque", "n_produtos", "Itens", "int"),
        ("Posição de estoque", "estoque", "Valor em estoque", "money"),
        ("Posição de estoque", "sugestao_nf", "A comprar (c/ impostos)", "money"),
        ("Posição de estoque", "n_ruptura", "Em ruptura", "int"),
        ("Posição de estoque", "valor_parado", "Capital parado", "money"),
        ("Venda no período", "venda", "Venda", "money"),
        ("Venda no período", "lucro", "Lucro", "money"),
        ("Venda no período", "margem", "Margem", "pct"),
        ("Ritmo de compra", "ciclo_dias", "Ciclo de compra 12m (d)", "int"),
        ("Ritmo de compra", "n_pedidos", "Compras no período", "int"),
        ("Ritmo de compra", "lead_real", "Lead time real (d)", "int"),
        ("Ritmo de compra", "lead_n", "Entradas medidas", "int"),
        # amostra fraca vira COLUNA, não nota de rodapé: o lead de 26d apurado em 2 entradas e o
        # apurado em 40 não valem a mesma decisão, e na planilha não há tooltip para avisar
        ("Ritmo de compra", "lead_confiavel", "Lead confiável", "bool"),
        ("Ritmo de compra", "prazo_entrega", "Prazo cadastrado (d)", "int"),
        ("Ritmo de compra", "ultima_compra", "Última compra", "date"),
        ("Verba", "verba", "Verba negociada", "money"),
        ("Verba", "verba_campanha", "· da qual campanha", "money"),
    ],
}


def _ficha_campos(tipo):
    """Só as chaves da ficha, na ordem — para montar colunas de export de lista."""
    return [c[1] for c in _FICHA_COLS[tipo]]


# Seções que o PDF omite (pedido do diretor 07/2026: "tira posição do estoque e coloca o gráfico
# no final da página"). Some só do PAPEL, onde o espaço é disputado e onde três dos quatro campos
# repetem o bloco Abastecimento logo abaixo (Disponível ≡ Estoque projetado quando não há pedido
# em aberto, Cobertura ≡ Cob. projetada). No Excel continuam: lá a largura não custa nada e a
# planilha é o registro completo.
_FICHA_PDF_OCULTA = {"produto": {"Posição de estoque"}}

# Rótulos encurtados SÓ no PDF (07/2026 → revisto em 08/2026, ficha de fornecedor saindo em 2
# páginas). O que estoura a página não é o número de campos — produto e fornecedor têm 13 linhas
# de tabela cada — e sim rótulo que não cabe na coluna e vira DUAS linhas: a `Table` estica a
# linha inteira pela célula mais alta, e a ficha de fornecedor tinha 8 dessas contra 5 da de
# produto. Encurtar aqui, e não em `_FICHA_COLS`, preserva o rótulo longo no Excel — lá a largura
# não custa nada e a planilha é o registro completo (mesmo princípio do `_FICHA_PDF_OCULTA`).
_FICHA_PDF_ROTULO = {
    "fornecedor": {
        "Ciclo de compra 12m (d)": "Ciclo 12m (d)",
        "Compras no período": "Compras (per.)",
        "Lead time real (d)": "Lead real (d)",
        "Prazo cadastrado (d)": "Prazo cad. (d)",
        "A comprar (c/ impostos)": "A comprar (c/ imp.)",
        "· da qual campanha": "· campanha",
    },
}


# ───────────────────────── export CSV ─────────────────────────
_CSV_COLS = {
    "produtos": ["codprod", "descricao", "fornecedor", "comprador", "curva_abc", "xyz", "abc_xyz",
                 "qtdisp", "qtd_ja_pedida", "qtbloq", "giro_mes", "cobertura", "dias_sem_venda",
                 "valor", "venda", "venda_ano_ant", "crescimento", "lucro", "margem",
                 "status_abast", "status_parado"],
    # ABC-XYZ tinha caído no fallback de "produtos" (21 colunas, a maioria irrelevante aqui).
    # Como o relatório passou a ser enviável por email, ganhou recorte próprio: a classificação
    # e só o que a sustenta (giro, variabilidade via cobertura, valor imobilizado).
    "abcxyz": ["codprod", "descricao", "fornecedor", "comprador", "curva_abc", "xyz", "abc_xyz",
               "qtdisp", "giro_mes", "cobertura", "valor", "venda", "status_abast"],
    "comprasvendas": ["codprod", "descricao", "fornecedor", "curva_abc", "comprador", "valor", "venda",
                      "venda_ano_ant", "crescimento", "lucro", "margem", "giro_mes", "cobertura",
                      "dias_sem_venda"],
    "reposicao": ["codprod", "descricao", "fornecedor", "comprador", "curva_abc", "giro_mes",
                  "qtdisp", "qtd_ja_pedida", "cobertura", "rop", "est_alvo", "sugestao_compra",
                  # as duas réguas: mercadoria (preço que vai na planilha) e NF (consome a meta)
                  "valor_sugerido_liq", "valor_sugerido_nf", "perc_ipi", "trib_fonte", "status_abast"],
    "parado": ["codprod", "descricao", "fornecedor", "comprador", "dtultsaida", "dias_sem_venda", "qtdisp",
               "valor", "cobertura", "parado_faixa"],
    "ruptura": ["codprod", "descricao", "fornecedor", "comprador", "qtdisp", "valor", "cobertura_dias",
                "cobertura_faixa", "qtd_ja_pedida", "giro_mes", "sugestao_compra"],
    "estoque_zero": ["codprod", "descricao", "fornecedor", "comprador", "qtdisp", "dias_sem_venda",
                     # `qt_transicao` = chegou e está em pré-entrada (aguardando liberação). Sem
                     # esta coluna o relatório mostra "estoque 0" sem dizer que a mercadoria já
                     # está no armazém — foi o que levou o comprador a pedir de novo.
                     "qtd_ja_pedida", "qt_transicao", "giro_mes", "sugestao_cx", "status_exec"],
}

# O Excel/CSV da aba Produtos leva TODOS os escalares do 360° (07/2026): o diretor pediu a ficha
# "em horizontal", e no Excel não há limite de largura que justifique escolher. O PDF da lista
# NÃO recebe o mesmo tratamento de propósito — 34 colunas em A4 paisagem dão ~0,8cm cada e viram
# borrão; para o detalhe completo em papel existe a ficha de um item (`/export/ficha/...`).
_CSV_COLS["produtos"] += [c for c in _ficha_campos("produto") if c not in _CSV_COLS["produtos"]]


def _margem_bucket(p):
    """Faixa de margem do produto — espelha o `margemBucket` do estoque.js (mesmos cortes e
    mesmos rótulos). Sem margem (produto sem venda no período) = 'sv'."""
    m = p.get("margem")
    if m is None:
        return "sv"
    return "neg" if m < 0 else ("b0" if m < 10 else ("b10" if m < 20 else ("b20" if m < 30 else "b30")))


def _aplicar_filtros_cliente(produtos, skip=()):
    """Aplica os filtros ativos da UI (mesma lógica do filtered() do front) p/ que os exports
    respeitem o que está na tela. Lê os params da querystring (enviados pelo exportQS).
    `skip` pula filtros específicos (ex.: {"curva"} nas visões por fornecedor, onde a Curva
    filtra pela ABC do FORNECEDOR e não do produto — Opção A)."""
    a = request.args

    def g(k):
        v = a.get(k)
        return v if v not in (None, "") else None

    out = produtos
    cc = g("comprador_cod")
    if cc:
        out = [p for p in out if str(p.get("codcomprador")) == cc]
    if "curva" not in skip and g("curva"):
        _cv = {x for x in g("curva").split(",") if x}   # multi-seleção de curva (A/B/C)
        out = [p for p in out if p.get("curva_abc") in _cv]
    if g("xyz"):
        _xy = {x for x in g("xyz").split(",") if x}   # multi-seleção de XYZ (X/Y/Z)
        out = [p for p in out if p.get("xyz") in _xy]
    if g("fornec"):
        out = [p for p in out if str(p.get("codfornec")) == g("fornec")]
    if g("depto"):
        out = [p for p in out if str(p.get("codepto")) == g("depto")]
    if g("abast"):
        _ab = {v for v in g("abast").split(",") if v}
        out = [p for p in out if p.get("status_abast") in _ab]
    # ── filtros do Explorador de produtos (aba Análise→Produtos) ──
    # ⚠️ Estes três existiam SÓ no front: o PDF/Excel saía com o universo inteiro enquanto a tela
    # mostrava 116 itens (reportado pelo diretor 07/2026). Espelham `filtered()` do estoque.js.
    if g("margem"):
        _mg = {v for v in g("margem").split(",") if v}
        out = [p for p in out if _margem_bucket(p) in _mg]
    if g("cob_max") is not None:
        try:
            _cm = float(g("cob_max"))
            # cobertura_dias oficial: inclui ruptura (0d) e exclui sem-giro (9999) — igual à tela
            out = [p for p in out if p.get("cobertura_dias") is not None and p["cobertura_dias"] <= _cm]
        except (TypeError, ValueError):
            pass
    if g("sem_ped") in ("1", "true", "sim"):
        out = [p for p in out if (p.get("qtd_ja_pedida") or 0) <= 0]
    # watchlist "Em desaceleração" (drill do card do Cockpit). Filtro COMPOSTO — janela sem venda
    # + cobertura + piso de valor — e é por isso que ele é um parâmetro próprio em vez de uma
    # combinação de `cob_max`/`par_faixa`: nenhum filtro existente expressa "cobertura MÍNIMA", e
    # montá-lo com as peças de hoje daria tela e export divergindo na primeira mudança de régua.
    if g("desacel") in ("1", "true", "sim"):
        _pr = core.merge_params(a.to_dict())
        out = [p for p in out if core.em_desaceleracao(p, _pr)]
    bs = g("busca")
    if bs:
        bs = bs.lower()
        out = [p for p in out if bs in str(p.get("codprod")) or bs in (p.get("descricao") or "").lower()]
    # filtros específicos de aba
    if g("ez_status"):
        out = [p for p in out if p.get("status_exec") == g("ez_status")]
    if g("cob_faixa"):
        _cf = {x for x in g("cob_faixa").split(",") if x}   # multi-seleção de faixas
        out = [p for p in out if p.get("cobertura_faixa") in _cf]
    if g("cob_sub") == "semgiro":
        out = [p for p in out if p.get("sem_giro")]
    elif g("cob_sub") == "excesso":
        out = [p for p in out if p.get("excesso_real")]
    cp = g("cob_ped")
    if cp == "com":
        out = [p for p in out if (p.get("qtd_ja_pedida") or 0) > 0]
    elif cp == "sem":
        out = [p for p in out if (p.get("qtd_ja_pedida") or 0) <= 0]
    if g("par_classe"):
        out = [p for p in out if p.get("status_parado") == g("par_classe")]
    return out


# checagens de qualidade da base (espelham QUAL_CHK do front)
_QUAL_CHECKS = [
    ("sem_custo",       "Sem custo",           lambda p: (p.get("custo_unit") or 0) <= 0),
    ("sem_fornecedor",  "Sem fornecedor",      lambda p: p.get("codfornec") is None),
    ("sem_comprador",   "Sem comprador",       lambda p: p.get("codcomprador") is None),
    ("sem_giro",        "Sem giro c/ estoque", lambda p: (p.get("giro_dia") or 0) <= 0 and (p.get("qtdisp") or 0) > 0),
    ("estoque_negativo","Estoque negativo",    lambda p: (p.get("qtdisp") or 0) < 0),
]


def _vazias_list(filiais):
    """Posições ocupadas sem estoque (reservadas), já com descrição — p/ export da tabela."""
    out = []
    _vaz = PS.ocupacao_vazias(filiais) if _pg() else pbi.run_dax(Q.q_ocupacao_vazias(filiais))
    for r in _vaz:
        cp = r.get("codprod")
        out.append({
            "endereco": "R%d·P%d·N%d·A%d" % (int(core._n(r.get("rua"))), int(core._n(r.get("predio"))),
                                             int(core._n(r.get("nivel"))), int(core._n(r.get("apto")))),
            "tipo": core.TIPO_WMS.get(r.get("tipo"), r.get("tipo") or "—"),
            "codprod": int(core._n(cp)) if cp is not None else None,
        })
    cods = sorted({x["codprod"] for x in out if x.get("codprod") is not None})
    if cods:
        try:
            _dd = PS.desc_de(cods) if _pg() else pbi.run_dax(Q.q_desc_de(cods))
            dm = {int(core._n(r["CODPROD"])): r.get("DESCRICAO") for r in _dd}
        except Exception:
            dm = {}
        for x in out:
            cp = x.get("codprod")
            x["descricao"] = dm.get(cp) or (f"Produto {cp}" if cp else "— sem produto")
    out.sort(key=lambda x: (x["codprod"] is None, x["endereco"]))
    return out


def _export_data(view):
    """Devolve (cols, linhas) para a view, reaproveitado por CSV e XLSX. Respeita os filtros da UI."""
    if view == "validade":
        produtos, params, filiais = _build_produtos()
        idx = {p["codprod"]: p for p in produtos}
        hoje = _hoje()
        cods = {p["codprod"] for p in _aplicar_filtros_cliente(produtos)}
        _jv = (hoje, hoje + timedelta(days=int(params["horizonte_val"])))
        lotes = PS.validade(*_jv, filiais) if _pg() else pbi.run_dax(Q.q_validade(*_jv, filiais))
        linhas = [l for l in core.validade_fefo(lotes, idx, params, hoje=hoje) if l["codprod"] in cods]
        _vd = request.args.get("val_dias")
        if _vd:
            try:
                linhas = [l for l in linhas if l["dias_para_vencer"] <= int(_vd)]
            except ValueError:
                pass
        # faixa de validade clicada no gráfico/cards (mesmo range da tela) — senão o export sai com tudo
        _flo, _fhi = request.args.get("val_faixa_lo"), request.args.get("val_faixa_hi")
        if _flo and _fhi:
            try:
                lo_i, hi_i = int(float(_flo)), int(float(_fhi))
                linhas = [l for l in linhas if lo_i <= l["dias_para_vencer"] <= hi_i]
            except ValueError:
                pass
        cols = ["codprod", "descricao", "curva_abc", "xyz", "comprador", "fornecedor", "numlote", "dtval",
                "dias_para_vencer", "qt", "saldo_proj", "valor_risco", "classificacao", "risco"]
    elif view == "fornecedores":
        produtos, params, _fil = _build_produtos()
        # Opção A: a Curva filtra pela ABC do FORNECEDOR (pula a curva do produto e filtra o
        # resultado agregado por curva_abc do fornecedor) — igual à tela.
        # mesmo `extra` da tela (ciclo + verba), senão o Excel/PDF sairia sem as colunas novas
        # curva do UNIVERSO (todos os produtos, antes de qualquer filtro) — senão o export
        # repetiria o bug que o diretor viu na tela: filtrar um fornecedor fazia ele virar C,
        # porque o Pareto era refeito sobre a lista recortada
        linhas = core.fornecedores(
            _aplicar_filtros_cliente(produtos, skip={"curva"}), params,
            extra=_forn_extra_map(_hoje(), request.args.get("venda_periodo", "mes"), _fil),
            curva_map=core.curva_abc_fornecedores(produtos, params))
        _cv = request.args.get("curva")
        if _cv:
            _cvset = {x for x in _cv.split(",") if x}
            linhas = [r for r in linhas if r.get("curva_abc") in _cvset]
        fc = request.args.get("forn_classe")
        if fc:
            linhas = [r for r in linhas if r.get("classificacao") == fc]
        # Campos que até 07/2026 só o drawer 360° mostrava: a comprar, em ruptura, capital parado
        # e o lead real. Agregados sobre os MESMOS produtos filtrados que alimentaram o
        # core.fornecedores (nunca sobre o universo inteiro — somar recortes diferentes nos dois
        # lados é como nasceu o bug do Cresc. AA). `ultima_compra`, ciclo e verba já vinham do
        # `_extra_fornecedor`. O lead sai do cache da aba Lead time: nenhuma query nova.
        _pf = _aplicar_filtros_cliente(produtos, skip={"curva"})
        _agg = {}
        for _p in _pf:
            _cf = _p.get("codfornec")
            if _cf is None:
                continue
            _a = _agg.setdefault(_cf, {"sugestao_nf": 0.0, "n_ruptura": 0, "valor_parado": 0.0})
            _a["sugestao_nf"] += core._valor_sugerido_compra(_p, "valor_sugerido_nf")
            if (_p.get("qtdisp") or 0) <= 0 and (_p.get("giro_dia") or 0) > 0:
                _a["n_ruptura"] += 1
            if core.eh_parado(_p):        # fonte única — `novo` não é capital parado
                _a["valor_parado"] += (_p.get("valor") or 0)
        try:
            _lead = {f.get("codfornec"): f for f in _leadtime_res(_hoje()).get("fornecedores", [])}
        except Exception as e:      # lead indisponível não pode derrubar o export inteiro
            print(f"[export fornecedores] lead time indisponível ({e}).")
            _lead = {}
        for r in linhas:
            _a = _agg.get(r.get("codfornec")) or {}
            _l = _lead.get(r.get("codfornec")) or {}
            r["sugestao_nf"] = core._round(_a.get("sugestao_nf") or 0)
            r["n_ruptura"] = _a.get("n_ruptura") or 0
            r["valor_parado"] = core._round(_a.get("valor_parado") or 0)
            r["lead_real"], r["lead_n"] = _l.get("lead_real"), _l.get("n")
            r["lead_confiavel"] = _l.get("confiavel")
        cols = ["codfornec", "fornecedor", "curva_abc", "comprador", "n_produtos", "valor", "giro", "cobertura",
                "venda_ano_ant", "crescimento",
                "venda", "lucro", "margem", "perc_venda", "perc_estoque", "indice", "classificacao",
                # os do 360° do fornecedor
                "sugestao_nf", "n_ruptura", "valor_parado", "n_pedidos", "ciclo_dias",
                "lead_real", "lead_n", "lead_confiavel", "ultima_compra", "verba", "verba_campanha"]
    elif view == "compradores":
        produtos, _, _ = _build_produtos()
        linhas = core.por_comprador(_aplicar_filtros_cliente(produtos))
        cols = ["codcomprador", "comprador", "n_produtos", "estoque", "venda", "lucro",
                "margem", "n_ruptura", "valor_parado", "sugestao_valor"]
    elif view == "ruptura_comprador":
        produtos, _, _ = _build_produtos()
        linhas = core.ruptura_por_comprador(_aplicar_filtros_cliente(produtos))
        cols = ["codcomprador", "comprador", "n_produtos", "n_ruptura", "pct_ruptura",
                "n_sem_pedido", "pct_sem_pedido", "venda_perdida", "custo_reposicao"]
    elif view == "desempenho":
        linhas = _desempenho_data(request.args.get("venda_periodo", "mes"), _hoje(), _filiais_venda())["compradores"]
        # única view do catálogo que não passa por `_aplicar_filtros_cliente` (o grão já é o
        # comprador, não o produto) — sem isto o email do comprador vinculado sairia com a
        # tabela de TODOS os compradores, expondo o desempenho dos colegas.
        cc = request.args.get("comprador_cod")
        if cc:
            linhas = [l for l in linhas if str(l.get("codcomprador")) == cc]
        cols = ["ranking", "comprador", "fornecedores", "clientes_pos", "venda_liquida",
                "lucro_bruto", "margem", "devolucao", "part_receita", "part_lucro",
                "yoy", "yoy_lucro", "status_lucro"]
    elif view == "leadtime":
        # espelha a tela: filtros globais de comprador e fornecedor + mín. de pedidos (lt_min)
        linhas = [dict(l, situacao=("ok" if l.get("confiavel") else "sem lead confiável"))
                  for l in _leadtime_res(_hoje())["fornecedores"]]
        cc = request.args.get("comprador_cod")
        if cc:
            linhas = [l for l in linhas if str(l.get("codcomprador")) == cc]
        fn = request.args.get("fornec")
        if fn:
            linhas = [l for l in linhas if str(l.get("codfornec")) == fn]
        try:
            lt_min = int(request.args.get("lt_min") or 0)
        except ValueError:
            lt_min = 0
        if lt_min:
            linhas = [l for l in linhas if (l.get("n") or 0) >= lt_min]
        cols = ["codfornec", "fornecedor", "comprador", "n", "na_hora", "pct_na_hora",
                "lead_todos", "lead_real", "prazo_manual", "delta", "situacao"]
    elif view == "verbas":
        # espelha a tela: filtros globais de comprador/fornecedor. O recorte é o MESMO do
        # endpoint (feito no core) — refiltrar aqui era uma 2ª implementação do mesmo conceito,
        # livre para divergir da tela sem ninguém perceber.
        linhas = _verbas_res(_hoje(), comprador=request.args.get("comprador_cod"),
                             fornec=request.args.get("fornec"))["fornecedores"]
        cols = ["codfornec", "fornecedor", "comprador", "n_verbas", "negociado", "aplicado",
                "saldo", "idade_saldo", "compra_12m", "pct_vc", "lead_real"]
    elif view == "vencidos":
        # perda por validade (conta 200042). Não passa por _aplicar_filtros_cliente: o grão é
        # item-da-nota (evento histórico), não produto — os filtros de estoque não se aplicam.
        # Espelha a tela: filtro de comprador (topo) + mês clicado no gráfico/tabela.
        produtos, _, filiais = _build_produtos()
        idx = {p["codprod"]: p for p in produtos}
        _venc = PS.vencidos(filiais) if _pg() else pbi.run_dax(Q.q_vencidos(filiais))
        linhas = core.vencidos_por_mes(_venc, idx)["itens"]
        cc = request.args.get("comprador_cod")
        if cc:
            linhas = [l for l in linhas if str(l.get("codcomprador")) == cc]
        fn = request.args.get("fornec")
        if fn:
            linhas = [l for l in linhas if str(l.get("codfornec")) == fn]
        vm = request.args.get("ven_mes")
        if vm:
            linhas = [l for l in linhas if l.get("mes") == vm]
        vp = request.args.get("ven_per")  # 2026 | 12m (tudo = sem filtro)
        if vp == "2026":
            linhas = [l for l in linhas if (l.get("mes") or "").startswith("2026")]
        elif vp == "12m":
            meses_ok = sorted({l.get("mes") for l in linhas if l.get("mes")})[-12:]
            linhas = [l for l in linhas if l.get("mes") in meses_ok]
        cols = ["dtsaida", "mes", "numnota", "codfornec", "fornecedor", "codprod", "descricao",
                "qt", "punit", "total", "comprador", "codfilial", "qtdisp"]
    elif view == "conferencia":
        # relatório de conferência de uma rua (ordem de caminhada).
        linhas = [dict(x) for x in _rua_conferencia(int(request.args.get("rua") or 0))]
        _tp = request.args.get("tipo")           # Picking | Pulmão (filtro da tela)
        if _tp:
            linhas = [l for l in linhas if l.get("tipo") == _tp]
        cols = ["endereco", "tipo", "codprod", "descricao", "qt", "dtval", "situacao"]
    elif view == "vazias":
        # posições ocupadas sem estoque (reservadas). Respeita o filtro Picking/Pulmão.
        linhas = _vazias_list(_filiais_estoque())
        _tp = request.args.get("tipo")
        if _tp:
            linhas = [l for l in linhas if l.get("tipo") == _tp]
        cols = ["endereco", "tipo", "codprod", "descricao"]
    elif view == "qualidade_cadastro":
        # base INTEIRA (nao o snapshot) — ver core.qualidade_cadastro. `cat` filtra 1 categoria.
        _q = _qualidade_cadastro_res(request.args.get("comprador_cod"))
        linhas = _q["produtos"]
        cat = request.args.get("cat")
        if cat:
            linhas = [l for l in linhas if cat in l.get("categorias", [])]
        cols = ["codprod", "descricao", "fornecedor", "comprador", "un_por_cx",
                "peso_un_kg", "volume_un_m3", "caixa_kg", "caixa_m3", "problemas"]
    elif view == "pesquisa":
        # Documento que VAI PARA O FORNECEDOR: leva o NOSSO PREÇO DE VENDA junto do pesquisado
        # ("pro fornecedor, poderia mandar nosso preço atual e o preço pesquisado", diretor).
        # ⚠️ Até 08/2026 a coluna era o `custo_unit` (CUSTOFIN) — mandava o nosso custo de
        # aquisição para quem negocia conosco. Preço de venda é público (está na gôndola e na
        # tabela); custo de compra é a única coisa que não se manda a fornecedor.
        dias = int(core._n(request.args.get("dias")) or 90)
        linhas = _pesquisa_enriquecida(store.pesquisa_lista(dias=dias) if store.ensure() else [],
                                       request.args.get("fornec"))
        cols = ["data_pesquisa", "codprod", "descricao", "fornecedor", "preco", "unidade",
                "preco_un", "preco_venda_unit", "delta", "delta_pct", "origem", "usuario", "obs"]
    elif view == "qualidade":
        # produtos com cadastro/saldo inconsistente. cat opcional filtra 1 categoria.
        produtos = _aplicar_filtros_cliente(_build_produtos()[0])
        cat = request.args.get("cat")
        linhas = []
        for p in produtos:
            probs = [(k, lbl) for k, lbl, fn in _QUAL_CHECKS if fn(p)]
            if not probs or (cat and cat not in {k for k, _ in probs}):
                continue
            linhas.append({**p, "problemas": " · ".join(lbl for _, lbl in probs)})
        cols = ["codprod", "descricao", "fornecedor", "comprador", "qtdisp", "custo_unit", "giro_mes", "problemas"]
    else:
        produtos, _, _ = _build_produtos()
        produtos = _aplicar_filtros_cliente(produtos)
        cols = _CSV_COLS.get(view, _CSV_COLS["produtos"])
        if view == "reposicao":
            # agrupado por fornecedor e, dentro dele, o item mais caro primeiro. O PDF reordena os
            # GRUPOS pelo total (`group_valor`); esta ordenação é o que o CSV/XLSX (sem grupo) usa.
            linhas = sorted((p for p in produtos
                             if (p["sugestao_compra"] or 0) > 0 and (p["giro_dia"] or 0) > 0),
                            key=lambda p: ((p.get("fornecedor") or "").upper(),
                                           -(p.get("valor_sugerido_nf") or 0)))
        elif view == "parado":
            # universo do parado = parado_faixa != None (≥15d, com estoque); filtro opcional por faixa
            # (mesmo critério da tela, p/ o export bater com o que aparece). Agrupa por fornecedor.
            _pf = request.args.get("par_faixa")
            faixas_sel = {x for x in _pf.split(",") if x} if _pf else None
            linhas = sorted((p for p in produtos if p.get("parado_faixa") and (faixas_sel is None or p["parado_faixa"] in faixas_sel)),
                            key=lambda p: ((p.get("fornecedor") or "").upper(), p.get("codprod") or 0))
        elif view == "ruptura":
            # cobertura de estoque por faixa (base inteira, métrica da planilha) — maior valor 1º
            linhas = sorted(produtos, key=lambda p: -(p.get("valor") or 0))
        elif view == "estoque_zero":
            linhas = [p for p in produtos if (p.get("qtdisp") or 0) <= 0]
        else:
            linhas = produtos
    return cols, linhas


def _ficha_dados(tipo, cod):
    """(spec, dict achatado, título) da ficha 360° de UM produto ou fornecedor.

    Lê exatamente as mesmas fontes do drawer — `_build_produtos` para o produto e
    `_resumo_fornecedor` para o fornecedor — e por isso responde aos filtros do topo
    (unidade/período) que viajam na querystring. Ficha que ignora o filtro sairia divergindo da
    tela que a originou, que é a armadilha do `exportQS()` já documentada no README."""
    produtos, params, filiais = _build_produtos()
    if tipo == "produto":
        p = next((x for x in produtos if x.get("codprod") == cod), None)
        if not p:
            return None, None, None
        return _FICHA_COLS["produto"], p, f"{p.get('codprod')} · {p.get('descricao') or ''}"
    itens = [p for p in produtos if p.get("codfornec") == cod]
    forn = _cadastro_fornecedores().get(cod) or {}
    if not itens and not forn:
        return None, None, None
    hoje = _hoje()
    extra = (_forn_extra_map(hoje, request.args.get("venda_periodo", "mes"), filiais) or {}).get(cod) or {}
    try:
        lead = next((f for f in _leadtime_res(hoje)["fornecedores"] if f.get("codfornec") == cod), None)
    except Exception as e:
        print(f"[ficha fornecedor {cod}] lead indisponível ({e}).")
        lead = None
    r = _resumo_fornecedor(cod, itens, forn, extra, lead)
    return _FICHA_COLS["fornecedor"], r, f"{r.get('codfornec')} · {r.get('fornecedor') or ''}"


@bp.route("/api/export/ficha/<tipo>/<int:cod>.xlsx")
def api_export_ficha_xlsx(tipo, cod):
    """Ficha em UMA linha horizontal: rótulos no cabeçalho, valores embaixo.

    Formato escolhido pelo diretor ("colocar em horizontal"). O ganho de colar várias fichas numa
    planilha só aparece assim — em pé, cada ficha seria um bloco novo que não empilha."""
    from openpyxl import Workbook
    if tipo not in _FICHA_COLS:
        return jsonify({"ok": False, "erro": "tipo inválido"}), 404
    spec, dados, titulo = _ficha_dados(tipo, cod)
    if not spec:
        return jsonify({"ok": False, "erro": f"{tipo} {cod} sem posição"}), 404
    wb = Workbook(); ws = wb.active; ws.title = tipo[:31]
    ws.append([c[2] for c in spec])
    ws.append([_fmt_xlsx(dados.get(c[1]), c[3]) for c in spec])
    for i, c in enumerate(spec, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, min(38, len(c[2]) + 4))
    ws.freeze_panes = "A2"
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return Response(bio.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="ficha_{tipo}_{cod}.xlsx"'})


@bp.route("/api/export/ficha/<tipo>/<int:cod>.pdf", methods=["GET", "POST"])
def api_export_ficha_pdf(tipo, cod):
    """O POST carrega o GRÁFICO já desenhado na tela (`grafico`, PNG em data-URL).

    Pedido do diretor 07/2026: "coloca o gráfico no final da página". Capturar o canvas do
    Chart.js em vez de redesenhar no reportlab põe no papel EXATAMENTE a curva que ele estava
    olhando — e evita manter duas implementações do mesmo gráfico, que divergiriam na primeira
    vez que alguém mexesse numa. O GET continua valendo e sai sem gráfico (é o caminho do email
    e de quem chama a URL direto)."""
    if tipo not in _FICHA_COLS:
        return jsonify({"ok": False, "erro": "tipo inválido"}), 404
    spec, dados, titulo = _ficha_dados(tipo, cod)
    if not spec:
        return jsonify({"ok": False, "erro": f"{tipo} {cod} sem posição"}), 404
    grafico = None
    if request.method == "POST":
        grafico = (request.get_json(silent=True) or {}).get("grafico")
    pdf = _gerar_pdf_ficha(tipo, spec, dados, titulo, grafico=grafico)
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="ficha_{tipo}_{cod}.pdf"'})


def _img_do_data_url(data_url, largura_cm, cm, altura_max_cm=None):
    """Converte o PNG capturado do canvas num Image do reportlab, preservando a proporção.

    ⚠️ Limita por LARGURA **e** ALTURA. Dimensionar só pela largura foi o bug da 1ª versão: a
    calibração usou um PNG 2,8:1 e o canvas real do Chart.js é bem mais quadrado, então a mesma
    largura virava uma imagem alta demais, que não cabia na sobra da página — o reportlab
    empurrava o gráfico para a folha 2 e deixava metade da 1ª em branco. A proporção da captura
    depende do tamanho do drawer e do devicePixelRatio de quem clicou, ou seja, NÃO é constante:
    o encaixe tem de ser calculado, não presumido.

    Blindado de propósito: o payload vem do navegador e um data-URL torto (recorte errado, canvas
    vazio, base64 truncado) não pode derrubar a geração — o PDF sai sem o gráfico, que é o que
    havia antes de existir esta funcionalidade."""
    import base64
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Image
    try:
        if not data_url or "," not in str(data_url):
            return None
        cabeca, b64 = str(data_url).split(",", 1)
        if "image/png" not in cabeca:
            return None
        # teto de 6MB no base64: imagem de dashboard não passa de algumas centenas de KB, e sem
        # limite um payload gigante viraria memória do processo (1 réplica serve todo mundo)
        if len(b64) > 6_000_000:
            print("[ficha] gráfico recusado: payload acima do teto.")
            return None
        bruto = io.BytesIO(base64.b64decode(b64))
        larg_px, alt_px = ImageReader(bruto).getSize()
        if not larg_px or not alt_px:
            return None
        bruto.seek(0)
        larg = largura_cm * cm
        alt = larg * alt_px / larg_px
        if altura_max_cm and alt > altura_max_cm * cm:      # alto demais → encolhe pela altura
            alt = altura_max_cm * cm
            larg = alt * larg_px / alt_px
        return Image(bruto, width=larg, height=alt)
    except Exception as e:
        print(f"[ficha] gráfico ignorado ({e}).")
        return None


def _fmt_xlsx(v, kind):
    """No Excel o número sai NÚMERO (para somar/filtrar); só data e booleano viram texto legível.
    Formatar tudo como string era o caminho fácil e entregaria uma planilha em que nada calcula."""
    if kind == "bool":
        return "" if v is None else ("sim" if v else "não")
    if v is None:
        return ""
    if kind == "date":
        s = str(v)[:10].split("-")
        return f"{s[2]}/{s[1]}/{s[0]}" if len(s) == 3 else str(v)
    return v


@bp.route("/api/export/<view>.csv")
def api_export_csv(view):
    cols, linhas = _export_data(view)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\n")
    w.writerow(cols)
    for r in linhas:
        w.writerow([r.get(c, "") for c in cols])
    return Response(
        "﻿" + buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="estoque_{view}.csv"'},
    )


@bp.route("/api/export/<view>.xlsx")
def api_export_xlsx(view):
    from openpyxl import Workbook
    cols, linhas = _export_data(view)
    wb = Workbook(); ws = wb.active; ws.title = view[:31]
    ws.append([c.upper() for c in cols])
    for r in linhas:
        ws.append([r.get(c, "") for c in cols])
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="estoque_{view}.xlsx"'},
    )


# ───────────────────────── export PDF (mesma estética do Multpel HTML) ─────────────
# colunas principais por tela: (chave, rótulo, tipo[, maxlen])
_PDF_COLS = {
    "conferencia": [("endereco", "Endereço", "text"), ("tipo", "Tipo", "text"), ("codprod", "Cód", "text"),
                    ("descricao", "Produto", "text", 38), ("qt", "Qtd (sist.)", "int"),
                    ("dtval", "Validade", "text"), ("situacao", "Situação", "text")],
    "vazias": [("endereco", "Endereço", "text"), ("tipo", "Tipo", "text"), ("codprod", "Cód", "text"),
               ("descricao", "Produto que reservou a vaga", "text", 46)],
    "qualidade_cadastro": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 34),
                  ("fornecedor", "Fornecedor", "text", 20), ("comprador", "Comprador", "text", 14),
                  ("un_por_cx", "Un/cx", "int"), ("peso_un_kg", "Peso un.", "num"),
                  ("volume_un_m3", "Vol. un.", "num"), ("caixa_kg", "Caixa kg", "num"),
                  ("caixa_m3", "Caixa m³", "num"), ("problemas", "Problemas", "text", 24)],
    "qualidade": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 36), ("fornecedor", "Fornecedor", "text", 22),
                  ("comprador", "Comprador", "text", 16), ("qtdisp", "Estoque", "int"), ("custo_unit", "Custo", "money"),
                  ("giro_mes", "Giro/mês", "int"), ("problemas", "Problemas", "text", 34)],
    "produtos": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 40), ("fornecedor", "Fornecedor", "text", 26),
                 ("curva_abc", "ABC", "text"), ("qtdisp", "Disp.", "int"), ("qtd_ja_pedida", "Já ped.", "int"),
                 ("giro_mes", "Giro/mês", "int"), ("cobertura", "Cob.(d)", "int"), ("valor", "Valor", "money"),
                 ("status_abast", "Abast.", "text")],
    "abcxyz": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 38), ("fornecedor", "Fornecedor", "text", 24),
               ("curva_abc", "ABC", "text"), ("xyz", "XYZ", "text"), ("abc_xyz", "Classe", "text"),
               ("qtdisp", "Disp.", "int"), ("giro_mes", "Giro/mês", "int"), ("cobertura", "Cob.(d)", "int"),
               ("valor", "Valor", "money")],
    "comprasvendas": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 40), ("fornecedor", "Fornecedor", "text", 26),
                      ("valor", "Estoque", "money"), ("venda", "Venda", "money"),
                      ("crescimento", "Cresc. AA", "pct"), ("lucro", "Lucro", "money"),
                      ("margem", "Margem", "pct"), ("cobertura", "Cob.(d)", "int")],
    # espelha o padrão da aba Análise→Produtos (ABC/Já ped./Cob.) + o VALOR a comprar, que é o
    # critério de ordenação pedido. Sai com impostos: é a régua do Orçamento (PCPEDIDO[VLTOTAL]).
    "reposicao": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 36), ("fornecedor", "Fornecedor", "text", 22),
                  ("curva_abc", "ABC", "text"), ("qtdisp", "Disp.", "int"), ("qtd_ja_pedida", "Já ped.", "int"),
                  ("giro_mes", "Giro/mês", "int"), ("cobertura", "Cob.(d)", "int"),
                  ("sugestao_compra", "Sugerido", "int"), ("valor_sugerido_nf", "A comprar", "money")],
    "parado": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 38), ("fornecedor", "Fornecedor", "text", 24),
               ("dtultsaida", "Últ. venda", "date"), ("dias_sem_venda", "Dias s/v", "int"), ("qtdisp", "Disp.", "int"),
               ("valor", "Valor", "money"), ("parado_faixa", "Faixa", "text")],
    "ruptura": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 36), ("fornecedor", "Fornecedor", "text", 22),
                ("qtdisp", "Disp.", "int"), ("valor", "Valor", "money"), ("cobertura_dias", "Cob.(d)", "int"),
                ("cobertura_faixa", "Faixa", "text"), ("qtd_ja_pedida", "Já ped.", "int"), ("giro_mes", "Giro/mês", "int"),
                ("sugestao_compra", "Sugerido", "int")],
    "estoque_zero": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 40), ("fornecedor", "Fornecedor", "text", 26),
                     ("qtdisp", "Estoque", "int"), ("dias_sem_venda", "Dias s/ venda", "int"), ("qtd_ja_pedida", "Já ped.", "int"),
                     ("giro_mes", "Giro/mês", "int"), ("sugestao_cx", "Sug.(cx)", "int"), ("status_exec", "Status", "text")],
    "validade": [("codprod", "Cód", "text"), ("descricao", "Produto", "text", 34), ("curva_abc", "ABC", "text"), ("xyz", "XYZ", "text"), ("comprador", "Comprador", "text", 18),
                 ("fornecedor", "Fornecedor", "text", 22),
                 ("dtval", "Validade", "date"), ("dias_para_vencer", "Dias", "int"), ("qt", "Qtd", "int"),
                 ("valor_risco", "Valor risco", "money"), ("classificacao", "Classe", "text")],
    "fornecedores": [("codfornec", "Cód", "text"), ("fornecedor", "Fornecedor", "text", 34), ("n_produtos", "Itens", "int"),
                     ("valor", "Estoque", "money"), ("venda", "Venda", "money"),
                     ("crescimento", "Cresc. AA", "pct"), ("margem", "Margem", "pct"),
                     ("n_pedidos", "Compras (nº)", "int"), ("ciclo_dias", "Ciclo 12m (d)", "num"),
                     ("lucro", "Lucro bruto", "money"), ("verba", "Verba", "money"),
                     ("lucro_verba", "Lucro c/ verba", "money"), ("margem_verba", "Margem c/ verba", "pct"),
                     ("indice", "Índice", "num"), ("classificacao", "Classe", "text")],
    "compradores": [("codcomprador", "Cód", "text"), ("comprador", "Comprador", "text", 30), ("n_produtos", "Itens", "int"),
                    ("estoque", "Estoque", "money"), ("venda", "Venda", "money"), ("lucro", "Lucro", "money"),
                    ("margem", "Margem", "pct")],
    "ruptura_comprador": [("comprador", "Comprador", "text", 30), ("n_produtos", "Produtos", "int"),
                          ("n_ruptura", "Em ruptura", "int"), ("pct_ruptura", "% Rupt.", "num"),
                          ("n_sem_pedido", "Sem pedido", "int"), ("pct_sem_pedido", "% s/ ped.", "num"),
                          ("venda_perdida", "Venda perdida/mês", "money"),
                          ("custo_reposicao", "Sugestão de compra", "money")],
    "desempenho": [("ranking", "#", "int"), ("comprador", "Comprador", "text", 28),
                   ("clientes_pos", "Positivação", "int"), ("venda_liquida", "Venda líq.", "money"),
                   ("lucro_bruto", "Lucro bruto", "money"), ("margem", "Margem", "pct"),
                   ("devolucao", "Devolução", "money"), ("part_lucro", "% Lucro", "num"),
                   ("yoy", "AA Venda", "pct"), ("yoy_lucro", "AA Lucro", "pct")],
    "leadtime": [("codfornec", "Cód", "text"), ("fornecedor", "Fornecedor", "text", 32),
                 ("comprador", "Comprador", "text", 18), ("n", "Pedidos", "int"),
                 ("na_hora", "Na hora", "int"), ("pct_na_hora", "% na hora", "pct"),
                 ("lead_todos", "Lead médio todos (d)", "num"), ("lead_real", "Lead real (d)", "num"),
                 ("prazo_manual", "Prazo manual (d)", "int"), ("delta", "Δ (d)", "num"),
                 ("situacao", "Situação", "text")],
    "verbas": [("codfornec", "Cód", "text"), ("fornecedor", "Fornecedor", "text", 30),
               ("comprador", "Comprador", "text", 16), ("n_verbas", "Verbas 12m", "int"),
               ("negociado", "Negociado 12m", "money"), ("aplicado", "Aplicado", "money"),
               ("saldo", "Saldo aberto", "money"), ("idade_saldo", "Idade (d)", "int"),
               ("compra_12m", "Compra 12m", "money"), ("pct_vc", "% Verba/Compra", "pct"),
               ("lead_real", "Lead (d)", "num")],
    # espelha a planilha VENCIDOS do diretor (+ qtdisp: o que ainda está na casa)
    "vencidos": [("dtsaida", "Data", "date"), ("numnota", "Nota", "text"),
                 ("fornecedor", "Fornecedor", "text", 24), ("codprod", "Cód", "text"),
                 ("descricao", "Produto", "text", 34), ("qt", "Qtd", "int"),
                 ("punit", "P.unit.", "money"), ("total", "Total", "money"),
                 ("comprador", "Comprador", "text", 18), ("qtdisp", "Em estoque", "int")],
}
_PDF_COLS["pesquisa"] = [
    ("data_pesquisa", "Data", "text"), ("codprod", "Cód", "text"),
    ("descricao", "Produto", "text", 38), ("preco", "Pesquisado", "money"),
    ("unidade", "Un", "text"), ("preco_venda_unit", "Nosso preço", "money"),
    ("delta_pct", "Dif.", "dec"), ("origem", "Onde pesquisou", "text", 18),
    ("usuario", "Quem pesquisou", "text", 16)]
_PDF_TITULO = {"pesquisa": "Pesquisa de preço", "produtos": "Produtos", "comprasvendas": "Compras × Vendas", "reposicao": "Reposição",
               "parado": "Estoque parado", "ruptura": "Cobertura de estoque", "validade": "Validade / FEFO",
               "fornecedores": "Fornecedores", "compradores": "Compradores", "estoque_zero": "Estoque zerado",
               "ruptura_comprador": "Ruptura por comprador", "desempenho": "Desempenho comercial",
               "vencidos": "Vencidos — perda por validade",
               "leadtime": "Lead time por fornecedor",
               "verbas": "Verbas por fornecedor"}


def _fmt_pdf(v, kind, maxlen=None):
    # `bool` antes do teste de vazio: False é informação ("lead NÃO confiável"), não ausência —
    # cair no "—" apagaria justamente o aviso que a coluna existe para dar.
    if kind == "bool":
        return "—" if v is None else ("sim" if v else "não")
    if v is None or v == "":
        return "—"
    try:
        if kind == "money":
            return ("R$ %s" % f"{float(v):,.2f}").replace(",", "X").replace(".", ",").replace("X", ".")
        if kind == "int":
            return f"{int(round(float(v))):,}".replace(",", ".")
        if kind == "num":
            return f"{float(v):.2f}".replace(".", ",")
        if kind == "pct":
            return f"{float(v):.1f}%"
        if kind == "date":
            s = str(v)[:10].split("-")
            return f"{s[2]}/{s[1]}/{s[0]}" if len(s) == 3 else str(v)
    except (ValueError, TypeError):
        pass
    s = str(v)
    return (s[:maxlen - 1] + "…") if (maxlen and len(s) > maxlen) else s


def _gerar_pdf(view, linhas, group_by=None, group_valor=None, group_rotulo="Estoque"):
    """`group_valor`: campo somado no cabeçalho do grupo — quando informado, os GRUPOS saem
    ordenados por esse total (maior → menor), não em ordem alfabética. É o que o relatório de
    Reposição pede: fornecedor com mais dinheiro a comprar primeiro."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT

    spec = _PDF_COLS.get(view) or _PDF_COLS["produtos"]
    titulo = _PDF_TITULO.get(view, view.capitalize())
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.2 * cm, rightMargin=1.2 * cm, topMargin=1.2 * cm, bottomMargin=1.5 * cm,
                            title=f"Estoque — {titulo}")
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('t', parent=styles['Heading1'], fontSize=14, alignment=TA_LEFT, textColor=colors.HexColor('#0a0e17'))
    sub_style = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#475569'))

    story = [Paragraph(f'<b>JOGA · Estoque</b> — {titulo}', titulo_style),
             Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')} · {len(linhas)} registros", sub_style),
             Spacer(1, 0.3 * cm)]

    header = [c[1] for c in spec]
    data = [header]
    group_rows = []   # índices das linhas-cabeçalho de grupo (fornecedor)

    def _row(r):
        return [_fmt_pdf(r.get(c[0]), c[2], c[3] if len(c) > 3 else None) for c in spec]

    if group_by:
        ncols = len(spec)
        grupos = {}
        for r in linhas:
            grupos.setdefault(str(r.get(group_by) or "—"), []).append(r)
        if group_valor:
            # grupos do MAIOR total para o menor (o comprador ataca o fornecedor mais pesado 1º)
            ordenados = sorted(grupos.items(),
                               key=lambda kv: -sum(core._n(x.get(group_valor)) for x in kv[1]))
        else:
            ordenados = sorted(grupos.items(), key=lambda kv: kv[0].upper())
        for gnome, grupo in ordenados:
            if group_valor:
                # dentro do grupo, item mais caro primeiro — mesmo critério do cabeçalho
                grupo = sorted(grupo, key=lambda r: -core._n(r.get(group_valor)))
                sub = sum(core._n(r.get(group_valor)) for r in grupo)
                gtxt = (f"{gnome}   ·   {len(grupo)} itens   ·   "
                        f"{group_rotulo} {_fmt_pdf(sub, 'money')}")
            else:
                sub_val = sum(core._n(r.get("valor")) for r in grupo)
                sub_ped = sum(core._n(r.get("qtd_ja_pedida")) for r in grupo)
                gtxt = (f"{gnome}   ·   {len(grupo)} itens   ·   Estoque {_fmt_pdf(sub_val, 'money')}"
                        f"   ·   Já pedido {_fmt_pdf(sub_ped, 'int')} un")
            data.append([gtxt] + [""] * (ncols - 1))
            group_rows.append(len(data) - 1)
            for r in grupo:
                data.append(_row(r))
    else:
        for r in linhas:
            data.append(_row(r))

    # larguras proporcionais (pesos por tipo; texto largo p/ descrição/fornecedor)
    def _peso(c):
        if len(c) > 3:
            return c[3] / 7.0
        return {"money": 2.2, "date": 1.8, "pct": 1.3, "int": 1.3, "num": 1.4}.get(c[2], 1.5)
    pesos = [_peso(c) for c in spec]
    usable = landscape(A4)[0] - 2.4 * cm
    soma = sum(pesos) or 1
    col_w = [usable * p / soma for p in pesos]

    estilo = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]
    for i, c in enumerate(spec):
        al = 'RIGHT' if c[2] in ("int", "money", "pct", "num") else ('CENTER' if c[2] == "date" else 'LEFT')
        if al != 'LEFT':
            estilo.append(('ALIGN', (i, 0), (i, -1), al))
    # cabeçalho de cada grupo de fornecedor: mescla a linha, fundo e negrito
    for gi in group_rows:
        estilo += [('SPAN', (0, gi), (-1, gi)),
                   ('BACKGROUND', (0, gi), (-1, gi), colors.HexColor('#dbe4f0')),
                   ('FONTNAME', (0, gi), (-1, gi), 'Helvetica-Bold'),
                   ('TEXTCOLOR', (0, gi), (-1, gi), colors.HexColor('#0f2a5c')),
                   ('ALIGN', (0, gi), (-1, gi), 'LEFT')]
    tbl = Table(data, repeatRows=1, colWidths=col_w)
    tbl.setStyle(TableStyle(estilo))
    story.append(tbl)

    def _rodape(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#94a3b8'))
        canvas.drawRightString(doc.pagesize[0] - 1.2 * cm, 0.8 * cm, f"Página {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_rodape, onLaterPages=_rodape)
    return buf.getvalue()


def _gerar_pdf_ficha(tipo, spec, dados, titulo, pares_por_linha=4, grafico=None):
    """Ficha 360° de UM item, em paisagem, com os pares rótulo/valor lado a lado numa grade.

    Gerador separado do `_gerar_pdf` porque ficha NÃO é tabela: os 34 campos do produto como 34
    colunas em A4 paisagem dariam ~0,8cm cada e sairiam ilegíveis. A grade transpõe a mesma
    informação em ~12 linhas e cabe numa página. Precedente no repo: `_gerar_pdf_pedido`."""
    from xml.sax.saxutils import escape as _x
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT

    rotulo = {"produto": "Ficha do produto", "fornecedor": "Ficha do fornecedor"}.get(tipo, "Ficha")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                            topMargin=1.2 * cm, bottomMargin=1.5 * cm, title=f"Estoque — {rotulo}")
    styles = getSampleStyleSheet()
    t_style = ParagraphStyle('t', parent=styles['Heading1'], fontSize=14, alignment=TA_LEFT,
                             textColor=colors.HexColor('#0a0e17'))
    s_style = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#475569'))
    # descrição de produto tem "&" e aspas — Paragraph interpreta markup, então escapa
    story = [Paragraph(f'<b>JOGA · Estoque</b> — {rotulo}', t_style),
             Paragraph(f"<b>{_x(str(titulo or ''))}</b>", s_style),
             Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}", s_style),
             Spacer(1, 0.35 * cm)]

    # ── quebra de texto: MEDIDA, não por contagem de caracteres ──
    # String crua no Table do reportlab não quebra — transborda por cima da coluna vizinha. Foi o
    # que aconteceu com "COPAPA COMPANHIA PADUANA DE PAPEIS" (valor) e com "A comprar (c/
    # impostos)" colado no número (rótulo). Mas quebrar por limite de caracteres chutado é o outro
    # extremo: com o corte em 14, quase todo rótulo virava duas linhas e a tabela crescia tanto que
    # empurrava o gráfico para a segunda página. Aqui a decisão é por LARGURA REAL do texto na
    # fonte, comparada com a da coluna: quebra só o que de fato não cabe.
    from reportlab.pdfbase.pdfmetrics import stringWidth
    usable = landscape(A4)[0] - 2.4 * cm
    pesos = [1.0, 1.45] * pares_por_linha
    col_w = [usable * p / sum(pesos) for p in pesos]
    PAD = 10          # LEFTPADDING + RIGHTPADDING aplicados no estilo abaixo
    val_style = ParagraphStyle('v', parent=styles['Normal'], fontSize=8, leading=9.5,
                               fontName='Helvetica-Bold', textColor=colors.black)
    rot_style = ParagraphStyle('r', parent=styles['Normal'], fontSize=8, leading=9.5,
                               textColor=colors.HexColor('#475569'))

    def _celula(txt, style, larg):
        s = str(txt)
        fonte = 'Helvetica-Bold' if style is val_style else 'Helvetica'
        return Paragraph(_x(s), style) if stringWidth(s, fonte, 8) > (larg - PAD) else s

    # agrupa os campos por seção, preservando a ordem da spec (a mesma ordem do drawer)
    ocultas = _FICHA_PDF_OCULTA.get(tipo) or set()
    curtos = _FICHA_PDF_ROTULO.get(tipo) or {}
    secoes = []
    for sec, chave, rot, kind in spec:
        if sec in ocultas:
            continue
        if not secoes or secoes[-1][0] != sec:
            secoes.append((sec, []))
        secoes[-1][1].append((_celula(curtos.get(rot, rot), rot_style, col_w[0]),
                              _celula(_fmt_pdf(dados.get(chave), kind), val_style, col_w[1])))

    ncols = pares_por_linha * 2
    data, linhas_secao = [], []
    for sec, campos in secoes:
        linhas_secao.append(len(data))
        data.append([sec] + [""] * (ncols - 1))
        for i in range(0, len(campos), pares_por_linha):
            linha = []
            for rot, val in campos[i:i + pares_por_linha]:
                linha += [rot, val]
            data.append(linha + [""] * (ncols - len(linha)))

    # col_w já foi calculado acima — é ele que decide o que quebra, então não pode ser recalculado
    # aqui (duas fontes da mesma largura acabariam divergindo numa mudança futura).
    estilo = [
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # respiro enxuto (2,5pt, como o PDF do pedido): com 17 linhas, cada ponto de padding custa
        # ~1,2cm de página. Foi o que abriu espaço para um gráfico legível em vez de um de 4cm
        ('TOPPADDING', (0, 0), (-1, -1), 2.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ('LEFTPADDING', (0, 0), (-1, -1), 5), ('RIGHTPADDING', (0, 0), (-1, -1), 5),
    ]
    for i in range(pares_por_linha):
        estilo.append(('TEXTCOLOR', (i * 2, 0), (i * 2, -1), colors.HexColor('#475569')))
        estilo.append(('FONTNAME', (i * 2 + 1, 0), (i * 2 + 1, -1), 'Helvetica-Bold'))
    for gi in linhas_secao:
        estilo += [('SPAN', (0, gi), (-1, gi)),
                   ('BACKGROUND', (0, gi), (-1, gi), colors.HexColor('#1e293b')),
                   ('TEXTCOLOR', (0, gi), (-1, gi), colors.white),
                   ('FONTNAME', (0, gi), (-1, gi), 'Helvetica-Bold'),
                   ('ALIGN', (0, gi), (-1, gi), 'LEFT')]
    tbl = Table(data, colWidths=col_w)
    tbl.setStyle(TableStyle(estilo))
    story.append(tbl)

    # gráfico no FIM da página, como ele pediu. Largura ~metade da faixa útil: o gráfico da tela é
    # largo e baixo, e esticá-lo na página inteira empurraria a ficha para uma segunda folha —
    # o ponto do formato paisagem era caber em uma.
    # Teto de ALTURA é o que garante a página única, não a largura: a proporção da captura varia
    # com o tamanho do drawer e o devicePixelRatio de quem clicou. A largura é só o limite
    # superior; um canvas quadrado encolhe pela altura e sai mais estreito, que é o certo.
    #
    # ⚠️ Esse teto era FIXO em 5,5cm — calibrado na ficha de produto e, portanto, uma aposta de que
    # nenhuma outra ficha ficaria mais alta. A de fornecedor ficou (9,3cm de tabela contra 8,5cm,
    # por ter uma seção a mais e o dobro de rótulos quebrando em duas linhas) e passou a sair em
    # DUAS páginas sempre que o canvas vinha mais quadrado. Agora o teto é a SOBRA REAL da página,
    # medida depois de montar a tabela: qualquer ficha cabe em uma folha, e quem paga a conta é o
    # gráfico encolhendo — não o leitor recebendo meia página em branco e um gráfico órfão atrás.
    rot_graf = Paragraph("<b>Venda dos últimos 12 meses</b>", s_style)
    usada = sum(f.wrap(doc.width, doc.height)[1] for f in story)
    sobra_cm = (doc.height - usada - 0.45 * cm - rot_graf.wrap(doc.width, doc.height)[1]
                - 0.15 * cm) / cm
    # piso de 2,6cm: abaixo disso o gráfico é ilegível e é melhor não fingir que informa.
    alt_max = min(5.5, sobra_cm - 0.15)
    img = _img_do_data_url(grafico, (landscape(A4)[0] - 2.4 * cm) / cm * 0.50, cm,
                           altura_max_cm=alt_max) if alt_max >= 2.6 else None
    if img is not None:
        # KeepTogether: sem isso o título cabe na página 1, a imagem não, e sobra um rótulo órfão
        # no rodapé apontando para um gráfico que está na folha seguinte (foi o que aconteceu).
        from reportlab.platypus import KeepTogether
        story += [Spacer(1, 0.45 * cm),
                  KeepTogether([rot_graf, Spacer(1, 0.15 * cm), img])]

    def _rodape(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#94a3b8'))
        canvas.drawRightString(doc.pagesize[0] - 1.2 * cm, 0.8 * cm, f"Página {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_rodape, onLaterPages=_rodape)
    return buf.getvalue()


@bp.route("/api/export/<view>.pdf")
def api_export_pdf(view):
    _, linhas = _export_data(view)
    # Reposição agrupa igual ao Produtos (pedido do diretor 07/2026), mas ordenada pelo DINHEIRO
    # a comprar (c/ impostos = a régua do Orçamento), não pelo nome do fornecedor.
    if view == "reposicao":
        pdf = _gerar_pdf(view, linhas, group_by="fornecedor",
                         group_valor="valor_sugerido_nf", group_rotulo="A comprar")
    else:
        pdf = _gerar_pdf(view, linhas, group_by="fornecedor" if view == "produtos" else None)
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="estoque_{view}.pdf"'})


# ───────────────────────── orçamento / pedidos (Postgres) ─────────────────────────
@bp.route("/api/orcamento")
def api_orcamento():
    """Orçamento de compras: meta = 65% da venda líq. 30d por comprador; realizado/aberto vêm
    do pedido REAL (Winthor). Pedidos manuais (nossa plataforma) entram à parte, pendentes de
    envio — não somam no realizado (evita dupla contagem quando voltarem do Winthor)."""
    mes = _mes_atual()
    comprador = request.args.get("comprador") or "TODOS"
    filiais = _filiais_estoque()
    hoje = _hoje()
    pct = float(request.args.get("pct") or 0.65)
    pdata = _pedidos_data(filiais, hoje)
    cab = pdata["cab"]
    venda_comp = _venda_comprador_30d(filiais, _filiais_venda(), hoje)
    # Base da meta que VALIA no fechamento do mês passado — a venda de 30d medida naquele dia,
    # não a de hoje. Reconstruir a meta de ontem com a venda de hoje produziria um "estouro"
    # contra uma meta que nunca existiu. Se falhar, o bloco do mês anterior sai vazio (None) em
    # vez de sair errado — e o arraste, sem base, não desconta nada.
    venda_comp_ant = None
    try:
        fim_mes_ant = date(hoje.year, hoje.month, 1) - timedelta(days=1)
        venda_comp_ant = _venda_comprador_30d(filiais, _filiais_venda(), fim_mes_ant)
    except Exception as e:
        print(f"[orcamento] venda do mês anterior indisponível ({e}).")
    # arraste do estouro: OPT-IN (?arrastar=1). Default desligado — ver core.orcamento_winthor.
    arrastar = request.args.get("arrastar") in ("1", "true", "True")
    # meta sempre automática (65% da venda líquida 30d por comprador) — sem override manual
    res = core.orcamento_winthor(cab, venda_comp, _compradores_map(), _cadastro_fornecedores(),
                                 mes, comprador, pct=pct, hoje=hoje, meta_override=None,
                                 cnpj_empresa=MULTPEL_EMPRESA["cnpj"],
                                 venda_comp_ant=venda_comp_ant, arrastar=arrastar)
    # peso + cubagem por pedido: mesma fonte (PCPEDIDO/PCITEM) já carregada → NENHUMA query nova.
    # Só existe p/ pedido em ABERTO (logistica_pedidos ignora o recebido) — que é o que a tabela mostra.
    # `?busca=` recorta o bloco de PEDIDOS EM ABERTO ao produto procurado (pedido do diretor
    # 08/2026: "saber se existe pedido para aquele item, a quantidade e quando foi feito").
    # Os KPIs de orçamento NÃO entram no recorte — ver core.recorta_abertos_por_produto.
    busca = (request.args.get("busca") or "").strip()
    try:
        _log = core.logistica_pedidos(pdata["cab"], pdata["itens"], _cadastro_produtos(),
                                      _embalagem_map(), _compradores_map(), _cadastro_fornecedores(),
                                      hoje=hoje, busca=busca or None)
        _lmap = {p["numped"]: p for p in _log["pedidos"]}
        for _pe in res["pedidos"]:      # 'abertos' referencia os MESMOS dicts
            _l = _lmap.get(_pe["numped"])
            if _l:
                _pe["cubagem_m3"] = _l["cubagem_m3"]
                _pe["peso_kg"] = _l["peso_kg"]
                _pe["sem_cubagem_itens"] = _l["sem_cubagem_itens"]
                _pe["sem_peso_itens"] = _l["sem_peso_itens"]
        if busca:
            res = core.recorta_abertos_por_produto(res, _log["do_produto"])
    except Exception as e:
        print(f"[orcamento] peso/cubagem indisponível ({e}).")
        if busca:
            # ⚠️ Degradar em silêncio aqui devolveria a lista INTEIRA para quem filtrou um
            # produto — o usuário leria "existem 46 pedidos deste item". A tela avisa.
            res["resumo"] = {**res["resumo"], "filtro_produto_falhou": True}
    manuais = store.pedidos_pendentes(mes, comprador) if store.ensure() else []
    return jsonify({"ok": True, "resumo": res["resumo"], "pedidos": res["pedidos"],
                    "abertos": res["abertos"], "por_comprador": res.get("por_comprador", []),
                    "manuais": manuais})


@bp.route("/api/logistica")
def api_logistica():
    """Cubagem/ocupação por pedido em aberto (o que ainda vai chegar). Capacidade do veículo
    e limite de baixa ocupação são parâmetros (cap_m3, baixa_ate)."""
    filiais = _filiais_estoque()
    hoje = _hoje()
    pdata = _pedidos_data(filiais, hoje)
    cap = float(request.args.get("cap_m3") or 60.0)
    baixa = float(request.args.get("baixa_ate") or 0.1)
    res = core.logistica_pedidos(pdata["cab"], pdata["itens"], _cadastro_produtos(),
                                 _embalagem_map(), _compradores_map(), _cadastro_fornecedores(),
                                 hoje=hoje, capacidade_m3=cap, baixa_ate=baixa)
    return jsonify({"ok": True, **res})


@bp.route("/api/orcamento/meta", methods=["POST"])
def api_orcamento_meta():
    d = request.get_json() or {}
    store.orcamento_set(d.get("mes") or date.today().strftime("%Y-%m"),
                        d.get("comprador") or "TODOS", d.get("meta_valor") or 0)
    return jsonify({"ok": True})


@bp.route("/api/pedidos", methods=["POST"])
def api_pedido_add():
    d = request.get_json() or {}
    d.setdefault("mes", d.get("data_pedido", date.today().isoformat())[:7])
    return jsonify({"ok": True, "id": store.pedido_add(d)})


@bp.route("/api/pedidos/<int:pid>", methods=["PUT", "DELETE"])
def api_pedido_edit(pid):
    if request.method == "DELETE":
        store.pedido_delete(pid)
    else:
        store.pedido_update(pid, request.get_json() or {})
    return jsonify({"ok": True})


def _gerar_pdf_pedido(pe, itens=None, forn=None):
    """Documento PDF de UM pedido de compra. Com itens → ordem de compra no estilo do
    relatório 211 do Winthor (logo + emitente + fornecedor + itens c/ IPI, retrato);
    sem itens → cabeçalho/valor (retrato). `forn` = cadastro do fornecedor (PCFORNEC)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.pdfgen import canvas as _rlcanvas
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT
    import math

    def _d(v):
        if not v:
            return "—"
        s = str(v)[:10].split("-")
        return f"{s[2]}/{s[1]}/{s[0]}" if len(s) == 3 else str(v)

    def _m(v):
        try:
            return ("R$ %s" % f"{float(v):,.2f}").replace(",", "X").replace(".", ",").replace("X", ".")
        except (ValueError, TypeError):
            return "—"

    def _e(v):
        return str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _i(v):
        try:
            return f"{int(round(float(v))):,}".replace(",", ".")
        except (ValueError, TypeError):
            return "—"

    def _dec2(v):
        """2 casas no padrão pt-BR — peso e volume do rodapé são conferidos contra o 211,
        que imprime 14.497,64; arredondar para inteiro faria parecer que não bate."""
        try:
            return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except (ValueError, TypeError):
            return "—"

    def _sug(it):
        q = core._n(it.get("qtd")); cx = core._n(it.get("qtunitcx"))
        if cx > 1 and q > 0:
            return f"{int(math.ceil(q / cx))} cx · {int(q)} un"
        return f"{int(q)} un" if q else "—"

    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('t', parent=styles['Heading1'], fontSize=14, alignment=TA_LEFT, textColor=colors.HexColor('#0a0e17'))
    sub_style = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#475569'))
    info_style = ParagraphStyle('i', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#0a0e17'), leading=14)
    buf = io.BytesIO()

    def _rodape(canvas, doc):
        canvas.saveState(); canvas.setFont('Helvetica', 7); canvas.setFillColor(colors.HexColor('#94a3b8'))
        canvas.drawRightString(doc.pagesize[0] - 1.2 * cm, 0.8 * cm, f"Página {doc.page}")
        canvas.restoreState()

    cab = (f"Nº pedido: <b>{_e(pe.get('n_pedido') or '—')}</b> · Data: <b>{_d(pe.get('data_pedido'))}</b> · "
           f"Fornecedor: <b>{_e(pe.get('fornecedor') or '—')}</b> · Comprador: <b>{_e(pe.get('comprador') or '—')}</b><br/>"
           f"Prazo: <b>{(str(pe.get('prazo_dias'))+'d') if pe.get('prazo_dias') else '—'}</b> · "
           f"Vencimento: <b>{_d(pe.get('dt_vencimento'))}</b> · Status: <b>{_e(pe.get('status') or '—')}</b>"
           + (f" · Forma pgto: <b>{_e(pe.get('forma_pgto'))}</b>" if pe.get('forma_pgto') else ""))

    if itens:
        # ordenado por código p/ facilitar a digitação no sistema interno (retrato, economiza folha).
        itens = sorted(itens, key=lambda it: core._n(it.get("codprod")))
        azul = colors.HexColor('#0f2a5c')

        # rodapé com "Página X de Y" (2 passadas: guarda estados e conta o total no save)
        class _NumCanvas(_rlcanvas.Canvas):
            def __init__(self, *a, **k):
                super().__init__(*a, **k); self._saved = []
            def showPage(self):
                self._saved.append(dict(self.__dict__)); self._startPage()
            def save(self):
                n = len(self._saved)
                for st in self._saved:
                    self.__dict__.update(st)
                    self.setFont('Helvetica', 7); self.setFillColor(colors.HexColor('#94a3b8'))
                    self.drawString(1.2 * cm, 0.8 * cm, "JOGA · Estoque — documento interno de compra")
                    self.drawRightString(A4[0] - 1.2 * cm, 0.8 * cm, f"Página {self._pageNumber} de {n}")
                    _rlcanvas.Canvas.showPage(self)
                _rlcanvas.Canvas.save(self)

        tit_bloco = ParagraphStyle('tb', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', textColor=colors.white)
        corpo = ParagraphStyle('cb', parent=styles['Normal'], fontSize=7.6, textColor=colors.HexColor('#0a0e17'), leading=11.5)
        cel_desc = ParagraphStyle('cd', parent=styles['Normal'], fontSize=6.3, leading=7.3, textColor=colors.HexColor('#0a0e17'))
        # Cód.Fab como Paragraph = REDE DE SEGURANÇA: a coluna (3,40cm) já cabe o maior código
        # do cadastro hoje (23 chars), então na prática não quebra; mas se um dia cadastrarem um
        # código maior, ele QUEBRA em 2 linhas em vez de vazar por cima da coluna Qtde.
        # splitLongWords: código não tem espaço, então precisa poder partir no meio da "palavra".
        cel_cod = ParagraphStyle('cf', parent=styles['Normal'], fontSize=6.5, leading=7.6,
                                 splitLongWords=1, textColor=colors.HexColor('#0a0e17'))

        def _bloco(titulo, corpo_html):
            t = Table([[Paragraph(titulo, tit_bloco)], [Paragraph(corpo_html, corpo)]], colWidths=[18.6 * cm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), azul),
                ('BOX', (0, 0), (-1, -1), 0.4, colors.HexColor('#94a3b8')),
                ('LINEBELOW', (0, 0), (0, 0), 0.4, colors.HexColor('#94a3b8')),
                ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            return t

        E, F = MULTPEL_EMPRESA, (forn or {})   # emitente da instância (env EMPRESA_*)
        # cabeçalho: logo do cliente + título/nº do pedido
        head_dir = Paragraph(
            f"<b>Pedido de Compra</b><br/><font size=9>Nº <b>{_e(pe.get('n_pedido') or pe.get('id') or '—')}</b> · "
            f"Emissão <b>{_d(pe.get('data_pedido'))}</b></font><br/>"
            f"<font size=7 color='#64748b'>Gerado em {date.today().strftime('%d/%m/%Y %H:%M')}</font>", titulo_style)
        logo_path = _logo_cliente()   # CLIENTE_LOGO (env) → fallback JOGA
        try:
            head_row = Table([[Image(logo_path, width=2.3 * cm, height=2.3 * cm), head_dir]], colWidths=[2.7 * cm, 15.9 * cm])
        except Exception:
            head_row = Table([[head_dir]], colWidths=[18.6 * cm])
        head_row.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                      ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0)]))

        emp_html = (f"<b>{_e(E['razao'])}</b> &nbsp;·&nbsp; CNPJ {_e(E['cnpj'])} &nbsp;·&nbsp; IE {_e(E['ie'])}<br/>"
                    f"{_e(E['endereco'])} · {_e(E['bairro'])} · CEP {_e(E['cep'])} · {_e(E['cidade'])}/{_e(E['uf'])}<br/>"
                    f"Tel {_e(E['tel'])} &nbsp;·&nbsp; {_e(E['email'])}")
        cidade_f = (f"{_e(F.get('CIDADE'))}/{_e(F.get('ESTADO'))}" if F.get('CIDADE') else '—')
        forn_html = (f"<b>{_i(pe.get('codfornec') or F.get('CODFORNEC'))} · {_e(pe.get('fornecedor') or F.get('FORNECEDOR') or '—')}</b><br/>"
                     f"CNPJ {_e(F.get('CGC') or '—')} &nbsp;·&nbsp; IE {_e(F.get('IE') or '—')}<br/>"
                     f"Nº {_e(F.get('NUMEROEND') or '—')} · Bairro {_e(F.get('BAIRRO') or '—')} · CEP {_e(F.get('CEP') or '—')} · {cidade_f}"
                     + (f" &nbsp;·&nbsp; {_e(F.get('EMAIL'))}" if F.get('EMAIL') else ""))
        ped_html = (f"Comprador: <b>{_e(pe.get('comprador') or '—')}</b> &nbsp;·&nbsp; "
                    f"Prazo pgto: <b>{(str(pe.get('prazo_dias'))+' dias') if pe.get('prazo_dias') else '—'}</b> &nbsp;·&nbsp; "
                    f"Vencimento: <b>{_d(pe.get('dt_vencimento'))}</b> &nbsp;·&nbsp; "
                    f"Forma pgto: <b>{_e(pe.get('forma_pgto') or '—')}</b> &nbsp;·&nbsp; "
                    f"Status: <b>{_e(pe.get('status') or '—')}</b>"
                    + (f"<br/>Obs: {_e(pe.get('obs'))}" if pe.get('obs') else ""))

        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                                topMargin=1.0 * cm, bottomMargin=1.4 * cm, title=f"Pedido {pe.get('n_pedido') or pe.get('id')}")
        story = [head_row, Spacer(1, 0.2 * cm),
                 _bloco("EMITENTE", emp_html), Spacer(1, 0.12 * cm),
                 _bloco("FORNECEDOR", forn_html), Spacer(1, 0.12 * cm),
                 _bloco("DADOS DO PEDIDO", ped_html), Spacer(1, 0.28 * cm)]

        header = ["Cód", "Descrição", "Embalagem", "Un", "Cód.Fab", "Qtde", "Custo un.", "IPI %", "Vlr. Total"]
        data = [header]
        total = 0.0; total_ipi = 0.0; total_st = 0.0
        # ⚠️ Os três totais logísticos são calculados sobre a quantidade em UNIDADES, que é
        # como o pedido é gravado — é assim que o Winthor chega aos números do rodapé do 211
        # (validado no pedido 565848: 14.482,02 / 14.497,64 / 23,50, exatos).
        # A 1ª versão só somava peso quando `un == "CX"`, então item sem fator de caixa
        # ficava fora do total inteiro, além do buraco da fonte.
        total_liq = 0.0; total_bru = 0.0; total_m3 = 0.0; sem_medida = 0
        for it in itens:
            _vl = core._n(it.get("valor"))
            total += _vl
            # IPI/ST da linha: o Winthor calcula assim na importação (é o que produziu a NF de
            # R$ 44.982,01 sobre R$ 39.536,28 de mercadoria no pedido 565684).
            total_ipi += _vl * core._n(it.get("percipi")) / 100.0
            total_st += _vl * core._n(it.get("percst")) / 100.0
            # MESMA conversão da planilha do Winthor (core.item_master) — os dois documentos
            # precisam bater linha a linha. `custo_master` é o preço da CAIXA nas linhas CX,
            # senão "Qtde × Custo un." não fecharia com o "Vlr. Total" impresso ao lado.
            q_master, custo_master, un = core.item_master(it.get("qtd"), it.get("qtunitcx"),
                                                          it.get("custo_unit"),
                                                          it.get("embalagem"))
            qtde = f"{q_master}" if q_master else "—"
            _qun = core._n(it.get("qtd"))
            total_liq += _qun * core._n(it.get("peso_liq_un"))
            total_bru += _qun * core._n(it.get("peso_bruto_un"))
            total_m3 += _qun * core._n(it.get("vol_un"))
            if core._n(it.get("peso_bruto_un")) <= 0:
                sem_medida += 1
            ipi = core._n(it.get("percipi"))
            data.append([_i(it.get("codprod")), Paragraph(_e(str(it.get("descricao") or "")[:52]), cel_desc),
                         _e(it.get("embalagem") or "—"), un,
                         Paragraph(_e(it.get("codfab") or "—"), cel_cod),
                         qtde, _m(custo_master),
                         (f"{ipi:.1f}".replace('.', ',') + "%" if ipi > 0 else "—"), _m(it.get("valor"))])
        # Totalização espelhando o rodapé do relatório 211: mercadoria, impostos e o TOTAL DA NF —
        # que é o número com que o Winthor vai receber o pedido e o que o Orçamento mede.
        data.append(["", "", "", "", "", "", "", "PRODUTOS", _m(total)])
        _linhas_tot = 1
        if total_ipi > 0:
            data.append(["", "", "", "", "", "", "", "IPI", _m(total_ipi)]); _linhas_tot += 1
        if total_st > 0:
            data.append(["", "", "", "", "", "", "", "ST", _m(total_st)]); _linhas_tot += 1
        total_nf = total + total_ipi + total_st
        if _linhas_tot > 1:
            data.append(["", "", "", "", "", "", "", "TOTAL NF", _m(total_nf)]); _linhas_tot += 1
        # Larguras dimensionadas pelo CONTEÚDO REAL (medido com pdfmetrics), não no olho:
        # Cód.Fab chega a 23 chars no cadastro (ex.: "ET40401TRB/SB1,5P20MBAI") e precisa de
        # 3,26cm — com os 1,8cm antigos o texto VAZAVA por cima da coluna Qtde (string em
        # célula do ReportLab não quebra nem corta). Redistribuí a folga das demais colunas;
        # a fonte segue 6,5pt (legibilidade — o pedido é impresso e vai ao fornecedor) e a
        # Descrição ainda GANHOU espaço. IPI % = 1,30cm porque os rótulos de totalização moram
        # nela ("PRODUTOS"/"TOTAL NF" a 6,5pt bold ≈ 1,22cm; o antigo 1,00cm cortava).
        col_w = [1.10 * cm, 6.15 * cm, 1.85 * cm, 0.60 * cm, 3.40 * cm,
                 0.85 * cm, 1.55 * cm, 1.30 * cm, 1.80 * cm]   # soma = 18,60cm (útil do A4)
        tbl = Table(data, repeatRows=1, colWidths=col_w)
        _r0 = len(data) - _linhas_tot        # 1ª linha do bloco de totais
        estilo = [
            ('BACKGROUND', (0, 0), (-1, 0), azul),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6.5),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (_r0 - 1, -1), [colors.white, colors.HexColor('#f4f7fb')]),
            ('ALIGN', (5, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3), ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
            ('FONTNAME', (0, _r0), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, _r0), (-1, -1), colors.HexColor('#eef2f7')),
        ]
        # cada linha de totalização vaza o rótulo pelas colunas 0..6
        # (nome _lin, não _i: `_i` é o helper de formatação de inteiro deste módulo)
        for _lin in range(_r0, len(data)):
            estilo.append(('SPAN', (0, _lin), (6, _lin)))
        # o TOTAL DA NF (última, quando há imposto) é o número que vale — destacado
        estilo.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e2e8f0')))
        tbl.setStyle(TableStyle(estilo))
        story.append(tbl)
        story.append(Spacer(1, 0.2 * cm))
        # Mesmo conjunto de medidas do rodapé do 211, para o comprador conferir documento
        # contra documento sem ter de calcular nada. Deixou de se chamar "estimado": bate com
        # o ERP ao centavo. Item sem cadastro confiável NÃO entra e a tela diz quantos são —
        # total incompleto apresentado como completo foi o defeito original.
        _p = []
        if total_liq > 0:
            _p.append(f"Peso líquido <b>{_dec2(total_liq)} kg</b>")
        if total_bru > 0:
            _p.append(f"Peso bruto <b>{_dec2(total_bru)} kg</b>")
        if total_m3 > 0:
            _p.append(f"Volume <b>{_dec2(total_m3)} m³</b>")
        if sem_medida:
            _p.append(f"<i>{sem_medida} item(ns) sem cadastro de peso — total incompleto</i>")
        _kg = (" &nbsp;·&nbsp; " + " &nbsp;·&nbsp; ".join(_p)) if _p else ""
        # o total que o comprador confere contra o ERP é o da NF; a mercadoria fica ao lado
        # porque é ela que vira preço na planilha de importação.
        _tot = (f"Total da NF: <b>{_m(total_nf)}</b> &nbsp;·&nbsp; mercadoria {_m(total)}"
                if total_nf > total else f"Total do pedido: <b>{_m(total)}</b>")
        story.append(Paragraph(f"{_tot} &nbsp;·&nbsp; {len(itens)} itens{_kg}", info_style))
        doc.build(story, canvasmaker=_NumCanvas)
    else:
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.6 * cm, rightMargin=1.6 * cm,
                                topMargin=1.4 * cm, bottomMargin=1.5 * cm, title=f"Pedido {pe.get('n_pedido') or pe.get('id')}")
        lbl = ParagraphStyle('l', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#475569'))
        val = ParagraphStyle('v', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#0a0e17'))
        story = [Paragraph('<b>JOGA · Estoque</b> — Pedido de Compra', titulo_style),
                 Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}", sub_style), Spacer(1, 0.5 * cm)]
        linhas = [("Nº do pedido", pe.get('n_pedido') or '—'), ("Data do pedido", _d(pe.get('data_pedido'))),
                  ("Fornecedor", pe.get('fornecedor') or '—'), ("Comprador", pe.get('comprador') or '—'),
                  ("Valor", _m(pe.get('valor'))), ("Prazo de pagamento", f"{pe.get('prazo_dias')} dias" if pe.get('prazo_dias') else '—'),
                  ("Vencimento", _d(pe.get('dt_vencimento'))), ("Forma de pagamento", pe.get('forma_pgto') or '—'),
                  ("Status", pe.get('status') or '—'), ("Observações", pe.get('obs') or '—')]
        data = [[Paragraph(k, lbl), Paragraph(_e(v), val)] for k, v in linhas]
        tbl = Table(data, colWidths=[5 * cm, 11.8 * cm])
        tbl.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl)
        doc.build(story)
    return buf.getvalue()


@bp.route("/api/pedido_itens/<int:numped>")
def api_pedido_itens_winthor(numped):
    """Itens comprados de um pedido REAL do Winthor (PCITEM) — drill do Orçamento."""
    rows = PS.pedido_itens_um(numped) if _pg() else pbi.run_dax(Q.q_pedido_itens_um(numped))
    cad = _cadastro_produtos()
    itens = []
    for r in rows:
        cod = int(core._n(r.get("CODPROD")))
        qp, qe = core._n(r.get("qtped")), core._n(r.get("qtentregue"))
        itens.append({"codprod": cod, "descricao": (cad.get(cod) or {}).get("DESCRICAO") or f"PRODUTO {cod}",
                      "qtped": core._round(qp), "qtentregue": core._round(qe), "aberto": core._round(max(0.0, qp - qe))})
    itens.sort(key=lambda x: x["codprod"])
    return jsonify({"ok": True, "numped": numped, "itens": itens})


@bp.route("/api/pedidos/<int:pid>.xlsx")
def api_pedido_xlsx(pid):
    """Planilha de importação de pedido de compra do Winthor (v26+): 3 colunas, SEM cabeçalho —
    A = código do produto · B = preço · C = quantidade pedida, ambos na **UNIDADE MASTER**
    (pedido do diretor 07/2026: "tem que sair igual ao PDF, tudo em unidade Master").
    Antes saía em unidades e divergia do PDF — o Winthor faz `B × C` literal, então o preço
    converte junto com a quantidade (`core.item_master`), senão o valor do pedido mudaria."""
    if not store.ensure():
        return jsonify({"ok": False, "error": "Postgres indisponível"}), 503
    pe = store.pedido_get(pid)
    if not pe:
        return jsonify({"ok": False, "error": "Pedido não encontrado"}), 404
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "PEDIDO"
    for it in store.pedido_itens(pid):
        cod = int(core._n(it.get("codprod")))
        qtd, preco, _un = core.item_master(it.get("qtd"), it.get("qtunitcx"), it.get("custo_unit"))
        if cod <= 0 or qtd <= 0:
            continue
        ws.append([cod, preco, qtd])   # ordem exata do modelo Winthor: cód · preço · qtd
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    base = re.sub(r'[^A-Za-z0-9 ._-]', '', str(pe.get("fornecedor") or "").strip()) or f"pedido_{pe.get('n_pedido') or pid}"
    return Response(bio.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="winthor_{base}.xlsx"'})


@bp.route("/api/pedidos/<int:pid>.pdf")
def api_pedido_pdf(pid):
    if not store.ensure():
        return jsonify({"ok": False, "error": "Postgres indisponível"}), 503
    pe = store.pedido_get(pid)
    if not pe:
        return jsonify({"ok": False, "error": "Pedido não encontrado"}), 404
    itens = store.pedido_itens(pid)
    # enriquece cada item com cód. de fábrica, % IPI e embalagem (do cadastro) p/ o PDF estilo 211
    if itens:
        cad = _cadastro_produtos()
        emb_map = _embalagem_map()
        for it in itens:
            cod = int(core._n(it.get("codprod")))
            c = cad.get(cod) or {}
            it["codfab"] = c.get("CODFAB")
            # % gravado no pedido (o praticado pelo FORNECEDOR na emissão) manda; o cadastro só
            # atende item sem snapshot (pedido anterior à migração) — ver DDL em store.py.
            it["percipi"] = it["perc_ipi"] if it.get("perc_ipi") is not None else c.get("PERCIPI")
            it["percst"] = it.get("perc_st") or 0
            # embalagem = a da CAIXA (PCEMBALAGEM, igual à tela do Abastecimento); fallback no cadastro
            it["embalagem"] = (emb_map.get(cod) or {}).get("embalagem") or c.get("EMBALAGEM")
            # peso/volume por UNIDADE, do PCPRODUT (core.medidas_unitarias). Até 08/2026 vinha
            # do PCEMBALAGEM[PESOBRUTO], vazio em 75,6% dos produtos: o PDF dizia 6.758 kg
            # onde o Winthor dizia 14.497,64. O fator segue a regra do resto do módulo
            # (PCEMBALAGEM só quando > 1, senão QTUNITCX).
            _cxe = core._n((emb_map.get(cod) or {}).get("qtunit"))
            _fat = (_cxe if _cxe > 1 else core._n(c.get("QTUNITCX"))) or 1
            _md = core.medidas_unitarias(c, _fat)
            it["peso_bruto_un"] = _md["bruto"] if _md["confiavel"] else 0.0
            it["peso_liq_un"] = _md["liq"] if _md["confiavel"] else 0.0
            it["vol_un"] = _md["vol"] if _md["confiavel"] else 0.0
    # dados do fornecedor (PCFORNEC) p/ o bloco do cabeçalho
    forn = None
    if pe.get("codfornec") not in (None, ""):
        forn = _cadastro_fornecedores().get(int(core._n(pe.get("codfornec"))))
    # arquivo com o nome do fornecedor (sanitizado); fallback no nº do pedido
    base = re.sub(r'[^A-Za-z0-9 ._-]', '', str(pe.get("fornecedor") or "").strip()) or f"pedido_{pe.get('n_pedido') or pid}"
    nome = f"{base}.pdf"
    return Response(_gerar_pdf_pedido(pe, itens, forn=forn), mimetype="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{nome}"'})


# ───────────────────────── planos de ação (Postgres) ─────────────────────────
@bp.route("/api/planos")
def api_planos():
    if not store.ensure():
        return jsonify({"ok": True, "planos": {}})
    return jsonify({"ok": True, "planos": store.planos_map(request.args.get("tipo"))})


@bp.route("/api/planos", methods=["POST"])
def api_plano_upsert():
    d = request.get_json() or {}
    if not d.get("chave"):
        return jsonify({"ok": False, "error": "chave obrigatória"}), 400
    store.plano_upsert(d)
    return jsonify({"ok": True})


@bp.route("/api/planos/<path:chave>", methods=["DELETE"])
def api_plano_delete(chave):
    store.plano_delete(chave)
    return jsonify({"ok": True})

