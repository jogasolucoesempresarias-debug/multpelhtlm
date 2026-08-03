"""Gate da ficha 360° exportável (pedido do diretor 07/2026: "exportar os dois [drawers] e
colocar em horizontal no Excel e PDF").

Decisões que estes testes travam:
  · só ESCALARES — as listas do drawer (lotes, endereços, pedidos em aberto, top produtos) têm
    tamanho variável e não cabem numa linha; foi decisão dele, não limitação;
  · HORIZONTAL — rótulos no cabeçalho, valores numa ÚNICA linha embaixo. Em pé, colar duas
    fichas na mesma planilha não empilharia;
  · a spec `_FICHA_COLS` é FONTE ÚNICA da ficha e das colunas extras do export de lista, para
    campo novo não precisar ser lembrado em dois lugares (a dívida da aba Fornecedores);
  · a ficha do fornecedor passa pelo MESMO `_resumo_fornecedor` do drawer — se divergir, a ficha
    e a tela diriam números diferentes do mesmo fornecedor.
"""
import io

import pytest
from openpyxl import load_workbook

from estoque import core, routes


@pytest.fixture
def app_ctx():
    import server  # noqa: F401
    from server import app
    return app


def _prod(cod, forn=1, **kw):
    base = {"codprod": cod, "descricao": f"P{cod}", "codfornec": forn, "venda": 100.0,
            "lucro": 20.0, "valor": 50.0, "qtdisp": 10, "giro_dia": 1.0, "giro_mes": 30,
            "sugestao_cx": 0, "cobertura": 30, "curva_abc": "A", "xyz": "X",
            "status_abast": "ok", "status_parado": None, "comprador": "JOAO",
            "fornecedor": "FORN TESTE", "valor_sugerido_liq": 0.0, "valor_sugerido_nf": 0.0}
    base.update(kw)
    return base


def _mock(monkeypatch, produtos, extra=None, lead=None,
          forn_nome="FORN TESTE", comp_nome="JOAO"):
    monkeypatch.setattr(routes, "_build_produtos",
                        lambda *a, **k: (produtos, core.merge_params({}), ["3"]))
    monkeypatch.setattr(routes, "_cadastro_fornecedores",
                        lambda: {1: {"FORNECEDOR": forn_nome, "ESTADO": "SP",
                                     "CODCOMPRADOR": 7, "PRAZOENTREGA": 12}})
    monkeypatch.setattr(routes, "_compradores_map", lambda: {7: comp_nome})
    monkeypatch.setattr(routes, "_forn_extra_map",
                        lambda *a, **k: {1: extra if extra is not None
                                         else {"ciclo_dias": 30.0, "n_pedidos": 4}})
    monkeypatch.setattr(routes, "_leadtime_res",
                        lambda *a, **k: {"fornecedores": lead if lead is not None else []})
    monkeypatch.setattr(routes, "_filiais_venda", lambda: ["3"])


def _xlsx_rows(resp):
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    return [[c.value for c in row] for row in ws.iter_rows()]


# ─────────────────────── formato horizontal ───────────────────────
@pytest.mark.parametrize("tipo,cod", [("produto", 1), ("fornecedor", 1)])
def test_ficha_sai_horizontal_uma_linha_de_valores(app_ctx, monkeypatch, tipo, cod):
    """Rótulos no cabeçalho e UMA linha de valores — é o "em horizontal" do pedido.
    Se alguém transpuser para o formato vertical (campo|valor por linha), este teste cai."""
    _mock(monkeypatch, [_prod(1)])
    with app_ctx.test_request_context(f"/estoque/api/export/ficha/{tipo}/{cod}.xlsx"):
        resp = routes.api_export_ficha_xlsx(tipo, cod)
    rows = _xlsx_rows(resp)
    assert len(rows) == 2, "ficha tem de ser cabeçalho + 1 linha de valores"
    esperado = [c[2] for c in routes._FICHA_COLS[tipo]]
    assert rows[0] == esperado
    assert len(rows[1]) == len(esperado)


def test_ficha_nao_leva_as_listas_do_drawer(app_ctx, monkeypatch):
    """Só escalares, por decisão do diretor. Lote/endereço/pedido em aberto têm tamanho
    variável — uma linha horizontal não os comporta sem inventar coluna por índice."""
    campos = {c[1] for c in routes._FICHA_COLS["produto"]}
    for proibido in ("lotes", "enderecos", "plano", "liberacoes", "serie_mensal", "serie_giro"):
        assert proibido not in campos
    campos_f = {c[1] for c in routes._FICHA_COLS["fornecedor"]}
    for proibido in ("pedidos_abertos", "top_produtos", "serie", "serie_ant"):
        assert proibido not in campos_f


# ─────────────────────── coerência com o drawer ───────────────────────
def test_ficha_do_fornecedor_bate_com_o_drawer(app_ctx, monkeypatch):
    """Mesma fonte (`_resumo_fornecedor`): ficha e tela não podem divergir."""
    p = _prod(1, sugestao_cx=2, valor_sugerido_liq=100.0, valor_sugerido_nf=115.0)
    _mock(monkeypatch, [p])
    with app_ctx.test_request_context("/estoque/api/fornecedor/1"):
        drawer = routes.api_fornecedor(1).get_json()["fornecedor"]
    _mock(monkeypatch, [p])
    with app_ctx.test_request_context("/estoque/api/export/ficha/fornecedor/1.xlsx"):
        rows = _xlsx_rows(routes.api_export_ficha_xlsx("fornecedor", 1))
    idx = {c[2]: i for i, c in enumerate(routes._FICHA_COLS["fornecedor"])}
    assert rows[1][idx["A comprar (c/ impostos)"]] == drawer["sugestao_nf"] == 115.0
    assert rows[1][idx["Valor em estoque"]] == drawer["estoque"]


def test_a_comprar_da_ficha_sai_na_regua_da_NF(app_ctx, monkeypatch):
    """Coerência com todo o app: o que responde "quanto vou gastar" sai c/ impostos."""
    _mock(monkeypatch, [_prod(1, sugestao_cx=3, valor_sugerido_liq=200.0, valor_sugerido_nf=232.0)])
    with app_ctx.test_request_context("/estoque/api/export/ficha/produto/1.xlsx"):
        rows = _xlsx_rows(routes.api_export_ficha_xlsx("produto", 1))
    idx = {c[2]: i for i, c in enumerate(routes._FICHA_COLS["produto"])}
    assert rows[1][idx["A comprar (c/ impostos)"]] == 232.0
    assert rows[1][idx["A comprar (mercadoria)"]] == 200.0


def test_excel_leva_numero_como_numero(app_ctx, monkeypatch):
    """Planilha em que nada soma é planilha morta — money/int saem numéricos, não string."""
    _mock(monkeypatch, [_prod(1, valor=1234.56)])
    with app_ctx.test_request_context("/estoque/api/export/ficha/produto/1.xlsx"):
        rows = _xlsx_rows(routes.api_export_ficha_xlsx("produto", 1))
    idx = {c[2]: i for i, c in enumerate(routes._FICHA_COLS["produto"])}
    assert rows[1][idx["Valor em estoque"]] == 1234.56
    assert isinstance(rows[1][idx["Valor em estoque"]], (int, float))


# ─────────────────────── o "não" que não pode virar "—" ───────────────────────
def test_lead_nao_confiavel_sai_como_NAO_e_nao_como_vazio():
    """False é informação ("a amostra não sustenta"), não ausência. Cair no travessão apagaria
    exatamente o aviso que a coluna existe para dar — e na planilha não há tooltip que salve."""
    assert routes._fmt_pdf(False, "bool") == "não"
    assert routes._fmt_pdf(True, "bool") == "sim"
    assert routes._fmt_pdf(None, "bool") == "—"


# ─────────────────────── spec como fonte única ───────────────────────
def test_export_da_lista_de_produtos_herda_os_campos_da_ficha():
    """Campo novo entra na spec e aparece nos dois — foi o motivo de existir `_ficha_campos`."""
    cols = routes._CSV_COLS["produtos"]
    assert len(cols) == len(set(cols)), "coluna duplicada no export de produtos"
    for campo in routes._ficha_campos("produto"):
        assert campo in cols


def test_export_de_fornecedores_ganhou_os_campos_do_360(app_ctx, monkeypatch):
    """A comprar / em ruptura / capital parado / lead real só existiam no drawer."""
    _mock(monkeypatch, [_prod(1, sugestao_cx=1, valor_sugerido_nf=115.0, qtdisp=0)],
          lead=[{"codfornec": 1, "lead_real": 9.0, "n": 12, "confiavel": True}])
    with app_ctx.test_request_context("/estoque/api/export/fornecedores.csv"):
        cols, linhas = routes._export_data("fornecedores")
    for c in ("sugestao_nf", "n_ruptura", "valor_parado", "lead_real", "lead_confiavel"):
        assert c in cols
    r = next(x for x in linhas if x["codfornec"] == 1)
    assert r["sugestao_nf"] == 115.0        # item zerado com giro → entra na sugestão
    assert r["n_ruptura"] == 1              # qtdisp 0 e giro > 0
    assert r["lead_real"] == 9.0 and r["lead_confiavel"] is True


def test_export_de_fornecedores_sobrevive_ao_lead_indisponivel(app_ctx, monkeypatch):
    """O lead vem de outro cache (aba Lead time). Se ele cair, o export inteiro não pode cair
    junto — degrada a coluna, não o relatório."""
    def _boom(*a, **k):
        raise RuntimeError("BI fora")
    _mock(monkeypatch, [_prod(1)])
    monkeypatch.setattr(routes, "_leadtime_res", _boom)
    with app_ctx.test_request_context("/estoque/api/export/fornecedores.csv"):
        _cols, linhas = routes._export_data("fornecedores")
    assert linhas and linhas[0]["lead_real"] is None


# ─────────────────────── PDF e bordas ───────────────────────
@pytest.mark.parametrize("tipo", ["produto", "fornecedor"])
def test_ficha_pdf_gera_documento(app_ctx, monkeypatch, tipo):
    _mock(monkeypatch, [_prod(1)])
    with app_ctx.test_request_context(f"/estoque/api/export/ficha/{tipo}/1.pdf"):
        resp = routes.api_export_ficha_pdf(tipo, 1)
    assert resp.data[:4] == b"%PDF"
    assert "ficha_" in resp.headers["Content-Disposition"]


def test_ficha_pdf_escapa_markup_da_descricao(app_ctx, monkeypatch):
    """`Paragraph` do reportlab interpreta markup: descrição com "&" ou "<" derrubaria a
    geração. Nome de produto do Winthor tem os dois."""
    _mock(monkeypatch, [_prod(1, descricao="SABAO & CIA <500ML>")])
    with app_ctx.test_request_context("/estoque/api/export/ficha/produto/1.pdf"):
        resp = routes.api_export_ficha_pdf("produto", 1)
    assert resp.data[:4] == b"%PDF"


def test_item_inexistente_devolve_404(app_ctx, monkeypatch):
    _mock(monkeypatch, [_prod(1)])
    with app_ctx.test_request_context("/estoque/api/export/ficha/produto/999.xlsx"):
        _resp, status = routes.api_export_ficha_xlsx("produto", 999)
    assert status == 404


def test_tipo_invalido_devolve_404(app_ctx, monkeypatch):
    _mock(monkeypatch, [_prod(1)])
    with app_ctx.test_request_context("/estoque/api/export/ficha/cliente/1.xlsx"):
        _resp, status = routes.api_export_ficha_xlsx("cliente", 1)
    assert status == 404


# ─────────────────── lead real na linha da Abastecimento ───────────────────
# Pedido do diretor 07/2026: "não preciso entrar na aba lead time para verificar o tempo de
# entrega do fornecedor, só olhar ali e ajustar no parametro". O lead viaja pelo endpoint LAZY
# (`/api/fornecedores_extra`) e não pelo snapshot — o snapshot é carregado por todas as telas.
#
# ⚠️ As chaves chegam como STRING: jsonify serializa chave numérica de dict como string. O front
# não sente (em JS `obj[1]` e `obj["1"]` são o mesmo slot), mas o teste sente.
def test_fornecedores_extra_leva_o_lead_real(app_ctx, monkeypatch):
    _mock(monkeypatch, [_prod(1)],
          lead=[{"codfornec": 1, "lead_real": 14.0, "n": 28, "confiavel": True}])
    with app_ctx.test_request_context("/estoque/api/fornecedores_extra"):
        extra = routes.api_fornecedores_extra().get_json()["extra"]
    assert extra["1"]["lead_real"] == 14.0
    assert extra["1"]["lead_n"] == 28 and extra["1"]["lead_confiavel"] is True
    # e NÃO pode ter apagado o ciclo/verba que já viajavam por aqui
    assert extra["1"]["ciclo_dias"] == 30.0


def test_amostra_fraca_viaja_junto_do_lead(app_ctx, monkeypatch):
    """O comprador vai MEXER no parâmetro com base neste número. 14d medidos em 2 entradas não
    sustentam a mesma decisão que 14d em 40 — sem a marca, a tela venderia palpite como fato."""
    _mock(monkeypatch, [_prod(1)],
          lead=[{"codfornec": 1, "lead_real": 14.0, "n": 2, "confiavel": False}])
    with app_ctx.test_request_context("/estoque/api/fornecedores_extra"):
        extra = routes.api_fornecedores_extra().get_json()["extra"]
    assert extra["1"]["lead_confiavel"] is False


def test_lead_fora_nao_derruba_ciclo_e_verba(app_ctx, monkeypatch):
    """O lead vem de outro cache. Se ele cair, a aba Fornecedores não pode perder junto as
    colunas que já funcionavam — degrada o campo, não o endpoint."""
    def _boom(*a, **k):
        raise RuntimeError("BI fora")
    _mock(monkeypatch, [_prod(1)])
    monkeypatch.setattr(routes, "_leadtime_res", _boom)
    with app_ctx.test_request_context("/estoque/api/fornecedores_extra"):
        j = routes.api_fornecedores_extra().get_json()
    assert j["ok"] is True
    assert j["extra"]["1"]["ciclo_dias"] == 30.0     # ciclo sobreviveu
    assert "lead_real" not in j["extra"]["1"]        # e o lead simplesmente não veio


# ─────────────────── gráfico no PDF + seção oculta (07/2026) ───────────────────
# "coloca o gráfico no final da página" e "tira posição do estoque". O gráfico vem CAPTURADO do
# canvas do Chart.js (POST) em vez de redesenhado no reportlab: põe no papel a curva que ele
# estava olhando e evita manter duas implementações do mesmo gráfico.
def _png_data_url(w=40, h=20):
    import base64
    from PIL import Image as PILImage
    bio = io.BytesIO()
    PILImage.new("RGB", (w, h), "white").save(bio, format="PNG")
    return "data:image/png;base64," + base64.b64encode(bio.getvalue()).decode()


def test_pdf_da_ficha_embute_o_grafico_enviado(app_ctx, monkeypatch):
    _mock(monkeypatch, [_prod(1)])
    with app_ctx.test_request_context("/estoque/api/export/ficha/produto/1.pdf",
                                      method="POST", json={"grafico": _png_data_url()}):
        com = routes.api_export_ficha_pdf("produto", 1)
    _mock(monkeypatch, [_prod(1)])
    with app_ctx.test_request_context("/estoque/api/export/ficha/produto/1.pdf"):
        sem = routes.api_export_ficha_pdf("produto", 1)
    assert com.data[:4] == b"%PDF" and sem.data[:4] == b"%PDF"
    assert len(com.data) > len(sem.data), "o PDF com gráfico tem de carregar a imagem"


def test_get_continua_valendo_sem_grafico(app_ctx, monkeypatch):
    """O GET é o caminho do email e de quem chama a URL direto — não pode exigir POST."""
    _mock(monkeypatch, [_prod(1)])
    with app_ctx.test_request_context("/estoque/api/export/ficha/fornecedor/1.pdf"):
        resp = routes.api_export_ficha_pdf("fornecedor", 1)
    assert resp.data[:4] == b"%PDF"


@pytest.mark.parametrize("ruim", [
    None, "", "nao-e-data-url", "data:image/svg+xml,<svg/>",      # tipo errado
    "data:image/png;base64,ISTO_NAO_E_BASE64_VALIDO!!!",          # base64 quebrado
    "data:image/png;base64," + "QQ" * 10,                          # decodifica, mas não é PNG
])
def test_grafico_torto_nao_derruba_a_ficha(app_ctx, monkeypatch, ruim):
    """O payload vem do navegador. Canvas vazio, recorte errado ou base64 truncado degradam para
    "PDF sem gráfico" — que é exatamente o que existia antes desta funcionalidade."""
    _mock(monkeypatch, [_prod(1)])
    with app_ctx.test_request_context("/estoque/api/export/ficha/produto/1.pdf",
                                      method="POST", json={"grafico": ruim}):
        resp = routes.api_export_ficha_pdf("produto", 1)
    assert resp.data[:4] == b"%PDF"


def test_grafico_gigante_e_recusado(app_ctx, monkeypatch):
    """Uma réplica serve todo mundo: payload sem teto viraria memória do processo."""
    from reportlab.lib.units import cm
    assert routes._img_do_data_url("data:image/png;base64," + "A" * 6_000_001, 10, cm) is None


def test_pdf_omite_a_secao_pedida_mas_o_excel_mantem(app_ctx, monkeypatch):
    """"tira posição do estoque" vale para o PAPEL, onde o espaço é disputado e três dos quatro
    campos repetem o bloco Abastecimento. No Excel a largura não custa nada e a planilha segue
    sendo o registro completo — por isso a omissão é do PDF, não da spec."""
    assert "Posição de estoque" in routes._FICHA_PDF_OCULTA["produto"]
    # a spec (e portanto o Excel) continua com os campos
    campos = [c[1] for c in routes._FICHA_COLS["produto"]]
    assert "qtdisp" in campos and "valor" in campos and "giro_mes" in campos
    _mock(monkeypatch, [_prod(1)])
    with app_ctx.test_request_context("/estoque/api/export/ficha/produto/1.xlsx"):
        rows = _xlsx_rows(routes.api_export_ficha_xlsx("produto", 1))
    assert "Valor em estoque" in rows[0]


@pytest.mark.parametrize("canvas", [
    (900, 320),    # gráfico largo
    (760, 350),    # proporção do drawer
    (800, 420),    # tela retina (devicePixelRatio 2)
    (600, 560),    # quase quadrado
    (500, 900),    # em pé (drawer estreito) — o pior caso de altura
])
def test_ficha_com_grafico_cabe_em_UMA_pagina(app_ctx, monkeypatch, canvas):
    """O formato paisagem existe para caber numa folha.

    A 1ª versão dimensionava o gráfico só pela LARGURA e foi calibrada contra um PNG 2,8:1. O
    canvas real do Chart.js é mais quadrado — mesma largura, imagem muito mais alta — e a ficha
    saía com o gráfico sozinho na página 2 e metade da 1ª em branco. A proporção depende do
    tamanho do drawer e do devicePixelRatio de quem clicou, ou seja NÃO é constante: por isso o
    caso de teste varre proporções em vez de fixar uma, e o encaixe limita altura E largura.

    ⚠️ Os NOMES aqui são longos de propósito. Este gate já existia quando o diretor reportou a
    ficha de FORNECEDOR saindo em 2 páginas (08/2026) — e passava, porque a fixture usava
    "FORN TESTE"/"JOAO". Nome curto não quebra em duas linhas, a tabela fica ~1cm mais baixa e o
    gráfico cabe: o teste media um caso que a produção não tem. Dado de teste benigno demais é
    gate que não protege nada.
    """
    grande = _png_data_url(*canvas)
    for tipo in ("produto", "fornecedor"):
        _mock(monkeypatch, [_prod(1, descricao="COPAPA COMPANHIA PADUANA DE PAPEIS LTDA ME",
                                  fornecedor="COPAPA COMPANHIA PADUANA DE PAPEIS")],
              forn_nome="COPAPA COMPANHIA PADUANA DE PAPEIS",
              comp_nome="MARIA APARECIDA GONCALVES DA SILVA",
              extra={"ciclo_dias": 27.0, "n_pedidos": 14, "verba": 208413.55,
                     "verba_campanha": 41220.10},
              lead=[{"codfornec": 1, "lead_real": 16.0, "n": 31, "confiavel": True}])
        with app_ctx.test_request_context(f"/estoque/api/export/ficha/{tipo}/1.pdf",
                                          method="POST", json={"grafico": grande}):
            data = routes.api_export_ficha_pdf(tipo, 1).data
        paginas = data.count(b"/Type /Page") - data.count(b"/Type /Pages")
        assert paginas == 1, f"ficha de {tipo} com canvas {canvas} saiu com {paginas} páginas"
        assert b"/Image" in data or b"/XObject" in data, (
            f"ficha de {tipo} coube em 1 página mas PERDEU o gráfico — encolher é aceitável, "
            f"sumir não")


def test_texto_longo_nao_transborda_da_celula(app_ctx, monkeypatch):
    """String crua no Table do reportlab não quebra: transborda por cima da coluna vizinha —
    aconteceu com o nome do fornecedor e, depois, com o rótulo "A comprar (c/ impostos)" colado
    no número. A quebra é decidida MEDINDO a largura do texto na fonte contra a da coluna;
    quebrar por contagem de caracteres chutada foi o erro seguinte (quase todo rótulo virava
    duas linhas e a tabela empurrava o gráfico para fora da página)."""
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    pares = 4
    usable = landscape(A4)[0] - 2.4 * cm
    pesos = [1.0, 1.45] * pares
    col_rot, col_val = [usable * x / sum(pesos) for x in (1.0, 1.45)]
    for _sec, _chave, rot, _kind in routes._FICHA_COLS["produto"] + routes._FICHA_COLS["fornecedor"]:
        cabe = stringWidth(rot, "Helvetica", 8) <= (col_rot - 10)
        # rótulo que não cabe TEM de ser quebrável (ter espaço) — senão nem o Paragraph salva
        assert cabe or " " in rot, f"rótulo {rot!r} não cabe na coluna e não tem onde quebrar"
    assert col_val > col_rot, "a coluna de valor tem de ser a mais larga do par"
